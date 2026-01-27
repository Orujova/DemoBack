# api/vacation_views.py - COMPLETE REFACTORED VERSION

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.db.models import Q, Count, Sum
from datetime import date, datetime, timedelta
import openpyxl
from .vacation_serializers import *
from .models import Employee
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from .vacation_permissions import (
    is_admin_user,
    get_vacation_access,
    filter_vacation_queryset,
    can_user_modify_vacation_request,
    can_user_modify_schedule,
    can_user_register_schedule,
    can_user_approve_request,
    check_vacation_access,
    is_uk_additional_approver
)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from .vacation_models import (
    VacationSetting,
    VacationType,
    VacationRequest,
    VacationSchedule,
    EmployeeVacationBalance,
    VacationAttachment
)
from .vacation_serializers import EmployeeVacationBalanceSerializer

import logging
from django.shortcuts import get_object_or_404
logger = logging.getLogger(__name__)
good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
from rest_framework import status as rest_status
from .vacation_notifications import notification_manager
from .models import UserGraphToken

def get_graph_access_token(user):
    """Get Microsoft Graph access token for the authenticated user"""
    try:
        token = UserGraphToken.get_valid_token(user)
        if token:
            logger.info(f"✅ Valid Graph token found for user {user.username}")
            return token
        else:
            return None
    except Exception as e:
        logger.error(f"❌ Error getting Graph token: {e}")
        return None

def get_notification_context(request):
    """Get notification context with Graph token status"""
    graph_token = get_graph_access_token(request.user)
    
    return {
        'can_send_emails': bool(graph_token),
        'graph_token': graph_token,
        'reason': 'Graph token available' if graph_token else 'No Microsoft Graph token. Login again to enable email notifications.',
        'user': request.user
    }



# ==================== DASHBOARD ====================
@swagger_auto_schema(
    method='get',
    operation_description="Dashboard məlumatları - 6 stat card və əsas statistika",
    operation_summary="Dashboard",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Dashboard data')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def vacation_dashboard(request):
    """✅ Dashboard - Employee öz balansını görür"""
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        emp = access['employee']
        year = date.today().year
        
        # Get employee's balance
        try:
            balance = EmployeeVacationBalance.objects.get(
                employee=emp, 
                year=year,
                is_deleted=False
            )
        except EmployeeVacationBalance.DoesNotExist:
            return Response({
                'balance': {
                    'total_balance': 0,
                    'yearly_balance': 0,
                    'used_days': 0,
                    'remaining_balance': 0,
                    'scheduled_days': 0,
                    'should_be_planned': 0
                }
            })
        
        balance.refresh_from_db()
        
        return Response({
           'balance': {
                'total_balance': float(balance.total_balance),           # 33.0
                'yearly_balance': float(balance.yearly_balance),         # 28.0
                'used_days': float(balance.used_days),                   # 7.0
                'scheduled_days': float(balance.scheduled_days),         # 10.0
                'remaining_balance': float(balance.remaining_balance),   # ✅ 26.0 (33-7)
                'available_for_planning': float(balance.available_for_planning), # ✅ NEW: 16.0 (26-10)
                'should_be_planned': float(balance.should_be_planned)    # 11.0
            }
        })
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# ==================== PRODUCTION CALENDAR SETTINGS ====================
# api/vacation_views.py - FIXED get_calendar_events

@swagger_auto_schema(
    method='get',
    operation_description="Calendar view - holidays və vacation events with automatic calendar selection",
    operation_summary="Get Calendar Events",
    tags=['Vacation'],
    manual_parameters=[
        openapi.Parameter('month', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Month (1-12)'),
        openapi.Parameter('year', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Year (e.g., 2025)'),
        openapi.Parameter('country', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Override calendar: "az" or "uk" (auto-detected from user business function)'),
        openapi.Parameter('employee_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Filter by employee ID'),
        openapi.Parameter('department_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Filter by department ID'),
        openapi.Parameter('business_function_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Filter by business function ID'),
    ],
    responses={200: openapi.Response(description='Calendar events')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_calendar_events(request):
    """✅ ENHANCED: Calendar view - automatic calendar selection based on user business function"""
    try:
        access = get_vacation_access(request.user)
        
        # Get filters
        month = request.GET.get('month')
        year = request.GET.get('year')
        employee_id = request.GET.get('employee_id')
        department_id = request.GET.get('department_id')
        business_function_id = request.GET.get('business_function_id')
        country_override = request.GET.get('country')  # ✅ Optional override
        
        # ✅ AUTO-DETECT calendar based on user's business function
        default_country = 'az'  # Default
        if access['employee'] and access['employee'].business_function:
            bf_code = getattr(access['employee'].business_function, 'code', '')
            if bf_code and bf_code.upper() == 'UK':
                default_country = 'uk'
        
        # Use override if provided, otherwise use auto-detected
        country = (country_override or default_country).lower()
        
        # Default to current month/year
        if not month or not year:
            today = date.today()
            month = month or today.month
            year = year or today.year
        
        month = int(month)
        year = int(year)
        
        # Calculate date range
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # Get holidays
        settings = VacationSetting.get_active()
        holidays = []
        
        if settings:
            # ✅ Use appropriate calendar
            if country == 'uk':
                holiday_calendar = settings.non_working_days_uk
            else:
                holiday_calendar = settings.non_working_days_az
            
            for holiday in holiday_calendar:
                if isinstance(holiday, dict):
                    try:
                        holiday_date = datetime.strptime(holiday['date'], '%Y-%m-%d').date()
                        if start_date <= holiday_date <= end_date:
                            holidays.append({
                                'date': holiday['date'],
                                'name': holiday.get('name', 'Holiday'),
                                'type': 'holiday',
                                'country': country.upper()
                            })
                    except (ValueError, KeyError) as e:
                        logger.warning(f"Invalid holiday date format: {holiday}")
                        continue
                elif isinstance(holiday, str):
                    try:
                        holiday_date = datetime.strptime(holiday, '%Y-%m-%d').date()
                        if start_date <= holiday_date <= end_date:
                            holidays.append({
                                'date': holiday,
                                'name': 'Holiday',
                                'type': 'holiday',
                                'country': country.upper()
                            })
                    except ValueError as e:
                        logger.warning(f"Invalid holiday date string: {holiday}")
                        continue
        
        # ✅ Get vacation requests - filtered by access
        requests_qs = VacationRequest.objects.filter(
            is_deleted=False,
            status__in=['PENDING_LINE_MANAGER', 'PENDING_UK_ADDITIONAL', 'PENDING_HR', 'APPROVED']
        ).filter(
            Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
        ).select_related('employee', 'employee__department', 'employee__business_function', 'vacation_type')
        
        # ✅ Filter by access level
        requests_qs = filter_vacation_queryset(request.user, requests_qs, 'request')
        
        # ✅ Get vacation schedules - filtered by access
        schedules_qs = VacationSchedule.objects.filter(
            is_deleted=False,
            status__in=['SCHEDULED', 'REGISTERED']
        ).filter(
            Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
        ).select_related('employee', 'employee__department', 'employee__business_function', 'vacation_type')
        
        # ✅ Filter by access level
        schedules_qs = filter_vacation_queryset(request.user, schedules_qs, 'schedule')
        
        # Apply additional filters
        if employee_id:
            requests_qs = requests_qs.filter(employee_id=employee_id)
            schedules_qs = schedules_qs.filter(employee_id=employee_id)
        
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
            schedules_qs = schedules_qs.filter(employee__department_id=department_id)
        
        if business_function_id:
            requests_qs = requests_qs.filter(employee__business_function_id=business_function_id)
            schedules_qs = schedules_qs.filter(employee__business_function_id=business_function_id)
        
        # Build vacation events
        vacations = []
        employee_ids_on_vacation = set()
        
        for req in requests_qs:
            # Get business function code
            bf_code = None
            if req.employee.business_function:
                bf_code = getattr(req.employee.business_function, 'code', '')
            
            # Half day display
            period_display = f"{req.start_date.strftime('%Y-%m-%d')} to {req.end_date.strftime('%Y-%m-%d')}"
            if req.is_half_day:
                period_display = f"{req.start_date.strftime('%Y-%m-%d')} (Half Day: {req.half_day_start_time.strftime('%H:%M')} - {req.half_day_end_time.strftime('%H:%M')})"
            
            vacations.append({
                'id': req.id,
                'type': 'request',
                'request_id': req.request_id,
                'employee_id': req.employee.id,
                'employee_name': req.employee.full_name,
                'employee_code': getattr(req.employee, 'employee_id', ''),
                'department': req.employee.department.name if req.employee.department else '',
                'business_function': req.employee.business_function.name if req.employee.business_function else '',
                'business_function_code': bf_code,
                'vacation_type': req.vacation_type.name,
                'vacation_type_id': req.vacation_type.id,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'period_display': period_display,
                'status': req.get_status_display(),
                'status_code': req.status,
                'days': float(req.number_of_days),
                'comment': req.comment,
                'is_half_day': req.is_half_day,
                'half_day_start_time': req.half_day_start_time.strftime('%H:%M') if req.half_day_start_time else None,
                'half_day_end_time': req.half_day_end_time.strftime('%H:%M') if req.half_day_end_time else None,
            })
            employee_ids_on_vacation.add(req.employee.id)
        
        for sch in schedules_qs:
            # Get business function code
            bf_code = None
            if sch.employee.business_function:
                bf_code = getattr(sch.employee.business_function, 'code', '')
            
            vacations.append({
                'id': sch.id,
                'type': 'schedule',
                'request_id': f'SCH{sch.id}',
                'employee_id': sch.employee.id,
                'employee_name': sch.employee.full_name,
                'employee_code': getattr(sch.employee, 'employee_id', ''),
                'department': sch.employee.department.name if sch.employee.department else '',
                'business_function': sch.employee.business_function.name if sch.employee.business_function else '',
                'business_function_code': bf_code,
                'vacation_type': sch.vacation_type.name,
                'vacation_type_id': sch.vacation_type.id,
                'start_date': sch.start_date.strftime('%Y-%m-%d'),
                'end_date': sch.end_date.strftime('%Y-%m-%d'),
                'period_display': f"{sch.start_date.strftime('%Y-%m-%d')} to {sch.end_date.strftime('%Y-%m-%d')}",
                'status': sch.get_status_display(),
                'status_code': sch.status,
                'days': float(sch.number_of_days),
                'comment': sch.comment,
                'is_half_day': False,
            })
            employee_ids_on_vacation.add(sch.employee.id)
        
        # Summary
        summary = {
            'total_holidays': len(holidays),
            'total_vacations': len(vacations),
            'employees_on_vacation': len(employee_ids_on_vacation),
            'month': month,
            'year': year,
            'country': country.upper(),
            'calendar_auto_detected': country_override is None,
            'user_business_function': access['employee'].business_function.name if access['employee'] and access['employee'].business_function else None,
            'access_level': access['access_level']
        }
        
        return Response({
            'holidays': holidays,
            'vacations': vacations,
            'summary': summary
        })
        
    except Exception as e:
        logger.error(f"Calendar events error: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

# ==================== SETTINGS - ADMIN ONLY ====================
@swagger_auto_schema(
    method='get',
    operation_description="✅ Get production calendars (AZ & UK)",
    operation_summary="Get Production Calendars",
    tags=['Vacation - Settings'],
    responses={200: openapi.Response(description='Production calendars')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_production_calendar(request):
    """✅ Production Calendar - AZ və UK"""
    try:
        settings = VacationSetting.get_active()
        
        return Response({
            'azerbaijan': settings.non_working_days_az if settings else [],
            'uk': settings.non_working_days_uk if settings else []
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='put',
    operation_description="✅ Update production calendars (ADMIN ONLY)",
    operation_summary="Update Production Calendars",
    tags=['Vacation - Settings'],
    request_body=ProductionCalendarSerializer,
    responses={200: openapi.Response(description='Calendars updated')}
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # Admin only
def update_production_calendar(request):
    """✅ Production Calendars - Admin only"""
    try:
        serializer = ProductionCalendarSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        if 'non_working_days_az' in serializer.validated_data:
            settings.non_working_days_az = serializer.validated_data['non_working_days_az']
        
        if 'non_working_days_uk' in serializer.validated_data:
            settings.non_working_days_uk = serializer.validated_data['non_working_days_uk']
        
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Production calendars updated successfully',
            'azerbaijan': settings.non_working_days_az,
            'uk': settings.non_working_days_uk,
            'updated_at': settings.updated_at,
            'updated_by': request.user.get_full_name() or request.user.username
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="Production Calendar - qeyri-iş günlərini təyin et (ADMIN ONLY)",
    operation_summary="Set Non-Working Days",
    tags=['Vacation - Settings'],
    request_body=ProductionCalendarSerializer,
    responses={200: openapi.Response(description='Production calendar yeniləndi')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def set_production_calendar(request):
    """✅ Production Calendar - Admin only"""
    try:
        serializer = ProductionCalendarSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        non_working_days = serializer.validated_data['non_working_days']
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        settings.non_working_days = non_working_days
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Production calendar uğurla yeniləndi',
            'non_working_days': settings.non_working_days
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== UK ADDITIONAL APPROVER SETTINGS ====================
@swagger_auto_schema(
    method='get',
    operation_description="✅ Get UK Additional Approver",
    operation_summary="Get UK Additional Approver",
    tags=['Vacation - Settings'],
    responses={200: openapi.Response(description='UK Additional Approver')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_uk_additional_approver(request):
    """✅ Get UK Additional Approver"""
    try:
        settings = VacationSetting.get_active()
        
        if not settings or not settings.uk_additional_approver:
            return Response({
                'uk_additional_approver': None
            })
        
        approver = settings.uk_additional_approver
        
        return Response({
            'uk_additional_approver': {
                'id': approver.id,
                'name': approver.full_name,
                'employee_id': getattr(approver, 'employee_id', ''),
                'position_group': approver.position_group.name if approver.position_group else '',
                'business_function': approver.business_function.name if approver.business_function else '',
                'email': approver.user.email if approver.user else ''
            }
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='put',
    operation_description="✅ Set UK Additional Approver (ADMIN ONLY)",
    operation_summary="Set UK Additional Approver",
    tags=['Vacation - Settings'],
    request_body=UKAdditionalApproverSerializer,
    responses={200: openapi.Response(description='UK Additional Approver set')}
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # Admin only
def set_uk_additional_approver(request):
    """✅ Set UK Additional Approver - Admin only"""
    try:
        serializer = UKAdditionalApproverSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        approver_id = serializer.validated_data['uk_additional_approver_id']
        approver = Employee.objects.get(id=approver_id, is_deleted=False)
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        previous_approver = settings.uk_additional_approver
        settings.uk_additional_approver = approver
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'UK Additional Approver set successfully',
            'previous_approver': {
                'id': previous_approver.id,
                'name': previous_approver.full_name
            } if previous_approver else None,
            'current_approver': {
                'id': approver.id,
                'name': approver.full_name,
                'position_group': approver.position_group.name if approver.position_group else '',
                'business_function': approver.business_function.name if approver.business_function else ''
            },
            'updated_at': settings.updated_at
        })
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'Approver not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)




# ==================== GENERAL VACATION SETTINGS - ADMIN ONLY ====================
@swagger_auto_schema(
    method='put',
    operation_description="Vacation ümumi parametrləri - balans, edit limiti, bildirişlər yenilə (ADMIN ONLY)",
    operation_summary="Update General Vacation Settings",
    tags=['Vacation - Settings'],
    request_body=GeneralVacationSettingsSerializer,
    responses={200: openapi.Response(description='Ümumi parametrlər yeniləndi')}
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def update_general_vacation_settings(request):
    """✅ Vacation ümumi parametrləri - Admin only"""
    try:
        serializer = GeneralVacationSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        if 'allow_negative_balance' in data:
            settings.allow_negative_balance = data['allow_negative_balance']
        
        if 'max_schedule_edits' in data:
            settings.max_schedule_edits = data['max_schedule_edits']
        
        if 'notification_days_before' in data:
            settings.notification_days_before = data['notification_days_before']
        
        if 'notification_frequency' in data:
            settings.notification_frequency = data['notification_frequency']
        
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Vacation parametrləri uğurla yeniləndi',
            'settings': {
                'allow_negative_balance': settings.allow_negative_balance,
                'max_schedule_edits': settings.max_schedule_edits,
                'notification_days_before': settings.notification_days_before,
                'notification_frequency': settings.notification_frequency
            },
            'updated_at': settings.updated_at,
            'updated_by': request.user.get_full_name() or request.user.username
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_description="Vacation ümumi parametrləri - balans, edit limiti, bildirişlər (ADMIN ONLY)",
    operation_summary="Set General Vacation Settings",
    tags=['Vacation - Settings'],
    request_body=GeneralVacationSettingsSerializer,
    responses={200: openapi.Response(description='Ümumi parametrlər yeniləndi')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def set_general_vacation_settings(request):
    """✅ Vacation ümumi parametrləri - Admin only"""
    try:
        serializer = GeneralVacationSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        if 'allow_negative_balance' in data:
            settings.allow_negative_balance = data['allow_negative_balance']
        
        if 'max_schedule_edits' in data:
            settings.max_schedule_edits = data['max_schedule_edits']
        
        if 'notification_days_before' in data:
            settings.notification_days_before = data['notification_days_before']
        
        if 'notification_frequency' in data:
            settings.notification_frequency = data['notification_frequency']
        
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Vacation parametrləri uğurla yeniləndi',
            'settings': {
                'allow_negative_balance': settings.allow_negative_balance,
                'max_schedule_edits': settings.max_schedule_edits,
                'notification_days_before': settings.notification_days_before,
                'notification_frequency': settings.notification_frequency
            }
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_description="Vacation ümumi parametrlərini əldə et",
    operation_summary="Get General Vacation Settings",
    tags=['Vacation - Settings'],
    responses={200: openapi.Response(description='Ümumi parametrlər')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_general_vacation_settings(request):
    """✅ Vacation ümumi parametrlərini əldə et - Hamı görə bilər"""
    try:
        settings = VacationSetting.get_active()
        
        if not settings:
            return Response({
                'allow_negative_balance': False,
                'max_schedule_edits': 3,
                'notification_days_before': 7,
                'notification_frequency': 2
            })
        
        return Response({
            'allow_negative_balance': settings.allow_negative_balance,
            'max_schedule_edits': settings.max_schedule_edits,
            'notification_days_before': settings.notification_days_before,
            'notification_frequency': settings.notification_frequency
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==================== HR REPRESENTATIVE SETTINGS - ADMIN ONLY ====================
@swagger_auto_schema(
    method='put',
    operation_description="Default HR nümayəndəsini yenilə (ADMIN ONLY)",
    operation_summary="Update Default HR Representative",
    tags=['Vacation - Settings'],
    request_body=HRRepresentativeSerializer,
    responses={200: openapi.Response(description='Default HR yeniləndi')}
)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def update_default_hr_representative(request):
    """✅ Default HR nümayəndəsini yenilə - Admin only"""
    try:
        serializer = HRRepresentativeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        hr_id = serializer.validated_data['default_hr_representative_id']
        hr_employee = Employee.objects.get(id=hr_id, is_deleted=False)
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        previous_hr = settings.default_hr_representative
        previous_hr_info = None
        if previous_hr:
            previous_hr_info = {
                'id': previous_hr.id,
                'name': previous_hr.full_name,
                'department': previous_hr.department.name if previous_hr.department else ''
            }
        
        settings.default_hr_representative = hr_employee
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Default HR nümayəndəsi uğurla yeniləndi',
            'previous_hr': previous_hr_info,
            'current_hr': {
                'id': hr_employee.id,
                'name': hr_employee.full_name,
                'department': hr_employee.department.name if hr_employee.department else ''
            },
            'updated_at': settings.updated_at,
            'updated_by': request.user.get_full_name() or request.user.username
        })
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'HR nümayəndəsi tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_description="Default HR nümayəndəsini təyin et (ADMIN ONLY)",
    operation_summary="Set Default HR Representative",
    tags=['Vacation - Settings'],
    request_body=HRRepresentativeSerializer,
    responses={200: openapi.Response(description='Default HR təyin edildi')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def set_default_hr_representative(request):
    """✅ Default HR nümayəndəsini təyin et - Admin only"""
    try:
        serializer = HRRepresentativeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        hr_id = serializer.validated_data['default_hr_representative_id']
        hr_employee = Employee.objects.get(id=hr_id, is_deleted=False)
        
        settings = VacationSetting.get_active()
        if not settings:
            settings = VacationSetting.objects.create(
                is_active=True,
                created_by=request.user
            )
        
        settings.default_hr_representative = hr_employee
        settings.updated_by = request.user
        settings.save()
        
        return Response({
            'message': 'Default HR nümayəndəsi uğurla təyin edildi',
            'hr_representative': {
                'id': hr_employee.id,
                'name': hr_employee.full_name,
                'department': hr_employee.department.name if hr_employee.department else ''
            }
        })
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'HR nümayəndəsi tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_description="Mövcud HR nümayəndələrini əldə et",
    operation_summary="Get HR Representatives",
    tags=['Vacation - Settings'],
    responses={200: openapi.Response(description='HR nümayəndələri siyahısı')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_hr_representatives(request):
    """✅ HR nümayəndələrini əldə et - Hamı görə bilər"""
    try:
        settings = VacationSetting.get_active()
        current_default = settings.default_hr_representative if settings else None
        
        # HR departamentindəki işçilər
        hr_employees = Employee.objects.filter(
            unit__name__icontains='HR',
            is_deleted=False
        )
        
        hr_list = []
        for emp in hr_employees:
            hr_list.append({
                'id': emp.id,
                'name': emp.full_name,
                'email': emp.user.email if emp.user else '',
                'phone': emp.phone,
                'department': emp.unit.name if emp.unit else '',
                'is_default': current_default and current_default.id == emp.id
            })
        
        return Response({
            'current_default': {
                'id': current_default.id,
                'name': current_default.full_name,
                'department': current_default.department.name if current_default.department else ''
            } if current_default else None,
            'hr_representatives': hr_list
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)



@swagger_auto_schema(
    method='post',
    operation_description="Excel faylı ilə vacation balanslarını toplu yüklə (ADMIN ONLY)",
    operation_summary="Bulk Upload Balances",
    tags=['Vacation - Settings'],
    responses={200: openapi.Response(description='Upload successful')}
)
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def bulk_upload_balances(request):
    """✅ Excel ilə balance upload et - Admin only - FIXED"""
    if 'file' not in request.FILES:
        return Response({'error': 'File yoxdur'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    year = int(request.data.get('year', date.today().year))
    
    try:
        # ✅ READ EXCEL: Start from row 7 (header on row 5, description on row 6)
        df = pd.read_excel(file, header=4, skiprows=[5])
        
        # ✅ Clean dataframe: remove completely empty rows
        df = df.dropna(how='all')
        
        # ✅ Strip whitespace and normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # ✅ Debug: Show what columns we found
        logger.info(f"📋 Found columns: {list(df.columns)}")
        logger.info(f"📊 Total rows to process: {len(df)}")
        
        # Required columns check
        required_cols = ['employee_id', 'start_balance', 'yearly_balance']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            return Response({
                'error': f'Missing columns: {", ".join(missing_cols)}',
                'found_columns': list(df.columns),
                'hint': 'Please use the downloaded template without modifications'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        results = {'successful': 0, 'failed': 0, 'errors': [], 'skipped': 0, 'replaced': 0}
        
        for idx, row in df.iterrows():
            try:
                emp_id_raw = row.get('employee_id')
                
                if pd.isna(emp_id_raw):
                    results['skipped'] += 1
                    continue
                
                emp_id = str(emp_id_raw).strip()
                
                if not emp_id or len(emp_id) > 20:
                    results['skipped'] += 1
                    continue
                
                emp = Employee.objects.get(
                    employee_id=emp_id, 
                    is_deleted=False
                )
                
                # ✅ ENHANCED: Clean parsing with rounding to prevent float precision errors
                start_bal_raw = row.get('start_balance')
                yearly_bal_raw = row.get('yearly_balance')
                
                # Parse start_balance
                if pd.isna(start_bal_raw) or str(start_bal_raw).strip() == '':
                    start_bal = 0
                else:
                    try:
                        # ✅ Round to 1 decimal to prevent float precision errors
                        start_bal = round(float(str(start_bal_raw).strip()), 1)
                    except (ValueError, TypeError):
                        start_bal = 0
                
                # Parse yearly_balance
                if pd.isna(yearly_bal_raw) or str(yearly_bal_raw).strip() == '':
                    yearly_bal = 0
                else:
                    try:
                        # ✅ Round to 1 decimal to prevent float precision errors
                        yearly_bal = round(float(str(yearly_bal_raw).strip()), 1)
                    except (ValueError, TypeError):
                        yearly_bal = 0
                
                # ✅ LOG: Debug what we're actually saving
                logger.info(f"📊 Processing {emp_id}: start={start_bal}, yearly={yearly_bal}, total={start_bal + yearly_bal}")
                
                # ✅ CRITICAL: Check if employee already has balance for this year
                existing_balance = EmployeeVacationBalance.objects.filter(
                    employee=emp,
                    year=year,
                    is_deleted=False
                ).first()
                
                if existing_balance:
                    logger.warning(f"⚠️  {emp_id} has existing balance - DELETING OLD DATA")
                    logger.warning(f"   Old: start={existing_balance.start_balance}, yearly={existing_balance.yearly_balance}, used={existing_balance.used_days}")
                    existing_balance.delete()
                    results['replaced'] += 1
                
                # ✅ Create fresh balance (completely new record)
                new_balance = EmployeeVacationBalance.objects.create(
                    employee=emp,
                    year=year,
                    start_balance=start_bal,
                    yearly_balance=yearly_bal,
                    used_days=0,
                    scheduled_days=0,
                    updated_by=request.user
                )
                
                logger.info(f"✅ {emp_id} new balance created: total={new_balance.total_balance}")
                
                results['successful'] += 1
                
            except Employee.DoesNotExist:
                results['errors'].append(f"Employee ID '{row['employee_id']}' sistemdə tapılmadı")
                results['failed'] += 1
            except ValueError as e:
                results['errors'].append(f"Employee ID '{row['employee_id']}': Rəqəm formatı səhvdir")
                results['failed'] += 1
            except Exception as e:
                results['errors'].append(f"Employee ID '{row['employee_id']}': {str(e)}")
                results['failed'] += 1
        
        if results['successful'] == 0 and results['failed'] == 0:
            return Response({
                'error': 'Faylda heç bir məlumat tapılmadı',
                'hint': 'Please add employee data starting from row 7',
                'rows_found': len(df),
                'columns_found': list(df.columns)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ Return detailed results with replacement info
        message_parts = [f"{results['successful']} uğurlu"]
        if results['replaced'] > 0:
            message_parts.append(f"{results['replaced']} köhnə balans silindi və yeniləndi")
        if results['failed'] > 0:
            message_parts.append(f"{results['failed']} səhv")
        if results['skipped'] > 0:
            message_parts.append(f"{results['skipped']} skipped")
        
        return Response({
            'message': ', '.join(message_parts),
            'results': results,
            'year': year,
            'total_rows_processed': len(df),
            'columns_used': required_cols
        })
    
    except Exception as e:
        return Response({
            'error': f'File processing error: {str(e)}',
            'hint': 'Make sure you are using the latest template file'
        }, status=status.HTTP_400_BAD_REQUEST)


# ==================== EXCEL TEMPLATE DOWNLOAD - FIXED ====================
@swagger_auto_schema(
    method='get',
    operation_description="Excel template endir",
    operation_summary="Download Balance Template",
    tags=['Vacation - Settings'],
    responses={200: openapi.Response(description='Excel file')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_balance_template(request):
    """✅ Excel template - Hamı endir bilər - FIXED"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Template"
    ws.sheet_view.showGridLines = False

    # Define styles
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(size=16, bold=True, color="305496")
    desc_font = Font(size=9, italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:C2')
    ws['A1'] = 'VACATION BALANCE TEMPLATE'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Date info
    ws['A3'] = f'Generated on: {datetime.now().strftime("%B %d, %Y - %H:%M")}'
    ws['A3'].font = Font(size=10, italic=True, color="808080")

    # Column headers
    headers = ['employee_id', 'start_balance', 'yearly_balance']
    descriptions = [
        'Employee ID from system',
        'Remaining balance from previous year',
        'Annual vacation allocation',
    ]

    start_row = 5
    for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        desc_cell = ws.cell(row=start_row + 1, column=col, value=desc)
        desc_cell.font = desc_font
        desc_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        desc_cell.border = border

    # ✅ Auto-adjust column widths and set number format
    from openpyxl.styles import numbers
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 25
        
        # ✅ Set number format for balance columns (B and C)
        if col in [2, 3]:  # start_balance and yearly_balance columns
            for row in range(7, 1000):  # Future rows
                cell = ws[f'{col_letter}{row}']
                cell.number_format = '0.0'  # ✅ Force 1 decimal format

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vacation_balance_template.xlsx'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_balances(request):
    """
    ✅ Get all employee vacation balances with filters
    - Employee: Own balance only
    - Manager: Own + team balances
    - Admin: All balances
    """
    try:
        access = get_vacation_access(request.user)
        
        # Get filters
        year = request.GET.get('year', datetime.now().year)
        department_id = request.GET.get('department_id', '')
        business_function_id = request.GET.get('business_function_id', '')
        min_remaining = request.GET.get('min_remaining', '')
        max_remaining = request.GET.get('max_remaining', '')
        
        # Build base queryset
        queryset = EmployeeVacationBalance.objects.filter(
            is_deleted=False,
            year=year
        ).select_related('employee', 'employee__department', 'employee__business_function')
        
        # ✅ Filter by access level
        if access['accessible_employee_ids'] is not None:
            queryset = queryset.filter(employee_id__in=access['accessible_employee_ids'])
        
        # Apply additional filters
        if department_id:
            queryset = queryset.filter(employee__department_id=department_id)
        
        if business_function_id:
            queryset = queryset.filter(employee__business_function_id=business_function_id)
        
        # Calculate filtered balances
        balances_list = []
        for balance in queryset:
            if min_remaining and balance.remaining_balance < float(min_remaining):
                continue
            if max_remaining and balance.remaining_balance > float(max_remaining):
                continue
            balances_list.append(balance)
        
        # Serialize
        serializer = EmployeeVacationBalanceSerializer(balances_list, many=True)
        balances = serializer.data
        
        # Calculate summary
        total_allocated = sum(float(b.total_balance) for b in balances_list)
        total_used = sum(float(b.used_days) for b in balances_list)
        total_scheduled = sum(float(b.scheduled_days) for b in balances_list)
        total_remaining = sum(float(b.remaining_balance) for b in balances_list)
        employee_count = len(balances_list)
        
        summary = {
            'total_employees': employee_count,
            'total_allocated': round(total_allocated, 1),
            'total_used': round(total_used, 1),
            'total_scheduled': round(total_scheduled, 1),
            'total_remaining': round(total_remaining, 1)
        }
        
        return Response({
            'balances': balances,
            'summary': summary,
            'access_level': access['access_level'],
            'filters_applied': {
                'year': year,
                'department_id': department_id,
                'business_function_id': business_function_id,
                'min_remaining': min_remaining,
                'max_remaining': max_remaining
            }
        })
    
    except Exception as e:
        logger.error(f"Error in get_all_balances: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_all_balances(request):
    """
    ✅ Export all balances to Excel - filtered by access level
    """
    try:
        access = get_vacation_access(request.user)
        
        # Get filters
        year = request.GET.get('year', datetime.now().year)
        department_id = request.GET.get('department_id', '')
        business_function_id = request.GET.get('business_function_id', '')
        min_remaining = request.GET.get('min_remaining', '')
        max_remaining = request.GET.get('max_remaining', '')
        
        # Build queryset
        queryset = EmployeeVacationBalance.objects.filter(
            is_deleted=False,
            year=year
        ).select_related('employee', 'employee__department').order_by(
            'employee__department__name', 'employee__full_name'
        )
        
        # ✅ Filter by access level
        if access['accessible_employee_ids'] is not None:
            queryset = queryset.filter(employee_id__in=access['accessible_employee_ids'])
        
        if department_id:
            queryset = queryset.filter(employee__department_id=department_id)
        
        if business_function_id:
            queryset = queryset.filter(employee__business_function_id=business_function_id)
        
        # Apply remaining balance filters
        balances_list = []
        for balance in queryset:
            if min_remaining and balance.remaining_balance < float(min_remaining):
                continue
            if max_remaining and balance.remaining_balance > float(max_remaining):
                continue
            balances_list.append(balance)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Vacation Balances {year}'
        
        # Header style
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'Employee Name', 'Employee ID', 'Department', 'Year',
            'Start Balance', 'Yearly Balance', 'Total Balance',
            'Used Days', 'Scheduled Days', 'Remaining Balance', 'To Plan'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, balance in enumerate(balances_list, 2):
            ws.cell(row=row_num, column=1, value=balance.employee.full_name)
            ws.cell(row=row_num, column=2, value=getattr(balance.employee, 'employee_id', ''))
            ws.cell(row=row_num, column=3, value=balance.employee.department.name if balance.employee.department else '')
            ws.cell(row=row_num, column=4, value=balance.year)
            ws.cell(row=row_num, column=5, value=float(balance.start_balance))
            ws.cell(row=row_num, column=6, value=float(balance.yearly_balance))
            ws.cell(row=row_num, column=7, value=float(balance.total_balance))
            ws.cell(row=row_num, column=8, value=float(balance.used_days))
            ws.cell(row=row_num, column=9, value=float(balance.scheduled_days))
            ws.cell(row=row_num, column=10, value=float(balance.remaining_balance))
            ws.cell(row=row_num, column=11, value=float(balance.should_be_planned))
            
            # Center align numeric columns
            for col in range(4, 12):
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=vacation_balances_{access["access_level"].replace(" ", "_")}_{year}.xlsx'
        wb.save(response)
        
        return response
    
    except Exception as e:
        logger.error(f"Error in export_all_balances: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def update_employee_balance(request):
    """
    ✅ Update individual employee balance - ADMIN ONLY
    """
    employee_id = request.data.get('employee_id')
    year = request.data.get('year', datetime.now().year)
    
    if not employee_id:
        return Response(
            {'error': 'employee_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        employee = Employee.objects.get(id=employee_id, is_deleted=False)
    except Employee.DoesNotExist:
        return Response(
            {'error': 'Employee not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    balance, created = EmployeeVacationBalance.objects.get_or_create(
        employee=employee,
        year=year,
        defaults={
            'start_balance': request.data.get('start_balance', 0),
            'yearly_balance': request.data.get('yearly_balance', 28),
            'used_days': request.data.get('used_days', 0),
            'scheduled_days': request.data.get('scheduled_days', 0),
            'updated_by': request.user
        }
    )
    
    if not created:
        if 'start_balance' in request.data:
            balance.start_balance = request.data['start_balance']
        if 'yearly_balance' in request.data:
            balance.yearly_balance = request.data['yearly_balance']
        if 'used_days' in request.data:
            balance.used_days = request.data['used_days']
        if 'scheduled_days' in request.data:
            balance.scheduled_days = request.data['scheduled_days']
        
        balance.updated_by = request.user
        balance.save()
    
    serializer = EmployeeVacationBalanceSerializer(balance)
    return Response({
        'message': 'Balance updated successfully',
        'balance': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def reset_balances(request):
    """
    ✅ Reset balances for a specific year - ADMIN ONLY
    """
    year = request.data.get('year', datetime.now().year)
    department_id = request.data.get('department_id', None)
    
    queryset = EmployeeVacationBalance.objects.filter(
        is_deleted=False,
        year=year
    )
    
    if department_id:
        queryset = queryset.filter(employee__department_id=department_id)
    
    updated_count = queryset.update(
        used_days=0,
        scheduled_days=0,
        updated_by=request.user
    )
    
    return Response({
        'message': f'Reset {updated_count} balances for year {year}',
        'count': updated_count
    })


# ==================== REQUEST IMMEDIATE ====================
@swagger_auto_schema(
    method='post',
    operation_description="✅ Create vacation request with half-day support",
    operation_summary="Create Immediate Request",
    tags=['Vacation'],
    responses={201: openapi.Response(description='Request created')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def create_immediate_request(request):
    """
    ✅ ENHANCED: Create vacation request
    - Half day support for UK employees
    - UK additional approval for 5+ days
    """
    import json
    
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profile not found for current user'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Parse JSON fields
        data = request.data.dict()
        
        if 'employee_manual' in data:
            try:
                data['employee_manual'] = json.loads(data['employee_manual'])
            except json.JSONDecodeError:
                return Response({
                    'error': 'Invalid employee_manual format. Must be valid JSON object.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Convert string booleans
        if 'is_half_day' in data:
            if isinstance(data['is_half_day'], str):
                data['is_half_day'] = data['is_half_day'].lower() == 'true'
        
        uploaded_files = request.FILES.getlist('files')
        
        serializer = VacationRequestCreateSerializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        
        requester_emp = access['employee']
        
        # Determine target employee
        if validated_data['requester_type'] == 'for_me':
            employee = requester_emp
        else:
            if validated_data.get('employee_id'):
                try:
                    employee = Employee.objects.get(id=validated_data['employee_id'], is_deleted=False)
                    
                    # Check access
                    if not access['can_view_all']:
                        if access['accessible_employee_ids'] is None:
                            pass  # Admin
                        elif employee.id not in access['accessible_employee_ids']:
                            return Response({
                                'error': 'This employee is not in your team'
                            }, status=status.HTTP_403_FORBIDDEN)
                    
                except Employee.DoesNotExist:
                    return Response({
                        'error': 'Employee not found'
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # Manual employee
                if not (access['is_manager'] or access['is_admin']):
                    return Response({
                        'error': 'Only managers and admins can create manual employees'
                    }, status=status.HTTP_403_FORBIDDEN)
                
                manual_data = validated_data.get('employee_manual', {})
                if not manual_data.get('name'):
                    return Response({
                        'error': 'Employee name is required'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                employee = Employee.objects.create(
                    full_name=manual_data.get('name', ''),
                    phone=manual_data.get('phone', ''),
                    line_manager=requester_emp,
                    created_by=request.user
                )
        
        # ✅ Validate vacation type for UK-only types
        vacation_type = VacationType.objects.get(
            id=validated_data['vacation_type_id'],
            is_active=True,
            is_deleted=False
        )
        
        # Check if employee can use this type
        is_uk = False
        if employee.business_function:
            code = getattr(employee.business_function, 'code', '')
            is_uk = code.upper() == 'UK'
        
        if vacation_type.is_uk_only and not is_uk:
            return Response({
                'error': 'This vacation type is only available for UK employees'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check conflicts
        temp_request = VacationRequest(
            employee=employee,
            start_date=validated_data['start_date'],
            end_date=validated_data['end_date']
        )
        
        has_conflict, conflicts = temp_request.check_date_conflicts()
        if has_conflict:
            return Response({
                'error': 'Bu tarixlərdə artıq vacation mövcuddur',
                'conflicts': conflicts,
                'message': 'Zəhmət olmasa başqa tarix seçin'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Balance check
        settings = VacationSetting.get_active()
        year = date.today().year
        
        balance, created = EmployeeVacationBalance.objects.get_or_create(
            employee=employee,
            year=year,
            defaults={
                'start_balance': 0,
                'yearly_balance': 28,
                'updated_by': request.user
            }
        )
        
        # ✅ Calculate working days with business function code
        working_days = 0
        bf_code = None
        if employee.business_function:
            bf_code = getattr(employee.business_function, 'code', None)
        
        if validated_data.get('is_half_day'):
            working_days = 0.5
        elif settings:
            working_days = settings.calculate_working_days(
                validated_data['start_date'], 
                validated_data['end_date'],
                bf_code
            )
        
        if settings and not settings.allow_negative_balance:
            if working_days > balance.remaining_balance:  # ✅ scheduled nəzərə alınmır
                return Response({
                    'error': f'Insufficient balance. You have {balance.remaining_balance} days remaining.',
                    'available_balance': float(balance.remaining_balance),
                    'requested_days': working_days
                }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            # Create request
            vac_req = VacationRequest.objects.create(
                employee=employee,
                requester=request.user,
                request_type='IMMEDIATE',
                vacation_type_id=validated_data['vacation_type_id'],
                start_date=validated_data['start_date'],
                end_date=validated_data['end_date'],
                comment=validated_data.get('comment', ''),
                hr_representative_id=validated_data.get('hr_representative_id'),
                is_half_day=validated_data.get('is_half_day', False),
                half_day_start_time=validated_data.get('half_day_start_time'),
                half_day_end_time=validated_data.get('half_day_end_time')
            )
            
            # Upload files
            uploaded_attachments = []
            file_errors = []
            
            for file in uploaded_files:
                try:
                    from .business_trip_serializers import TripAttachmentUploadSerializer
                    upload_serializer = TripAttachmentUploadSerializer(data={'file': file})
                    if not upload_serializer.is_valid():
                        file_errors.append({
                            'filename': file.name,
                            'errors': upload_serializer.errors
                        })
                        continue
                    
                    from .vacation_models import VacationAttachment
                    attachment = VacationAttachment.objects.create(
                        vacation_request=vac_req,
                        file=file,
                        original_filename=file.name,
                        file_size=file.size,
                        file_type=file.content_type,
                        uploaded_by=request.user
                    )
                    uploaded_attachments.append(attachment)
                    
                except Exception as e:
                    file_errors.append({
                        'filename': file.name,
                        'error': str(e)
                    })
            
            # Submit request
            vac_req.submit_request(request.user)
            
            # ✅ Send appropriate notification based on status
            graph_token = get_graph_access_token(request.user)
            notification_sent = False
            
            if graph_token:
                if vac_req.status == 'PENDING_LINE_MANAGER':
                    notification_sent = notification_manager.notify_request_created(vac_req, graph_token)
                elif vac_req.status == 'PENDING_UK_ADDITIONAL':
                    notification_sent = notification_manager.notify_uk_additional_approval_needed(
                        vac_req, 
                        graph_token
                    )
                elif vac_req.status == 'PENDING_HR':
                    notification_sent = notification_manager.notify_hr_approval_needed(vac_req, graph_token)
            
            balance.refresh_from_db()
            
            response_data = {
                'message': 'Vacation request created and submitted successfully.',
                'notification_sent': notification_sent,
                'request': VacationRequestDetailSerializer(vac_req).data,
                'files_uploaded': len(uploaded_attachments),
                'files_failed': len(file_errors),
                'balance': {
                    'total_balance': float(balance.total_balance),
                    'yearly_balance': float(balance.yearly_balance),
                    'used_days': float(balance.used_days),
                    'remaining_balance': float(balance.remaining_balance),
                    'scheduled_days': float(balance.scheduled_days),
                    'should_be_planned': float(balance.should_be_planned)
                }
            }
            
            if uploaded_attachments:
                from .vacation_serializers import VacationAttachmentSerializer
                response_data['attachments'] = VacationAttachmentSerializer(
                    uploaded_attachments,
                    many=True,
                    context={'request': request}
                ).data
            
            if file_errors:
                response_data['file_errors'] = file_errors
            
            return Response(response_data, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        logger.error(f"Error creating vacation request: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def edit_schedule(request, pk):
    """✅ ENHANCED: Edit schedule + HR notification"""
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Only owner can edit
        if schedule.employee != access['employee']:
            return Response({
                'error': 'Bu schedule-i edit etmək hüququnuz yoxdur'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Edit limit check
        if not schedule.can_edit():
            return Response({
                'error': 'Bu schedule-i daha edit edə bilməzsiniz'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse dates
        if 'start_date' in request.data:
            start_date_str = request.data['start_date']
            if isinstance(start_date_str, str):
                schedule.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            else:
                schedule.start_date = start_date_str
        
        if 'end_date' in request.data:
            end_date_str = request.data['end_date']
            if isinstance(end_date_str, str):
                schedule.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            else:
                schedule.end_date = end_date_str
        
        if 'vacation_type_id' in request.data:
            schedule.vacation_type_id = request.data['vacation_type_id']
        
        if 'comment' in request.data:
            schedule.comment = request.data['comment']
        
        # Check conflicts
        has_conflict, conflicts = schedule.check_date_conflicts()
        if has_conflict:
            return Response({
                'error': 'Bu tarixlərdə artıq vacation mövcuddur',
                'conflicts': conflicts,
                'message': 'Zəhmət olmasa başqa tarix seçin'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        schedule.edit_count += 1
        schedule.last_edited_at = timezone.now()
        schedule.last_edited_by = request.user
        schedule.save()
        
        # ✅ Send notification to HR
        graph_token = get_graph_access_token(request.user)
        notification_sent = False
        if graph_token:
            notification_sent = notification_manager.notify_schedule_edited(
                schedule,
                request.user,
                graph_token
            )
        
        return Response({
            'message': 'Schedule yeniləndi',
            'notification_sent': notification_sent,
            'schedule': VacationScheduleSerializer(schedule).data
        })
        
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error editing schedule: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# api/vacation_views.py - approve_schedule UPDATE

@swagger_auto_schema(
    method='post',
    operation_description="✅ Manager/Admin approve schedule",
    operation_summary="Approve Schedule",
    tags=['Vacation'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'action': openapi.Schema(type=openapi.TYPE_STRING, enum=['approve', 'reject']),
            'comment': openapi.Schema(type=openapi.TYPE_STRING)
        }
    ),
    responses={200: openapi.Response(description='Schedule approved')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_schedule(request, pk):
    """✅ Manager/Admin approve/reject schedule"""
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # ✅ Check permission - Manager OR Admin can approve
        can_approve = False
        
        # Admin can always approve
        if access['is_admin']:
            can_approve = True
            logger.info(f"✅ Admin {request.user.username} approving schedule {schedule.id}")
        
        # Manager can approve if they are the line manager
        elif schedule.line_manager == access['employee']:
            can_approve = True
            logger.info(f"✅ Manager {request.user.username} approving schedule {schedule.id}")
        
        if not can_approve:
            return Response({
                'error': 'Yalnız line manager və ya admin təsdiq edə bilər'
            }, status=status.HTTP_403_FORBIDDEN)
        
        action = request.data.get('action')
        comment = request.data.get('comment', '')
        
        if action == 'approve':
            # Approve schedule
            schedule.approve_by_manager(request.user, comment)
            
            # Send notification to HR
            graph_token = get_graph_access_token(request.user)
            notification_sent = False
            if graph_token:
                notification_sent = notification_manager.notify_schedule_approved_by_manager(
                    schedule,
                    graph_token
                )
            
            return Response({
                'message': 'Schedule təsdiq edildi',
                'notification_sent': notification_sent,
                'approved_by': 'Admin' if access['is_admin'] else 'Line Manager',
                'schedule': VacationScheduleSerializer(schedule).data
            })
        
        elif action == 'reject':
            # Reject - soft delete
            schedule.is_deleted = True
            schedule.deleted_by = request.user
            schedule.deleted_at = timezone.now()
            schedule.save()
            
            return Response({
                'message': 'Schedule rədd edildi',
                'rejected_by': 'Admin' if access['is_admin'] else 'Line Manager'
            })
        
        else:
            return Response({
                'error': 'Invalid action. Use "approve" or "reject"'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error approving schedule: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST) 
    


@swagger_auto_schema(
    method='post',
    operation_description="✅ ENHANCED: Vacation Schedule - Manager approve with notification",
    operation_summary="Create Schedule with Approval",
    tags=['Vacation'],
    request_body=VacationScheduleCreateSerializer,
    responses={201: openapi.Response(description='Schedule created')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_schedule(request):
    """
    ✅ ENHANCED: Schedule yarat - Manager approve + HR notification
    """
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profile not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        serializer = VacationScheduleCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        requester_emp = access['employee']
        year = date.today().year
        
        # Determine target employee
        if data['requester_type'] == 'for_me':
            employee = requester_emp
        else:
            if data.get('employee_id'):
                employee = Employee.objects.get(id=data['employee_id'])
                
                # Check access
                if not access['can_view_all']:
                    if access['accessible_employee_ids'] is None:
                        pass
                    elif employee.id not in access['accessible_employee_ids']:
                        return Response({
                            'error': 'Bu işçi sizin tabeliyinizdə deyil'
                        }, status=status.HTTP_403_FORBIDDEN)
            else:
                # Manual employee
                if not (access['is_manager'] or access['is_admin']):
                    return Response({
                        'error': 'Only managers and admins can create manual employees'
                    }, status=status.HTTP_403_FORBIDDEN)
                
                manual_data = data.get('employee_manual', {})
                if not manual_data.get('name'):
                    return Response({
                        'error': 'Employee adı mütləqdir'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                employee = Employee.objects.create(
                    full_name=manual_data.get('name', ''),
                    phone=manual_data.get('phone', ''),
                    line_manager=requester_emp,
                    created_by=request.user
                )
        
        # Check conflicts
        temp_schedule = VacationSchedule(
            employee=employee,
            start_date=data['start_date'],
            end_date=data['end_date']
        )
        
        has_conflict, conflicts = temp_schedule.check_date_conflicts()
        if has_conflict:
            return Response({
                'error': 'Bu tarixlərdə artıq vacation mövcuddur',
                'conflicts': conflicts,
                'message': 'Zəhmət olmasa başqa tarix seçin'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        balance, created = EmployeeVacationBalance.objects.get_or_create(
            employee=employee,
            year=year,
            defaults={
                'start_balance': 0,
                'yearly_balance': 28,
                'updated_by': request.user
            }
        )
        
        settings = VacationSetting.get_active()
        working_days = 0
        if settings:
            bf_code = None
            if employee.business_function:
                bf_code = getattr(employee.business_function, 'code', None)
            working_days = settings.calculate_working_days(
                data['start_date'], 
                data['end_date'],
                bf_code
            )
        
        # ✅ Planning limit check
        if settings and not settings.allow_negative_balance:
            available = balance.available_for_planning
            
            if working_days > available:
                return Response({
                    'error': f'Planlaşdırma limiti aşıldı',
                    'details': {
                        'remaining_balance': float(balance.remaining_balance),
                        'already_scheduled': float(balance.scheduled_days),
                        'available_for_planning': float(available),
                        'requested_days': working_days,
                        'message': f'Maksimum {available} gün planlaşdıra bilərsiniz'
                    }
                }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            # ✅ Determine if needs approval
            is_manager_creating = (
                access['is_manager'] and 
                employee.id in access.get('accessible_employee_ids', [])
            )
            
            # Create schedule with appropriate status
            if is_manager_creating or access['is_admin']:
                # Manager/Admin creates for employee → auto-approve → SCHEDULED
                schedule = VacationSchedule.objects.create(
                    employee=employee,
                    vacation_type_id=data['vacation_type_id'],
                    start_date=data['start_date'],
                    end_date=data['end_date'],
                    comment=data.get('comment', ''),
                    created_by=request.user,
                    status='SCHEDULED',  # ✅ Direct to SCHEDULED
                    manager_approved_by=request.user,
                    manager_approved_at=timezone.now()
                )
                
                # ✅ Send notification to HR
                graph_token = get_graph_access_token(request.user)
                notification_sent = False
                if graph_token:
                    notification_sent = notification_manager.notify_schedule_approved_by_manager(
                        schedule, 
                        graph_token
                    )
                
                message = 'Schedule yaradıldı və təsdiq edildi'
            
            else:
                # Employee creates for self → needs manager approval → PENDING_MANAGER
                schedule = VacationSchedule.objects.create(
                    employee=employee,
                    vacation_type_id=data['vacation_type_id'],
                    start_date=data['start_date'],
                    end_date=data['end_date'],
                    comment=data.get('comment', ''),
                    created_by=request.user,
                    status='PENDING_MANAGER',  # ✅ Needs approval
                    line_manager=employee.line_manager
                )
                
                # ✅ Send notification to MANAGER
                graph_token = get_graph_access_token(request.user)
                notification_sent = False
                if graph_token:
                    logger.info(f"📧 Sending notification to manager: {employee.line_manager.full_name if employee.line_manager else 'N/A'}")
                    notification_sent = notification_manager.notify_schedule_created(
                        schedule,
                        graph_token
                    )
                    logger.info(f"📧 Notification sent: {notification_sent}")
                else:
                    logger.warning(f"⚠️ No Graph token available for user {request.user.username}")
                
                message = 'Schedule yaradıldı və təsdiq gözləyir'
            
            balance.refresh_from_db()
            
            return Response({
                'message': message,
                'notification_sent': notification_sent,
                'schedule': VacationScheduleSerializer(schedule).data,
                'balance': {
                    'total_balance': float(balance.total_balance),
                    'yearly_balance': float(balance.yearly_balance),
                    'used_days': float(balance.used_days),
                    'remaining_balance': float(balance.remaining_balance),
                    'scheduled_days': float(balance.scheduled_days),
                    'available_for_planning': float(balance.available_for_planning),
                    'should_be_planned': float(balance.should_be_planned)
                }
            }, status=status.HTTP_201_CREATED)
    
    except Employee.DoesNotExist:
        return Response({
            'error': 'Employee tapılmadı'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error creating schedule: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
# ==================== MY SCHEDULE TABS ====================
@swagger_auto_schema(
    method='get',
    operation_description="Schedule tabları - upcoming, peers, all",
    operation_summary="My Schedule Tabs",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Schedule tabs data')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_schedule_tabs(request):
    """
    ✅ Schedule tabları - filtered by access
    - Include PENDING_MANAGER and SCHEDULED
    """
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        emp = access['employee']
        
        # ✅ Upcoming schedules - Include PENDING_MANAGER + SCHEDULED
        upcoming_qs = VacationSchedule.objects.filter(
            start_date__gte=date.today(),
            status__in=['PENDING_MANAGER', 'SCHEDULED'],  # ✅ Include both
            is_deleted=False
        )
        
        # ✅ Filter by access
        upcoming_qs = filter_vacation_queryset(request.user, upcoming_qs, 'schedule')
        
        # Peers - only for employees (manager/admin see team in 'all')
        peers_schedules = []
        if not access['is_manager'] and not access['is_admin']:
            peers = Employee.objects.filter(
                Q(department=emp.department) | Q(line_manager=emp.line_manager),
                is_deleted=False
            ).exclude(id=emp.id)
            
            peers_schedules_qs = VacationSchedule.objects.filter(
                employee__in=peers,
                start_date__gte=date.today(),
                status__in=['PENDING_MANAGER', 'SCHEDULED'],  # ✅ Include both
                is_deleted=False
            )
            peers_schedules = VacationScheduleSerializer(peers_schedules_qs, many=True).data
        
        # All schedules - filtered by access
        all_schedules_qs = VacationSchedule.objects.filter(is_deleted=False)
        all_schedules_qs = filter_vacation_queryset(request.user, all_schedules_qs, 'schedule')
        
        return Response({
            'upcoming': VacationScheduleSerializer(upcoming_qs, many=True).data,
            'peers': peers_schedules,
            'all': VacationScheduleSerializer(all_schedules_qs, many=True).data,
            'access_level': access['access_level']
        })
    except Exception as e:
        logger.error(f"Error in my_schedule_tabs: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== REGISTER SCHEDULE ====================
@swagger_auto_schema(
    method='post',
    operation_description="Register schedule as taken with notification (ADMIN ONLY)",
    operation_summary="Register Schedule",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Schedule registered')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def register_schedule(request, pk):
    """✅ Register schedule as taken - ADMIN ONLY"""
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        
        # Register schedule
        schedule.register_as_taken(request.user)
        
        # Send notification
        graph_token = get_graph_access_token(request.user)
        notification_sent = False
        if graph_token:
            notification_sent = notification_manager.notify_schedule_registered(schedule, graph_token)
        
        # Refresh balance
        balance = EmployeeVacationBalance.objects.get(
            employee=schedule.employee,
            year=schedule.start_date.year
        )
        
        return Response({
            'message': 'Schedule registered successfully',
            'notification_sent': notification_sent,
            'schedule': VacationScheduleSerializer(schedule).data,
            'updated_balance': {
                'total_balance': float(balance.total_balance),
                'yearly_balance': float(balance.yearly_balance),
                'used_days': float(balance.used_days),
                'scheduled_days': float(balance.scheduled_days),
                'remaining_balance': float(balance.remaining_balance),
                'should_be_planned': float(balance.should_be_planned)
            }
        })
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error registering schedule: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)



# ==================== DELETE SCHEDULE ====================
@swagger_auto_schema(
    method='delete',
    operation_description="Schedule-i sil (ADMIN ONLY)",
    operation_summary="Delete Schedule",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Schedule silindi')}
)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@check_vacation_access('all')  # ✅ Admin only
def delete_schedule(request, pk):
    """
    ✅ Schedule-i sil - ADMIN ONLY
    - Cannot delete REGISTERED schedules
    """
    try:
        schedule = VacationSchedule.objects.get(pk=pk, is_deleted=False)
        
        # Registered schedule-i silmək olmaz
        if schedule.status == 'REGISTERED':
            return Response({
                'error': 'Registered schedule-i silmək olmaz'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Soft delete
        schedule.is_deleted = True
        schedule.deleted_by = request.user
        schedule.deleted_at = timezone.now()
        schedule.save()
        
        # Scheduled days balansını azalt
        schedule._update_scheduled_balance(add=False)
        
        return Response({'message': 'Schedule silindi'})
        
    except VacationSchedule.DoesNotExist:
        return Response({'error': 'Schedule tapılmadı'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error deleting schedule: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==================== APPROVAL PENDING ====================
@swagger_auto_schema(
    method='get',
    operation_description="✅ Approval - Pending requests (UK additional included)",
    operation_summary="Pending Requests",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Pending requests')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def approval_pending_requests(request):
    """
    ✅ ENHANCED: Approval - Pending requests
    - Manager: Line Manager stage
    - UK Additional Approver: UK Additional stage
    - HR: HR stage
    - Admin: All stages
    """
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        emp = access['employee']
        
        # Admin sees ALL pending
        if access['is_admin']:
            lm_requests = VacationRequest.objects.filter(
                status='PENDING_LINE_MANAGER',
                is_deleted=False
            ).order_by('-created_at')
            
            uk_requests = VacationRequest.objects.filter(
                status='PENDING_UK_ADDITIONAL',
                is_deleted=False
            ).order_by('-created_at')
            
            hr_requests = VacationRequest.objects.filter(
                status='PENDING_HR',
                is_deleted=False
            ).order_by('-created_at')
        else:
            # Manager sees only THEIR team's requests
            lm_requests = VacationRequest.objects.filter(
                line_manager=emp,
                status='PENDING_LINE_MANAGER',
                is_deleted=False
            ).order_by('-created_at')
            
            if access['accessible_employee_ids']:
                lm_requests = lm_requests.filter(
                    employee_id__in=access['accessible_employee_ids']
                )
            
            # ✅ UK Additional Approver sees UK requests
            uk_requests = VacationRequest.objects.none()
            if is_uk_additional_approver(request.user):
                uk_requests = VacationRequest.objects.filter(
                    uk_additional_approver=emp,
                    status='PENDING_UK_ADDITIONAL',
                    is_deleted=False
                ).order_by('-created_at')
            
            # HR representative sees HR stage requests
            hr_requests = VacationRequest.objects.filter(
                hr_representative=emp,
                status='PENDING_HR',
                is_deleted=False
            ).order_by('-created_at')
        
        return Response({
            'line_manager_requests': VacationRequestListSerializer(lm_requests, many=True).data,
            'uk_additional_requests': VacationRequestListSerializer(uk_requests, many=True).data,
            'hr_requests': VacationRequestListSerializer(hr_requests, many=True).data,
            'total_pending': lm_requests.count() + uk_requests.count() + hr_requests.count(),
            'access_level': access['access_level'],
            'is_uk_additional_approver': is_uk_additional_approver(request.user)
        })
    except Exception as e:
        logger.error(f"Error in approval_pending_requests: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_description="✅ Approve/Reject vacation request (UK stage included)",
    tags=['Vacation'],
    request_body=VacationApprovalSerializer,
    responses={200: openapi.Response(description='Action completed')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_reject_request(request, pk):
    """
    ✅ ENHANCED: Approve/Reject with detailed logging
    """
    try:
        vac_req = VacationRequest.objects.select_related(
            'hr_representative',
            'uk_additional_approver'
        ).get(pk=pk, is_deleted=False)
        
        logger.info(f"🔵 APPROVAL REQUEST - ID: {vac_req.request_id}")
        logger.info(f"   Current Status: {vac_req.status}")
        logger.info(f"   HR Rep: {vac_req.hr_representative}")
        logger.info(f"   UK Approver: {vac_req.uk_additional_approver}")
        
        # Check permission
        can_approve, reason = can_user_approve_request(request.user, vac_req)
        
        if not can_approve:
            return Response({
                'error': 'Permission denied',
                'detail': reason
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = VacationApprovalSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        # Get notification context
        notification_ctx = get_notification_context(request)
        graph_token = notification_ctx['graph_token']
        notification_sent = False
        
        # Store old status for comparison
        old_status = vac_req.status
        
        # ✅ UK ADDITIONAL APPROVER APPROVAL/REJECTION
        if vac_req.status == 'PENDING_UK_ADDITIONAL':
            if data['action'] == 'approve':
                logger.info(f"⚙️ Calling approve_by_uk_additional...")
                
                vac_req.approve_by_uk_additional(request.user, data.get('comment', ''))
                
                # ✅ Refresh to get latest status
                vac_req.refresh_from_db()
                
                logger.info(f"✅ UK Approval Complete")
                logger.info(f"   Old Status: {old_status}")
                logger.info(f"   New Status: {vac_req.status}")
                logger.info(f"   HR Rep: {vac_req.hr_representative}")
                
                msg = f'Approved by UK Additional Approver - Now {vac_req.get_status_display()}'
                
                # ✅ Send notification based on NEW status
                if graph_token:
                    try:
                        if vac_req.status == 'PENDING_HR':
                            logger.info(f"📧 Sending HR notification...")
                            notification_sent = notification_manager.notify_uk_additional_approved(
                                vac_req, graph_token
                            )
                        elif vac_req.status == 'APPROVED':
                            logger.info(f"📧 Sending final approval notification...")
                            notification_sent = notification_manager.notify_hr_approved(
                                vac_req, graph_token
                            )
                    except Exception as e:
                        logger.error(f"❌ Notification error: {e}")
            else:
                vac_req.reject_by_uk_additional(request.user, data.get('reason', ''))
                msg = 'Rejected by UK Additional Approver'
                
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_request_rejected(
                            vac_req, graph_token
                        )
                    except Exception as e:
                        logger.error(f"Notification error: {e}")
        
        # ✅ LINE MANAGER APPROVAL/REJECTION
        if vac_req.status == 'PENDING_LINE_MANAGER':
            if data['action'] == 'approve':
                vac_req.approve_by_line_manager(request.user, data.get('comment', ''))
                msg = 'Approved by Line Manager'
                
                # ✅ Send appropriate notification based on NEXT status
                if graph_token:
                    try:
                        if vac_req.status == 'PENDING_UK_ADDITIONAL':
                            notification_sent = notification_manager.notify_uk_additional_approval_needed(
                                vac_req, graph_token
                            )
                        elif vac_req.status == 'PENDING_HR':
                            notification_sent = notification_manager.notify_line_manager_approved(
                                vac_req, graph_token
                            )
                        elif vac_req.status == 'APPROVED':
                            notification_sent = notification_manager.notify_hr_approved(
                                vac_req, graph_token
                            )
                    except Exception as e:
                        logger.error(f"Notification error: {e}")
            else:
                vac_req.reject_by_line_manager(request.user, data.get('reason', ''))
                msg = 'Rejected by Line Manager'
                
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_request_rejected(
                            vac_req, graph_token
                        )
                    except Exception as e:
                        logger.error(f"Notification error: {e}")
        
        
        # ✅ HR APPROVAL/REJECTION
        elif vac_req.status == 'PENDING_HR':
            if data['action'] == 'approve':
                vac_req.approve_by_hr(request.user, data.get('comment', ''))
                msg = 'Approved by HR - Request is now APPROVED'
                
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_hr_approved(
                            vac_req, graph_token
                        )
                    except Exception as e:
                        logger.error(f"Notification error: {e}")
            else:
                vac_req.reject_by_hr(request.user, data.get('reason', ''))
                msg = 'Rejected by HR'
                
                if graph_token:
                    try:
                        notification_sent = notification_manager.notify_request_rejected(
                            vac_req, graph_token
                        )
                    except Exception as e:
                        logger.error(f"Notification error: {e}")
        else:
            return Response({
                'error': 'Request is not pending approval'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'message': msg,
            'notification_sent': notification_sent,
            'notification_available': notification_ctx['can_send_emails'],
            'current_status': vac_req.status,  # ✅ Debug info
            'next_step': vac_req.get_status_display(),  # ✅ Debug info
            'request': VacationRequestDetailSerializer(vac_req).data
        })
    
    except VacationRequest.DoesNotExist:
        return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in approve/reject: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="✅ Create multiple schedules at once (Planning feature)",
    operation_summary="Bulk Create Schedules",
    tags=['Vacation'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['schedules'],
        properties={
            'schedules': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    required=['vacation_type_id', 'start_date', 'end_date'],
                    properties={
                        'vacation_type_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'start_date': openapi.Schema(type=openapi.TYPE_STRING, format='date'),
                        'end_date': openapi.Schema(type=openapi.TYPE_STRING, format='date'),
                        'comment': openapi.Schema(type=openapi.TYPE_STRING),
                    }
                )
            ),
            'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER, description='For manager creating for employee')
        }
    ),
    responses={201: openapi.Response(description='Schedules created')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_create_schedules(request):
    """✅ ENHANCED: Multiple schedule yaratmaq - Planning feature + Notifications"""
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profile not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        schedules_data = request.data.get('schedules', [])
        employee_id = request.data.get('employee_id')
        
        if not schedules_data:
            return Response({
                'error': 'No schedules provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Determine target employee
        if employee_id:
            # Manager/Admin creating for employee
            if not (access['is_manager'] or access['is_admin']):
                return Response({
                    'error': 'Only managers and admins can create schedules for others'
                }, status=status.HTTP_403_FORBIDDEN)
            
            employee = Employee.objects.get(id=employee_id, is_deleted=False)
            
            # Check access
            if not access['can_view_all']:
                if access['accessible_employee_ids'] is None:
                    pass  # Admin
                elif employee.id not in access['accessible_employee_ids']:
                    return Response({
                        'error': 'This employee is not in your team'
                    }, status=status.HTTP_403_FORBIDDEN)
        else:
            employee = access['employee']
        
        year = date.today().year
        settings = VacationSetting.get_active()
        
        # Get current balance
        balance, created = EmployeeVacationBalance.objects.get_or_create(
            employee=employee,
            year=year,
            defaults={
                'start_balance': 0,
                'yearly_balance': 28,
                'updated_by': request.user
            }
        )
        
        created_schedules = []
        errors = []
        total_days = 0
        
        # ✅ Determine if needs approval
        is_manager_creating = (
            (access['is_manager'] or access['is_admin']) and 
            employee_id is not None
        )
        
        with transaction.atomic():
            for idx, schedule_data in enumerate(schedules_data):
                try:
                    # Validate required fields
                    if not all(k in schedule_data for k in ['vacation_type_id', 'start_date', 'end_date']):
                        errors.append({
                            'index': idx,
                            'error': 'Missing required fields'
                        })
                        continue
                    
                    # Parse dates
                    start_dt = datetime.strptime(schedule_data['start_date'], '%Y-%m-%d').date()
                    end_dt = datetime.strptime(schedule_data['end_date'], '%Y-%m-%d').date()
                    
                    if start_dt >= end_dt:
                        errors.append({
                            'index': idx,
                            'error': 'End date must be after start date'
                        })
                        continue
                    
                    # Check vacation type
                    vacation_type = VacationType.objects.get(
                        id=schedule_data['vacation_type_id'],
                        is_active=True,
                        is_deleted=False
                    )
                    
                    # Calculate days
                    if settings:
                        bf_code = None
                        if employee.business_function:
                            bf_code = getattr(employee.business_function, 'code', None)
                        days = settings.calculate_working_days(start_dt, end_dt, bf_code)
                    else:
                        days = (end_dt - start_dt).days + 1
                    
                    total_days += days
                    
                    # Create temp schedule to check conflicts
                    temp_schedule = VacationSchedule(
                        employee=employee,
                        start_date=start_dt,
                        end_date=end_dt
                    )
                    
                    has_conflict, conflicts = temp_schedule.check_date_conflicts()
                    if has_conflict:
                        errors.append({
                            'index': idx,
                            'error': 'Date conflict',
                            'dates': f"{start_dt} - {end_dt}",
                            'conflicts': conflicts
                        })
                        continue
                    
                    # Planning limit check
                    available = balance.available_for_planning
                    if settings and not settings.allow_negative_balance:
                        current_scheduled = float(balance.scheduled_days) + sum([
                            float(s.number_of_days) for s in created_schedules
                        ])
                        still_available = float(balance.remaining_balance) - current_scheduled
                        
                        if days > still_available:
                            errors.append({
                                'index': idx,
                                'error': 'Planlaşdırma limiti aşıldı',
                                'dates': f"{start_dt} - {end_dt}",
                                'days': days,
                                'available': still_available
                            })
                            continue
                    
                    # ✅ Create schedule with appropriate status
                    if is_manager_creating:
                        # Manager/Admin creating → auto-approve
                        schedule = VacationSchedule.objects.create(
                            employee=employee,
                            vacation_type=vacation_type,
                            start_date=start_dt,
                            end_date=end_dt,
                            comment=schedule_data.get('comment', ''),
                            created_by=request.user,
                            status='SCHEDULED',  # ✅ Direct approval
                            manager_approved_by=request.user,
                            manager_approved_at=timezone.now()
                        )
                    else:
                        # Employee creating → needs approval
                        schedule = VacationSchedule.objects.create(
                            employee=employee,
                            vacation_type=vacation_type,
                            start_date=start_dt,
                            end_date=end_dt,
                            comment=schedule_data.get('comment', ''),
                            created_by=request.user,
                            status='PENDING_MANAGER',  # ✅ Needs approval
                            line_manager=employee.line_manager
                        )
                    
                    created_schedules.append(schedule)
                    
                except VacationType.DoesNotExist:
                    errors.append({
                        'index': idx,
                        'error': 'Vacation type not found'
                    })
                except ValueError as e:
                    errors.append({
                        'index': idx,
                        'error': f'Invalid date format: {str(e)}'
                    })
                except Exception as e:
                    errors.append({
                        'index': idx,
                        'error': str(e)
                    })
            
            # Balance check
            if settings and not settings.allow_negative_balance:
                total_planned = balance.scheduled_days + total_days
                if total_planned > balance.total_balance:
                    return Response({
                        'error': f'Insufficient balance. Total planned ({total_planned}) exceeds balance ({balance.total_balance}).',
                        'available_balance': float(balance.remaining_balance),
                        'requested_days': total_days,
                        'current_scheduled': float(balance.scheduled_days)
                    }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ Send notifications AFTER successful creation
        graph_token = get_graph_access_token(request.user)
        notification_sent = False
        
        if created_schedules and graph_token:
            if is_manager_creating:
                # Manager/Admin created → notify HR (only once for bulk)
                logger.info(f"📧 Sending bulk approval notification to HR for {len(created_schedules)} schedules")
           
                settings_obj = VacationSetting.get_active()
                
                if settings_obj and settings_obj.default_hr_representative:
                    hr = settings_obj.default_hr_representative
                    if hr.user and hr.user.email:
                        try:
                            # Send bulk notification
                            subject = f"[VACATION SCHEDULE] {len(created_schedules)} Schedules Approved"
                            
                            schedules_list = '\n'.join([
                                f"- {s.start_date.strftime('%Y-%m-%d')} to {s.end_date.strftime('%Y-%m-%d')} ({s.number_of_days} days)"
                                for s in created_schedules
                            ])
                            
                            body_html = f"""
                            <html>
                            <body style="font-family: Arial, sans-serif;">
                                <div style="max-width: 600px; margin: 0 auto;">
                                    <div style="background-color: #28a745; color: white; padding: 20px; text-align: center;">
                                        <h2>✅ Bulk Vacation Schedules Approved</h2>
                                    </div>
                                    <div style="padding: 20px; background-color: #f9f9f9;">
                                        <p>Dear {hr.full_name},</p>
                                        <p><strong>{len(created_schedules)} vacation schedules</strong> have been approved by manager.</p>
                                        
                                        <div style="margin: 20px 0;">
                                            <strong>Employee:</strong> {employee.full_name}<br>
                                            <strong>Department:</strong> {employee.department.name if employee.department else 'N/A'}<br>
                                            <strong>Total Days:</strong> {total_days} days
                                        </div>
                                        
                                        <div style="background-color: white; padding: 15px; border: 1px solid #ddd; margin: 20px 0;">
                                            <strong>Schedules:</strong><br>
                                            <pre style="white-space: pre-wrap;">{schedules_list}</pre>
                                        </div>
                                        
                                        <div style="margin-top: 20px; text-align: center;">
                                            <a href="https://myalmet.com/requests/vacation/" 
                                               style="display: inline-block; padding: 12px 24px; background-color: #28a745; color: white; text-decoration: none; border-radius: 5px;">
                                                View Schedules
                                            </a>
                                        </div>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """
                            
                            from .notification_service import notification_service
                            notification_sent = notification_service.send_email(
                                recipient_email=hr.user.email,
                                subject=subject,
                                body_html=body_html,
                                access_token=graph_token,
                                related_model='VacationSchedule',
                                related_object_id=created_schedules[0].id if created_schedules else None,
                                sent_by=request.user
                            )
                            logger.info(f"📧 Bulk HR notification sent: {notification_sent}")
                        except Exception as e:
                            logger.error(f"Error sending bulk HR notification: {e}")
            else:
                # Employee created → notify MANAGER (only once for bulk)
                if employee.line_manager and employee.line_manager.user and employee.line_manager.user.email:
                    logger.info(f"📧 Sending bulk schedule notification to manager: {employee.line_manager.full_name}")
                    try:
                        schedules_list = '\n'.join([
                            f"- {s.start_date.strftime('%Y-%m-%d')} to {s.end_date.strftime('%Y-%m-%d')} ({s.number_of_days} days)"
                            for s in created_schedules
                        ])
                        
                        subject = f"[VACATION SCHEDULE] {len(created_schedules)} Schedules Pending Your Approval"
                        
                        body_html = f"""
                        <html>
                        <body style="font-family: Arial, sans-serif;">
                            <div style="max-width: 600px; margin: 0 auto;">
                                <div style="background-color: #366092; color: white; padding: 20px; text-align: center;">
                                    <h2>📅 New Vacation Schedules Pending Approval</h2>
                                </div>
                                <div style="padding: 20px; background-color: #f9f9f9;">
                                    <p>Dear {employee.line_manager.full_name},</p>
                                    <p><strong>{len(created_schedules)} vacation schedules</strong> have been created and require your approval.</p>
                                    
                                    <div style="margin: 20px 0;">
                                        <strong>Employee:</strong> {employee.full_name}<br>
                                        <strong>Department:</strong> {employee.department.name if employee.department else 'N/A'}<br>
                                        <strong>Total Days:</strong> {total_days} days
                                    </div>
                                    
                                    <div style="background-color: white; padding: 15px; border: 1px solid #ddd; margin: 20px 0;">
                                        <strong>Schedules:</strong><br>
                                        <pre style="white-space: pre-wrap;">{schedules_list}</pre>
                                    </div>
                                    
                                    <div style="margin-top: 20px; text-align: center;">
                                        <a href="https://myalmet.com/requests/vacation/" 
                                           style="display: inline-block; padding: 12px 24px; background-color: #366092; color: white; text-decoration: none; border-radius: 5px;">
                                            Review Schedules
                                        </a>
                                    </div>
                                </div>
                            </div>
                        </body>
                        </html>
                        """
                        
                        from .notification_service import notification_service
                        notification_sent = notification_service.send_email(
                            recipient_email=employee.line_manager.user.email,
                            subject=subject,
                            body_html=body_html,
                            access_token=graph_token,
                            related_model='VacationSchedule',
                            related_object_id=created_schedules[0].id if created_schedules else None,
                            sent_by=request.user
                        )
                        logger.info(f"📧 Bulk manager notification sent: {notification_sent}")
                    except Exception as e:
                        logger.error(f"Error sending bulk manager notification: {e}")
        
        # Refresh balance
        balance.refresh_from_db()
        
        message_parts = [f'{len(created_schedules)} schedules created successfully']
        if is_manager_creating:
            message_parts.append('and approved')
        else:
            message_parts.append('and pending manager approval')
        
        return Response({
            'message': ' '.join(message_parts),
            'created_count': len(created_schedules),
            'error_count': len(errors),
            'total_days_planned': total_days,
            'notification_sent': notification_sent,
            'schedules': VacationScheduleSerializer(created_schedules, many=True).data,
            'errors': errors if errors else None,
            'balance': {
                'total_balance': float(balance.total_balance),
                'yearly_balance': float(balance.yearly_balance),
                'used_days': float(balance.used_days),
                'remaining_balance': float(balance.remaining_balance),
                'scheduled_days': float(balance.scheduled_days),
                'should_be_planned': float(balance.should_be_planned)
            }
        }, status=status.HTTP_201_CREATED)
        
    except Employee.DoesNotExist:
        return Response({
            'error': 'Employee not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Bulk schedule creation error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
        
        
# ==================== APPROVAL HISTORY ====================
@swagger_auto_schema(
    method='get',
    operation_description="Approval History",
    operation_summary="Approval History",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Approval history')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def approval_history(request):
    """✅ Approval History - User's approval actions"""
    try:
        # Line Manager kimi təsdiq etdiklərim
        lm_approved = VacationRequest.objects.filter(
            line_manager_approved_by=request.user,
            is_deleted=False
        ).order_by('-line_manager_approved_at')[:20]
        
        # HR kimi təsdiq etdiklərim
        hr_approved = VacationRequest.objects.filter(
            hr_approved_by=request.user,
            is_deleted=False
        ).order_by('-hr_approved_at')[:20]
        
        # Reject etdiklərim
        rejected = VacationRequest.objects.filter(
            rejected_by=request.user,
            is_deleted=False
        ).order_by('-rejected_at')[:20]
        
        history = []
        
        for req in lm_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'days': float(req.number_of_days),
                'status': 'Approved (Line Manager)',
                'action': 'Approved',
                'comment': req.line_manager_comment,
                'date': req.line_manager_approved_at
            })
        
        for req in hr_approved:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'days': float(req.number_of_days),
                'status': 'Approved (HR)',
                'action': 'Approved',
                'comment': req.hr_comment,
                'date': req.hr_approved_at
            })
        
        for req in rejected:
            history.append({
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'days': float(req.number_of_days),
                'status': req.get_status_display(),
                'action': 'Rejected',
                'comment': req.rejection_reason,
                'date': req.rejected_at
            })
        
        history.sort(key=lambda x: x['date'] if x['date'] else datetime.min, reverse=True)
        
        return Response({'history': history[:20]})
    except Exception as e:
        logger.error(f"Error in approval_history: {e}")
        return Response({
            'error': f'History yüklənmədi: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)




# ==================== MY ALL REQUESTS & SCHEDULES ====================
@swagger_auto_schema(
    method='get',
    operation_description="İstifadəçinin bütün vacation request və schedule-lərini göstər",
    operation_summary="My All Requests & Schedules",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Bütün records')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_all_requests_schedules(request):
    """
    ✅ My All Requests & Schedules - filtered by access
    - Employee: Own records only
    - Manager: Own + team records
    - Admin: All records
    """
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # All requests - filtered by access
        requests_qs = VacationRequest.objects.filter(
            is_deleted=False
        ).select_related(
            'vacation_type',
            'line_manager',
            'hr_representative',
            'line_manager_approved_by',
            'hr_approved_by',
            'rejected_by'
        ).prefetch_related('attachments')
        
        requests_qs = filter_vacation_queryset(request.user, requests_qs, 'request')
        requests = requests_qs.order_by('-created_at')
        
        # All schedules - filtered by access
        schedules_qs = VacationSchedule.objects.filter(
            is_deleted=False
        ).select_related(
            'vacation_type',
            'created_by',
            'last_edited_by'
        )
        
        schedules_qs = filter_vacation_queryset(request.user, schedules_qs, 'schedule')
        schedules = schedules_qs.order_by('-start_date')
        
        # Combine
        combined = []
        
        for req in requests:
            attachments_count = req.attachments.filter(is_deleted=False).count()
            
            combined.append({
                'id': req.id,
                'type': 'request',
                'request_id': req.request_id,
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'return_date': req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                'days': float(req.number_of_days),
                'status': req.get_status_display(),
                'status_code': req.status,
                'comment': req.comment,
                'attachments_count': attachments_count,
                'has_attachments': attachments_count > 0,
                'line_manager': req.line_manager.full_name if req.line_manager else '',
                'hr_representative': req.hr_representative.full_name if req.hr_representative else '',
                'line_manager_comment': req.line_manager_comment,
                'hr_comment': req.hr_comment,
                'rejection_reason': req.rejection_reason,
                'created_at': req.created_at.isoformat() if req.created_at else None
            })
        
        for sch in schedules:
            combined.append({
                'id': sch.id,
                'type': 'schedule',
                'request_id': f'SCH{sch.id}',
                'vacation_type': sch.vacation_type.name,
                'start_date': sch.start_date.strftime('%Y-%m-%d'),
                'end_date': sch.end_date.strftime('%Y-%m-%d'),
                'return_date': sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                'days': float(sch.number_of_days),
                'status': sch.get_status_display(),
                'status_code': sch.status,
                'comment': sch.comment,
                'attachments_count': 0,
                'has_attachments': False,
                'created_at': sch.created_at.isoformat() if sch.created_at else None,
                'can_edit': sch.can_edit(),
                'edit_count': sch.edit_count,
                'created_by': sch.created_by.get_full_name() if sch.created_by else '',
                'last_edited_by': sch.last_edited_by.get_full_name() if sch.last_edited_by else '',
                'last_edited_at': sch.last_edited_at.isoformat() if sch.last_edited_at else None
            })
        
        combined.sort(key=lambda x: x['created_at'] if x['created_at'] else '', reverse=True)
        
        return Response({
            'records': combined,
            'total_count': len(combined),
            'requests_count': requests.count(),
            'schedules_count': schedules.count(),
            'access_level': access['access_level']
        })
        
    except Exception as e:
        logger.error(f"Error in my_all_requests_schedules: {e}")
        return Response({
            'error': str(e)
        }, status=rest_status.HTTP_400_BAD_REQUEST)


# ==================== REQUEST DETAIL ====================
@swagger_auto_schema(
    method='get',
    operation_description="Get detailed information of a vacation request including attachments and approval history",
    operation_summary="Get Vacation Request Detail",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Vacation request details')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_request_detail(request, pk):
    """
    ✅ Get detailed information of a vacation request
    Access control applied
    """
    try:
        vac_req = VacationRequest.objects.select_related(
            'employee', 
            'employee__department',
            'employee__business_function',
            'employee__unit',
            'employee__job_function',
            'vacation_type',
            'line_manager',
            'hr_representative',
            'requester'
        ).prefetch_related(
            'attachments'
        ).get(pk=pk, is_deleted=False)
        
        access = get_vacation_access(request.user)
        
        # ✅ Check if user can view this request
        can_view = False
        
        if access['is_admin']:
            can_view = True
        elif access['employee'] and vac_req.employee == access['employee']:
            can_view = True
        elif vac_req.requester == request.user:
            can_view = True
        elif access['employee'] and (
            vac_req.line_manager == access['employee'] or 
            vac_req.hr_representative == access['employee']
        ):
            can_view = True
        elif access['accessible_employee_ids'] and vac_req.employee_id in access['accessible_employee_ids']:
            can_view = True
        
        if not can_view:
            return Response({
                'error': 'Permission denied',
                'detail': 'You do not have permission to view this vacation request'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Serialize the data
        serializer = VacationRequestDetailSerializer(
            vac_req, 
            context={'request': request}
        )
        
        response_data = serializer.data
        
        # Add approval workflow status
        response_data['workflow'] = {
            'current_step': vac_req.status,
            'steps': [
                {
                    'name': 'Line Manager Approval',
                    'status': 'completed' if vac_req.line_manager_approved_at else (
                        'rejected' if vac_req.status == 'REJECTED_LINE_MANAGER' else (
                            'pending' if vac_req.status == 'PENDING_LINE_MANAGER' else 'not_started'
                        )
                    ),
                    'approver': vac_req.line_manager.full_name if vac_req.line_manager else None,
                    'approved_at': vac_req.line_manager_approved_at,
                    'comment': vac_req.line_manager_comment
                },
                {
                    'name': 'HR Processing',
                    'status': 'completed' if vac_req.hr_approved_at else (
                        'rejected' if vac_req.status == 'REJECTED_HR' else (
                            'pending' if vac_req.status == 'PENDING_HR' else 'not_started'
                        )
                    ),
                    'approver': vac_req.hr_representative.full_name if vac_req.hr_representative else None,
                    'approved_at': vac_req.hr_approved_at,
                    'comment': vac_req.hr_comment
                }
            ]
        }
        
        # Add requester information
        response_data['requester_info'] = {
            'type': vac_req.get_request_type_display(),
            'name': vac_req.requester.get_full_name() if vac_req.requester else None,
            'email': vac_req.requester.email if vac_req.requester else None
        }
        
        # ✅ Add permission flags
        can_approve, approve_reason = can_user_approve_request(request.user, vac_req)
        can_modify, modify_reason = can_user_modify_vacation_request(request.user, vac_req)
        
        response_data['permissions'] = {
            'can_approve': can_approve,
            'can_modify': can_modify,
            'is_admin': access['is_admin'],
            'access_level': access['access_level']
        }
        
        return Response(response_data)
        
    except VacationRequest.DoesNotExist:
        return Response({
            'error': 'Vacation request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error fetching vacation request detail: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# ==================== ALL VACATION RECORDS ====================
@swagger_auto_schema(
    method='get',
    operation_description="Bütün vacation request və schedule-ları göstər (hamı görə bilər)",
    operation_summary="All Vacation Records",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Bütün vacation records')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def all_vacation_records(request):
    """
    ✅ Bütün vacation records-u JSON formatında list qaytarır
    - Employee: Own records only
    - Manager: Own + team records
    - Admin: All records
    """
    try:
        access = get_vacation_access(request.user)
        
        # Filter parameters
        status_filter = request.GET.get('status')
        vacation_type_id = request.GET.get('vacation_type_id')
        department_id = request.GET.get('department_id')
        business_function_id = request.GET.get('business_function_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_name = request.GET.get('employee_name')
        year = request.GET.get('year')
        
        # All requests
        requests_qs = VacationRequest.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'line_manager', 'hr_representative',
            'line_manager_approved_by', 'hr_approved_by', 'rejected_by'
        ).prefetch_related('attachments')
        
        # ✅ Filter by access level
        requests_qs = filter_vacation_queryset(request.user, requests_qs, 'request')
        
        # All schedules  
        schedules_qs = VacationSchedule.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'created_by', 'last_edited_by'
        )
        
        # ✅ Filter by access level
        schedules_qs = filter_vacation_queryset(request.user, schedules_qs, 'schedule')
        
        # Apply filters
        if status_filter:
            requests_qs = requests_qs.filter(status=status_filter)
            schedules_qs = schedules_qs.filter(status=status_filter)
        
        if vacation_type_id:
            requests_qs = requests_qs.filter(vacation_type_id=vacation_type_id)
            schedules_qs = schedules_qs.filter(vacation_type_id=vacation_type_id)
        
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
            schedules_qs = schedules_qs.filter(employee__department_id=department_id)
        
        if business_function_id:
            requests_qs = requests_qs.filter(employee__business_function_id=business_function_id)
            schedules_qs = schedules_qs.filter(employee__business_function_id=business_function_id)
        
        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
            schedules_qs = schedules_qs.filter(start_date__gte=start_date)
        
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)
            schedules_qs = schedules_qs.filter(end_date__lte=end_date)
        
        if employee_name:
            requests_qs = requests_qs.filter(employee__full_name__icontains=employee_name)
            schedules_qs = schedules_qs.filter(employee__full_name__icontains=employee_name)
        
        if year:
            requests_qs = requests_qs.filter(start_date__year=year)
            schedules_qs = schedules_qs.filter(start_date__year=year)
        
        # Get data
        requests = requests_qs.order_by('-created_at')
        schedules = schedules_qs.order_by('-created_at')
        
        # Combine all records
        all_records = []
        
        # Add requests
        for req in requests:
            attachments_count = req.attachments.filter(is_deleted=False).count()
            
            all_records.append({
                'id': req.id,
                'type': 'request',
                'request_id': req.request_id,
                'employee_name': req.employee.full_name,
                'employee_id': getattr(req.employee, 'employee_id', ''),
                'department': req.employee.department.name if req.employee.department else '',
                'business_function': req.employee.business_function.name if req.employee.business_function else '',
                'vacation_type': req.vacation_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'return_date': req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                'days': float(req.number_of_days),
                'status': req.get_status_display(),
                'status_code': req.status,
                'comment': req.comment,
                'line_manager': req.line_manager.full_name if req.line_manager else '',
                'hr_representative': req.hr_representative.full_name if req.hr_representative else '',
                'attachments_count': attachments_count,
                'has_attachments': attachments_count > 0,
                'created_at': req.created_at.isoformat() if req.created_at else None,
                'updated_at': req.updated_at.isoformat() if req.updated_at else None
            })
        
        # Add schedules
        for sch in schedules:
            all_records.append({
                'id': sch.id,
                'type': 'schedule',
                'request_id': f'SCH{sch.id}',
                'employee_name': sch.employee.full_name,
                'employee_id': getattr(sch.employee, 'employee_id', ''),
                'department': sch.employee.department.name if sch.employee.department else '',
                'business_function': sch.employee.business_function.name if sch.employee.business_function else '',
                'vacation_type': sch.vacation_type.name,
                'start_date': sch.start_date.strftime('%Y-%m-%d'),
                'end_date': sch.end_date.strftime('%Y-%m-%d'),
                'return_date': sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                'days': float(sch.number_of_days),
                'status': sch.get_status_display(),
                'status_code': sch.status,
                'comment': sch.comment,
                'can_edit': sch.can_edit(),
                'edit_count': sch.edit_count,
                'created_by': sch.created_by.get_full_name() if sch.created_by else '',
                'attachments_count': 0,
                'has_attachments': False,
                'created_at': sch.created_at.isoformat() if sch.created_at else None,
                'updated_at': sch.updated_at.isoformat() if sch.updated_at else None
            })
        
        # Sort by created_at desc
        all_records.sort(key=lambda x: x['created_at'] if x['created_at'] else '', reverse=True)
        
        return Response({
            'records': all_records,
            'total_count': len(all_records),
            'requests_count': requests.count(),
            'schedules_count': schedules.count(),
            'access_level': access['access_level'],
            'filters_applied': {
                'status': status_filter,
                'vacation_type_id': vacation_type_id,
                'department_id': department_id,
                'business_function_id': business_function_id,
                'start_date': start_date,
                'end_date': end_date,
                'employee_name': employee_name,
                'year': year
            }
        })
        
    except Exception as e:
        logger.error(f"Error in all_vacation_records: {e}")
        return Response({
            'error': f'Error retrieving records: {str(e)}'
        }, status=rest_status.HTTP_400_BAD_REQUEST)


# ==================== EXPORT ALL VACATION RECORDS ====================
@swagger_auto_schema(
    method='get',
    operation_description="Bütün vacation records-u Excel formatında export et (filterlər dəstəklənir)",
    operation_summary="Export All Vacation Records",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Excel file')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_all_vacation_records(request):
    """✅ Bütün vacation records-u enhanced formatda export et - filtered by access"""
    try:
        access = get_vacation_access(request.user)
        
        # Filter parameters
        status_filter = request.GET.get('status')
        vacation_type_id = request.GET.get('vacation_type_id')
        department_id = request.GET.get('department_id')
        business_function_id = request.GET.get('business_function_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_name = request.GET.get('employee_name')
        year = request.GET.get('year')
        export_format = request.GET.get('format', 'combined')
        
        # All requests base query
        requests_qs = VacationRequest.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'line_manager', 'hr_representative',
            'line_manager_approved_by', 'hr_approved_by', 'rejected_by'
        )
        
        # All schedules base query
        schedules_qs = VacationSchedule.objects.filter(is_deleted=False).select_related(
            'employee', 'employee__department', 'employee__business_function', 
            'vacation_type', 'created_by', 'last_edited_by'
        )
        
        # ✅ Filter by access level
        requests_qs = filter_vacation_queryset(request.user, requests_qs, 'request')
        schedules_qs = filter_vacation_queryset(request.user, schedules_qs, 'schedule')
        
        # Apply other filters
        if status_filter:
            requests_qs = requests_qs.filter(status=status_filter)
            schedules_qs = schedules_qs.filter(status=status_filter)
        
        if vacation_type_id:
            requests_qs = requests_qs.filter(vacation_type_id=vacation_type_id)
            schedules_qs = schedules_qs.filter(vacation_type_id=vacation_type_id)
        
        if department_id:
            requests_qs = requests_qs.filter(employee__department_id=department_id)
            schedules_qs = schedules_qs.filter(employee__department_id=department_id)
        
        if business_function_id:
            requests_qs = requests_qs.filter(employee__business_function_id=business_function_id)
            schedules_qs = schedules_qs.filter(employee__business_function_id=business_function_id)
        
        if start_date:
            requests_qs = requests_qs.filter(start_date__gte=start_date)
            schedules_qs = schedules_qs.filter(start_date__gte=start_date)
        
        if end_date:
            requests_qs = requests_qs.filter(end_date__lte=end_date)
            schedules_qs = schedules_qs.filter(end_date__lte=end_date)
        
        if employee_name:
            requests_qs = requests_qs.filter(employee__full_name__icontains=employee_name)
            schedules_qs = schedules_qs.filter(employee__full_name__icontains=employee_name)
        
        if year:
            requests_qs = requests_qs.filter(start_date__year=year)
            schedules_qs = schedules_qs.filter(start_date__year=year)
        
        # Get data
        requests = requests_qs.order_by('-created_at')
        schedules = schedules_qs.order_by('-created_at')
        
        wb = Workbook()
        
        if export_format == 'separated':
            # Ayrı sheet-lər
            ws_req = wb.active
            ws_req.title = "Vacation Requests"
            
            req_headers = [
                'Request ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
                'Vacation Type', 'Start Date', 'End Date', 'Return Date', 'Working Days',
                'Status', 'Comment', 'Request Type',
                'Line Manager', 'LM Comment', 'LM Approved At', 'LM Approved By',
                'HR Representative', 'HR Comment', 'HR Approved At', 'HR Approved By',
                'Rejected By', 'Rejection Reason', 'Rejected At',
                'Created At', 'Updated At'
            ]
            ws_req.append(req_headers)
            
            for req in requests:
                ws_req.append([
                    req.request_id,
                    req.employee.full_name,
                    getattr(req.employee, 'employee_id', ''),
                    req.employee.department.name if req.employee.department else '',
                    req.employee.business_function.name if req.employee.business_function else '',
                    req.vacation_type.name,
                    req.start_date.strftime('%Y-%m-%d'),
                    req.end_date.strftime('%Y-%m-%d'),
                    req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                    float(req.number_of_days),
                    req.get_status_display(),
                    req.comment,
                    req.get_request_type_display(),
                    req.line_manager.full_name if req.line_manager else '',
                    req.line_manager_comment,
                    req.line_manager_approved_at.strftime('%Y-%m-%d %H:%M') if req.line_manager_approved_at else '',
                    req.line_manager_approved_by.get_full_name() if req.line_manager_approved_by else '',
                    req.hr_representative.full_name if req.hr_representative else '',
                    req.hr_comment,
                    req.hr_approved_at.strftime('%Y-%m-%d %H:%M') if req.hr_approved_at else '',
                    req.hr_approved_by.get_full_name() if req.hr_approved_by else '',
                    req.rejected_by.get_full_name() if req.rejected_by else '',
                    req.rejection_reason,
                    req.rejected_at.strftime('%Y-%m-%d %H:%M') if req.rejected_at else '',
                    req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else '',
                    req.updated_at.strftime('%Y-%m-%d %H:%M') if req.updated_at else ''
                ])
            
            # Schedules sheet
            ws_sch = wb.create_sheet("Vacation Schedules")
            
            sch_headers = [
                'Schedule ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
                'Vacation Type', 'Start Date', 'End Date', 'Return Date', 'Working Days',
                'Status', 'Comment', 
                'Edit Count', 'Can Edit', 'Last Edited By', 'Last Edited At',
                'Created By', 'Created At', 'Updated At'
            ]
            ws_sch.append(sch_headers)
            
            for sch in schedules:
                ws_sch.append([
                    f'SCH{sch.id}',
                    sch.employee.full_name,
                    getattr(sch.employee, 'employee_id', ''),
                    sch.employee.department.name if sch.employee.department else '',
                    sch.employee.business_function.name if sch.employee.business_function else '',
                    sch.vacation_type.name,
                    sch.start_date.strftime('%Y-%m-%d'),
                    sch.end_date.strftime('%Y-%m-%d'),
                    sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                    float(sch.number_of_days),
                    sch.get_status_display(),
                    sch.comment,
                    sch.edit_count,
                    'Yes' if sch.can_edit() else 'No',
                    sch.last_edited_by.get_full_name() if sch.last_edited_by else '',
                    sch.last_edited_at.strftime('%Y-%m-%d %H:%M') if sch.last_edited_at else '',
                    sch.created_by.get_full_name() if sch.created_by else '',
                    sch.created_at.strftime('%Y-%m-%d %H:%M') if sch.created_at else '',
                    sch.updated_at.strftime('%Y-%m-%d %H:%M') if sch.updated_at else ''
                ])
            
        else:
            # Combined sheet (default)
            ws = wb.active
            ws.title = "All Vacation Records"
            
            headers = [
                'Type', 'ID', 'Employee Name', 'Employee ID', 'Department', 'Business Function',
                'Vacation Type', 'Start Date', 'End Date', 'Return Date', 'Working Days',
                'Status', 'Comment', 
                'Line Manager/Created By', 'HR Representative', 'Approval Status',
                'Edit Count', 'Created At', 'Updated At'
            ]
            ws.append(headers)
            
            # Combine and sort all records
            all_records = []
            
            # Add requests
            for req in requests:
                approval_status = []
                if req.line_manager_approved_at:
                    approval_status.append('LM ✓')
                elif req.status == 'PENDING_LINE_MANAGER':
                    approval_status.append('LM ⏳')
                elif req.status == 'REJECTED_LINE_MANAGER':
                    approval_status.append('LM ✗')
                
                if req.hr_approved_at:
                    approval_status.append('HR ✓')
                elif req.status == 'PENDING_HR':
                    approval_status.append('HR ⏳')
                elif req.status == 'REJECTED_HR':
                    approval_status.append('HR ✗')
                
                all_records.append({
                    'type': 'Request',
                    'id': req.request_id,
                    'employee_name': req.employee.full_name,
                    'employee_id': getattr(req.employee, 'employee_id', ''),
                    'department': req.employee.department.name if req.employee.department else '',
                    'business_function': req.employee.business_function.name if req.employee.business_function else '',
                    'vacation_type': req.vacation_type.name,
                    'start_date': req.start_date.strftime('%Y-%m-%d'),
                    'end_date': req.end_date.strftime('%Y-%m-%d'),
                    'return_date': req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                    'working_days': float(req.number_of_days),
                    'status': req.get_status_display(),
                    'comment': req.comment,
                    'manager_created_by': req.line_manager.full_name if req.line_manager else '',
                    'hr_representative': req.hr_representative.full_name if req.hr_representative else '',
                    'approval_status': ' | '.join(approval_status),
                    'edit_count': '',
                    'created_at': req.created_at,
                    'updated_at': req.updated_at
                })
            
            # Add schedules
            for sch in schedules:
                all_records.append({
                    'type': 'Schedule',
                    'id': f'SCH{sch.id}',
                    'employee_name': sch.employee.full_name,
                    'employee_id': getattr(sch.employee, 'employee_id', ''),
                    'department': sch.employee.department.name if sch.employee.department else '',
                    'business_function': sch.employee.business_function.name if sch.employee.business_function else '',
                    'vacation_type': sch.vacation_type.name,
                    'start_date': sch.start_date.strftime('%Y-%m-%d'),
                    'end_date': sch.end_date.strftime('%Y-%m-%d'),
                    'return_date': sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                    'working_days': float(sch.number_of_days),
                    'status': sch.get_status_display(),
                    'comment': sch.comment,
                    'manager_created_by': sch.created_by.get_full_name() if sch.created_by else '',
                    'hr_representative': '',
                    'approval_status': 'No Approval Needed',
                    'edit_count': sch.edit_count,
                    'created_at': sch.created_at,
                    'updated_at': sch.updated_at
                })
            
            # Sort by created_at desc
            all_records.sort(key=lambda x: x['created_at'] if x['created_at'] else datetime.min, reverse=True)
            
            # Add data to sheet
            for record in all_records:
                ws.append([
                    record['type'],
                    record['id'],
                    record['employee_name'],
                    record['employee_id'],
                    record['department'],
                    record['business_function'],
                    record['vacation_type'],
                    record['start_date'],
                    record['end_date'],
                    record['return_date'],
                    record['working_days'],
                    record['status'],
                    record['comment'],
                    record['manager_created_by'],
                    record['hr_representative'],
                    record['approval_status'],
                    record['edit_count'],
                    record['created_at'].strftime('%Y-%m-%d %H:%M') if record['created_at'] else '',
                    record['updated_at'].strftime('%Y-%m-%d %H:%M') if record['updated_at'] else ''
                ])
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # Generate filename
        access_indicator = access['access_level'].replace(' ', '_').replace('-', '')
        filename = f'vacation_records_{access_indicator}_{export_format}_{date.today().strftime("%Y%m%d")}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        return response
        
    except Exception as e:
        logger.error(f"Export error: {e}")
        return Response({'error': f'Export error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


# ==================== EXPORT MY VACATIONS ====================
@swagger_auto_schema(
    method='get',
    operation_description="İstifadəçinin bütün vacation məlumatlarını Excel formatında export et",
    operation_summary="Export My Vacations",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Excel file')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_my_vacations(request):
    """✅ İstifadəçinin vacation məlumatlarını enhanced formatda export et"""
    try:
        access = get_vacation_access(request.user)
        
        if not access['employee']:
            return Response({
                'error': 'Employee profili tapılmadı'
            }, status=status.HTTP_404_NOT_FOUND)
        
        emp = access['employee']
        
        # All requests and schedules - filtered by access
        requests_qs = VacationRequest.objects.filter(is_deleted=False)
        requests_qs = filter_vacation_queryset(request.user, requests_qs, 'request')
        requests = requests_qs.order_by('-created_at')
        
        schedules_qs = VacationSchedule.objects.filter(is_deleted=False)
        schedules_qs = filter_vacation_queryset(request.user, schedules_qs, 'schedule')
        schedules = schedules_qs.order_by('-start_date')
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Personal Summary"
        ws_summary.sheet_view.showGridLines = False
        
        # Styles
        title_font = Font(size=16, bold=True, color="2B4C7E")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        stat_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        
        # Personal info header
        ws_summary['A1'] = f'VACATION SUMMARY - {emp.full_name}'
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:E1')
        
        ws_summary['A2'] = f'Employee ID: {getattr(emp, "employee_id", "N/A")}'
        ws_summary['A3'] = f'Department: {emp.department.name if emp.department else "N/A"}'
        ws_summary['A4'] = f'Access Level: {access["access_level"]}'
        ws_summary['A5'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        
        # Get years with data
        years_with_data = set()
        
        for req in requests:
            years_with_data.add(req.start_date.year)
        
        for sch in schedules:
            years_with_data.add(sch.start_date.year)
        
        balances_with_data = EmployeeVacationBalance.objects.filter(
            is_deleted=False
        ).exclude(
            yearly_balance=0,
            used_days=0,
            scheduled_days=0
        ).values_list('year', flat=True)
        
        # ✅ Filter balances by access
        if access['accessible_employee_ids'] is not None:
            balances_with_data = EmployeeVacationBalance.objects.filter(
                employee_id__in=access['accessible_employee_ids'],
                is_deleted=False
            ).exclude(
                yearly_balance=0,
                used_days=0,
                scheduled_days=0
            ).values_list('year', flat=True)
        
        years_with_data.update(balances_with_data)
        
        if not years_with_data:
            years_with_data.add(date.today().year)
        
        sorted_years = sorted(years_with_data, reverse=True)
        
        ws_summary['A7'] = 'YEARLY STATISTICS'
        ws_summary['A7'].font = Font(size=14, bold=True, color="2B4C7E")
        
        year_headers = ['Year', 'Total Balance', 'Used Days', 'Scheduled Days', 'Remaining', 'Requests', 'Schedules']
        for col, header in enumerate(year_headers, 1):
            cell = ws_summary.cell(row=9, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for i, year in enumerate(sorted_years, 10):
            # Get balance for accessible employees
            balance_qs = EmployeeVacationBalance.objects.filter(year=year, is_deleted=False)
            if access['accessible_employee_ids'] is not None:
                balance_qs = balance_qs.filter(employee_id__in=access['accessible_employee_ids'])
            
            balance = balance_qs.first()
            year_requests = requests.filter(start_date__year=year).count()
            year_schedules = schedules.filter(start_date__year=year).count()
            
            if balance:
                data = [
                    year,
                    float(balance.total_balance),
                    float(balance.used_days),
                    float(balance.scheduled_days),
                    float(balance.remaining_balance),
                    year_requests,
                    year_schedules
                ]
            else:
                data = [year, 0, 0, 0, 0, year_requests, year_schedules]
            
            for col, value in enumerate(data, 1):
                cell = ws_summary.cell(row=i, column=col, value=value)
                if col > 1:
                    cell.fill = stat_fill
                cell.alignment = Alignment(horizontal='center' if col == 1 else 'right')
        
        # Requests sheet
        ws_requests = wb.create_sheet("Vacation Requests")
        
        req_headers = [
            'Request ID', 'Type', 'Vacation Type', 'Start Date', 'End Date', 'Return Date',
            'Days', 'Status', 'Comment', 'Line Manager', 'LM Comment', 'HR Comment', 
            'Created At', 'Timeline'
        ]
        
        for col, header in enumerate(req_headers, 1):
            cell = ws_requests.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Status colors
        status_colors = {
            'APPROVED': good_fill,
            'PENDING_LINE_MANAGER': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'PENDING_HR': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
            'REJECTED_LINE_MANAGER': warning_fill,
            'REJECTED_HR': warning_fill,
        }
        
        for row, req in enumerate(requests, 2):
            timeline = f"Created: {req.created_at.strftime('%Y-%m-%d')}"
            if req.line_manager_approved_at:
                timeline += f" → LM: {req.line_manager_approved_at.strftime('%Y-%m-%d')}"
            if req.hr_approved_at:
                timeline += f" → HR: {req.hr_approved_at.strftime('%Y-%m-%d')}"
            if req.rejected_at:
                timeline += f" → Rejected: {req.rejected_at.strftime('%Y-%m-%d')}"
            
            data = [
                req.request_id,
                req.get_request_type_display(),
                req.vacation_type.name,
                req.start_date.strftime('%Y-%m-%d'),
                req.end_date.strftime('%Y-%m-%d'),
                req.return_date.strftime('%Y-%m-%d') if req.return_date else '',
                float(req.number_of_days),
                req.get_status_display(),
                req.comment,
                req.line_manager.full_name if req.line_manager else '',
                req.line_manager_comment,
                req.hr_comment,
                req.created_at.strftime('%Y-%m-%d %H:%M'),
                timeline
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_requests.cell(row=row, column=col, value=value)
                if col == 8:  # Status column
                    cell.fill = status_colors.get(req.status, PatternFill())
        
        # Schedules sheet
        ws_schedules = wb.create_sheet("Vacation Schedules")
        
        sch_headers = [
            'Schedule ID', 'Vacation Type', 'Start Date', 'End Date', 'Return Date',
            'Days', 'Status', 'Comment', 'Edit Count', 'Can Edit', 'Created At',
            'Last Edited', 'Timeline'
        ]
        
        for col, header in enumerate(sch_headers, 1):
            cell = ws_schedules.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, sch in enumerate(schedules, 2):
            timeline = f"Created: {sch.created_at.strftime('%Y-%m-%d')}"
            if sch.last_edited_at:
                timeline += f" → Last Edit: {sch.last_edited_at.strftime('%Y-%m-%d')}"
            
            data = [
                f'SCH{sch.id}',
                sch.vacation_type.name,
                sch.start_date.strftime('%Y-%m-%d'),
                sch.end_date.strftime('%Y-%m-%d'),
                sch.return_date.strftime('%Y-%m-%d') if sch.return_date else '',
                float(sch.number_of_days),
                sch.get_status_display(),
                sch.comment,
                sch.edit_count,
                'Yes' if sch.can_edit() else 'No',
                sch.created_at.strftime('%Y-%m-%d %H:%M'),
                sch.last_edited_at.strftime('%Y-%m-%d %H:%M') if sch.last_edited_at else '',
                timeline
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_schedules.cell(row=row, column=col, value=value)
                if col == 7:  # Status column
                    if sch.status == 'REGISTERED':
                        cell.fill = good_fill
                    else:
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Auto-adjust column widths
        for ws_current in [ws_summary, ws_requests, ws_schedules]:
            for column in ws_current.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws_current.column_dimensions[column_letter].width = adjusted_width
        
        filename = f'my_vacations_{emp.full_name.replace(" ", "_")}_{date.today().strftime("%Y%m%d")}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        return response
        
    except Exception as e:
        logger.error(f"Error in export_my_vacations: {e}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ==================== SETTINGS VIEWSETS ====================
class VacationSettingViewSet(viewsets.ModelViewSet):
    """✅ Vacation Settings CRUD - ADMIN ONLY"""
    queryset = VacationSetting.objects.filter(is_deleted=False)
    serializer_class = VacationSettingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # ✅ Admin only can manage settings
        if not is_admin_user(self.request.user):
            return VacationSetting.objects.none()
        return super().get_queryset()
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


class VacationTypeViewSet(viewsets.ModelViewSet):
    """✅ Vacation Types - ADMIN can manage, everyone can view"""
    queryset = VacationType.objects.filter(is_deleted=False)
    serializer_class = VacationTypeSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        # Everyone can view vacation types
        return super().list(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        # ✅ Admin only
        if not is_admin_user(request.user):
            return Response({
                'error': 'Permission denied',
                'detail': 'Only admin can create vacation types'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        # ✅ Admin only
        if not is_admin_user(request.user):
            return Response({
                'error': 'Permission denied',
                'detail': 'Only admin can update vacation types'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        # ✅ Admin only
        if not is_admin_user(request.user):
            return Response({
                'error': 'Permission denied',
                'detail': 'Only admin can update vacation types'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        # ✅ Admin only
        if not is_admin_user(request.user):
            return Response({
                'error': 'Permission denied',
                'detail': 'Only admin can delete vacation types'
            }, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

@swagger_auto_schema(
    method='get',
    operation_description="✅ Get vacation types filtered by business function",
    operation_summary="Get Vacation Types (Filtered)",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Vacation types')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_types_filtered(request):
    """
    ✅ Vacation types - UK employee-lar üçün Half Day görünür
    """
    try:
        access = get_vacation_access(request.user)
        
        # Base queryset
        types_qs = VacationType.objects.filter(is_active=True, is_deleted=False)
        
        # Check if UK employee
        is_uk = False
        if access['employee'] and access['employee'].business_function:
            code = getattr(access['employee'].business_function, 'code', '')
            is_uk = code.upper() == 'UK'
        
        # Filter UK-only types
        if not is_uk:
            types_qs = types_qs.exclude(is_uk_only=True)
        
        serializer = VacationTypeSerializer(types_qs, many=True)
        
        return Response({
            'types': serializer.data,
            'is_uk_employee': is_uk,
            'total_count': types_qs.count()
        })
        
    except Exception as e:
        logger.error(f"Error getting vacation types: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)

# ==================== FILE UPLOAD ENDPOINTS ====================
@swagger_auto_schema(
    method='get',
    operation_description="Get all attachments for a vacation request",
    operation_summary="List Vacation Request Attachments",
    tags=['Vacation - Files'],
    responses={200: openapi.Response(description='List of attachments')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_vacation_request_attachments(request, request_id):
    """✅ Get all attachments for a vacation request - access controlled"""
    try:
        vacation_request = get_object_or_404(
            VacationRequest, 
            request_id=request_id, 
            is_deleted=False
        )
        
        access = get_vacation_access(request.user)
        
        # ✅ Check access permission
        can_view = False
        if access['is_admin']:
            can_view = True
        elif access['employee'] and vacation_request.employee == access['employee']:
            can_view = True
        elif vacation_request.requester == request.user:
            can_view = True
        elif access['employee'] and (vacation_request.line_manager == access['employee'] or vacation_request.hr_representative == access['employee']):
            can_view = True
        elif access['accessible_employee_ids'] and vacation_request.employee_id in access['accessible_employee_ids']:
            can_view = True
        
        if not can_view:
            return Response({
                'error': 'Permission denied'
            }, status=status.HTTP_403_FORBIDDEN)
        
        attachments = vacation_request.attachments.filter(is_deleted=False).order_by('-uploaded_at')
        
        return Response({
            'request_id': request_id,
            'count': attachments.count(),
            'attachments': VacationAttachmentSerializer(
                attachments, 
                many=True, 
                context={'request': request}
            ).data
        })
        
    except VacationRequest.DoesNotExist:
        return Response({
            'error': 'Vacation request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error listing attachments: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='delete',
    operation_description="Delete a file attachment (Only uploader or admin can delete)",
    operation_summary="Delete Vacation Attachment",
    tags=['Vacation - Files'],
    responses={200: 'File deleted successfully'}
)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_vacation_attachment(request, attachment_id):
    """✅ Delete a file attachment - Only uploader or admin"""
    try:
        from .vacation_models import VacationAttachment
        
        attachment = get_object_or_404(
            VacationAttachment, 
            id=attachment_id, 
            is_deleted=False
        )
        
        # ✅ Check permission - only uploader or admin can delete
        if attachment.uploaded_by != request.user and not is_admin_user(request.user):
            return Response({
                'error': 'You can only delete files you uploaded'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Soft delete
        attachment.is_deleted = True
        attachment.save()
        
        return Response({
            'message': 'File deleted successfully',
            'filename': attachment.original_filename
        })
        
    except Exception as e:
        logger.error(f"Error deleting file: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='post',
    operation_description="Upload multiple files at once to vacation request",
    operation_summary="Bulk Upload Vacation Attachments",
    tags=['Vacation - Files'],
    responses={201: 'Files uploaded successfully'}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_vacation_attachments(request, request_id):
    """✅ Upload multiple files at once - access controlled"""
    try:
        vacation_request = get_object_or_404(
            VacationRequest, 
            request_id=request_id, 
            is_deleted=False
        )
        
        access = get_vacation_access(request.user)
        
        # ✅ Check permission
        can_upload = False
        if access['employee'] and vacation_request.employee == access['employee']:
            can_upload = True
        elif vacation_request.requester == request.user:
            can_upload = True
        elif is_admin_user(request.user):
            can_upload = True
        
        if not can_upload:
            return Response({
                'error': 'Permission denied'
            }, status=status.HTTP_403_FORBIDDEN)
        
        files = request.FILES.getlist('files')
        if not files:
            return Response({
                'error': 'No files provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_attachments = []
        errors = []
        
        for file in files:
            try:
                from .business_trip_serializers import TripAttachmentUploadSerializer
                upload_serializer = TripAttachmentUploadSerializer(data={'file': file})
                if not upload_serializer.is_valid():
                    errors.append({
                        'filename': file.name,
                        'errors': upload_serializer.errors
                    })
                    continue
                
                from .vacation_models import VacationAttachment
                attachment = VacationAttachment.objects.create(
                    vacation_request=vacation_request,
                    file=file,
                    original_filename=file.name,
                    file_size=file.size,
                    file_type=file.content_type,
                    uploaded_by=request.user
                )
                uploaded_attachments.append(attachment)
                
            except Exception as e:
                errors.append({
                    'filename': file.name,
                    'error': str(e)
                })
        
        return Response({
            'message': f'{len(uploaded_attachments)} files uploaded successfully',
            'uploaded': VacationAttachmentSerializer(
                uploaded_attachments, 
                many=True, 
                context={'request': request}
            ).data,
            'errors': errors,
            'success_count': len(uploaded_attachments),
            'error_count': len(errors)
        }, status=status.HTTP_201_CREATED)
        
    except VacationRequest.DoesNotExist:
        return Response({
            'error': 'Vacation request not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error in bulk upload: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method='get',
    operation_description="Get attachment details",
    operation_summary="Get Attachment Details",
    tags=['Vacation - Files'],
    responses={200: openapi.Response(description='Attachment details')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_attachment_details(request, attachment_id):
    """✅ Get details of a specific attachment"""
    try:
        from .vacation_models import VacationAttachment
        
        attachment = get_object_or_404(
            VacationAttachment, 
            id=attachment_id, 
            is_deleted=False
        )
        
        return Response({
            'attachment': VacationAttachmentSerializer(
                attachment, 
                context={'request': request}
            ).data
        })
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# ==================== SCHEDULE DETAIL ====================
@swagger_auto_schema(
    method='get',
    operation_description="Get detailed information of a vacation schedule",
    operation_summary="Get Vacation Schedule Detail",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Vacation schedule details')}
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vacation_schedule_detail(request, pk):
    """✅ Get detailed information of a vacation schedule - access controlled"""
    try:
        schedule = VacationSchedule.objects.select_related(
            'employee', 
            'employee__department',
            'employee__business_function',
            'employee__unit',
            'employee__job_function',
            'vacation_type',
            'created_by',
            'last_edited_by'
        ).get(pk=pk, is_deleted=False)
        
        access = get_vacation_access(request.user)
        
        # ✅ Check access permission
        can_view = False
        
        if access['is_admin']:
            can_view = True
        elif access['employee'] and schedule.employee == access['employee']:
            can_view = True
        elif schedule.created_by == request.user:
            can_view = True
        elif access['employee'] and schedule.employee.line_manager == access['employee']:
            can_view = True
        elif access['accessible_employee_ids'] and schedule.employee_id in access['accessible_employee_ids']:
            can_view = True
        
        if not can_view:
            return Response({
                'error': 'Permission denied',
                'detail': 'You do not have permission to view this vacation schedule'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Serialize the data
        serializer = VacationScheduleSerializer(
            schedule, 
            context={'request': request}
        )
        
        response_data = serializer.data
        
        # Add employee information
        response_data['employee_info'] = {
            'id': schedule.employee.id,
            'name': schedule.employee.full_name,
            'employee_id': getattr(schedule.employee, 'employee_id', ''),
            'department': schedule.employee.department.name if schedule.employee.department else None,
            'business_function': schedule.employee.business_function.name if schedule.employee.business_function else None,
            'phone': schedule.employee.phone
        }
        
        # Add edit history
        response_data['edit_history'] = {
            'edit_count': schedule.edit_count,
            'can_edit': schedule.can_edit(),
            'last_edited_by': schedule.last_edited_by.get_full_name() if schedule.last_edited_by else None,
            'last_edited_at': schedule.last_edited_at,
            'max_edits_allowed': VacationSetting.get_active().max_schedule_edits if VacationSetting.get_active() else 3
        }
        
        # Add creator information
        response_data['creator_info'] = {
            'name': schedule.created_by.get_full_name() if schedule.created_by else None,
            'email': schedule.created_by.email if schedule.created_by else None
        }
        
        # ✅ Add permission flags for frontend
        can_modify, modify_reason = can_user_modify_schedule(request.user, schedule)
        can_register, register_reason = can_user_register_schedule(request.user)
        
        response_data['permissions'] = {
            'can_edit': (
                schedule.status == 'SCHEDULED' and
                access['employee'] and schedule.employee == access['employee'] and
                schedule.can_edit()
            ),
            'can_delete': can_modify,
            'can_register': can_register,
            'is_admin': access['is_admin'],
            'access_level': access['access_level']
        }
        
        return Response(response_data)
        
    except VacationSchedule.DoesNotExist:
        return Response({
            'error': 'Vacation schedule not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error fetching vacation schedule detail: {e}")
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# ==================== UTILITIES ====================
@swagger_auto_schema(
    method='post',
    operation_description="✅ Calculate working days with business function support",
    operation_summary="Calculate Working Days",
    tags=['Vacation'],
    responses={200: openapi.Response(description='Calculation result')}
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculate_working_days(request):
    """✅ İş günlərini hesabla - business function dəstəyi"""
    start = request.data.get('start_date')
    end = request.data.get('end_date')
    business_function_code = request.data.get('business_function_code')  # ✅ NEW
    
    if not start or not end:
        return Response({
            'error': 'start_date və end_date mütləqdir'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    settings = VacationSetting.get_active()
    if settings:
        try:
            start_dt = datetime.strptime(start, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end, '%Y-%m-%d').date()
            
            if start_dt > end_dt:
                return Response({
                    'error': 'start_date end_date-dən kiçik olmalıdır'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # ✅ Calculate with business function
            days = settings.calculate_working_days(start_dt, end_dt, business_function_code)
            return_date = settings.calculate_return_date(end_dt, business_function_code)
            
            return Response({
                'working_days': days,
                'return_date': return_date.strftime('%Y-%m-%d'),
                'total_calendar_days': (end_dt - start_dt).days + 1,
                'business_function_code': business_function_code,
                'calculation_method': 'UK (excludes weekends & holidays)' if business_function_code and business_function_code.upper() == 'UK' else 'Azerbaijan (excludes holidays only)'
            })
            
        except ValueError:
            return Response({
                'error': 'Tarix formatı səhvdir. YYYY-MM-DD istifadə edin'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({
        'error': 'Settings tapılmadı'
    }, status=status.HTTP_404_NOT_FOUND)