from django.utils import timezone
from rest_framework import status, viewsets
from django.db.models import Q, Count
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status, viewsets, filters
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from django.db import models 
import logging
from .job_description_models import JobDescription,JobDescriptionAssignment
import traceback
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from datetime import datetime, timedelta, date
from django.utils.dateparse import parse_date
from django.db import transaction
from django.http import HttpResponse
import csv
import io
import pandas as pd
from django.contrib.auth.models import User
from .headcount_permissions import get_headcount_access, filter_headcount_queryset

from rest_framework.exceptions import AuthenticationFailed  # For 401 errors

from .models import (
    Employee, BusinessFunction, Department, Unit, JobFunction, 
    PositionGroup, EmployeeTag, EmployeeStatus,
    EmployeeActivity, VacantPosition, ContractTypeConfig,
     EmployeeArchive,EmployeeDocument,
    UserGraphToken, JobTitle
)

from .serializers import (
    EmployeeListSerializer, EmployeeDetailSerializer,
    BusinessFunctionSerializer, DepartmentSerializer, UnitSerializer,
    JobFunctionSerializer, PositionGroupSerializer, EmployeeTagSerializer,
    EmployeeStatusSerializer,  EmployeeActivitySerializer,
    UserSerializer, OrgChartNodeSerializer,
    VacantPositionListSerializer, VacantPositionDetailSerializer, VacantPositionCreateSerializer,
     ProfileImageDeleteSerializer,BulkHardDeleteSerializer,
    ProfileImageUploadSerializer, EmployeeDocumentSerializer,
    ContractTypeConfigSerializer, BulkContractExtensionSerializer, ContractExtensionSerializer,
    SingleEmployeeTagUpdateSerializer, SingleLineManagerAssignmentSerializer,
    BulkEmployeeTagUpdateSerializer, JobTitleSerializer,
    BulkLineManagerAssignmentSerializer,VacancyToEmployeeConversionSerializer,EmployeeJobDescriptionSerializer,ManagerJobDescriptionSerializer
)

from .asset_permissions import get_asset_access_level
from .auth import MicrosoftTokenValidator
from drf_yasg.inspectors import SwaggerAutoSchema
logger = logging.getLogger(__name__)


class ModernPagination(PageNumberPagination):
    """Modern, user-friendly pagination - DEFAULT: No pagination unless requested"""
    page_size = 20  # Default page size when pagination is used
    page_size_query_param = 'page_size'
    max_page_size = 1000  # Increased max page size
    page_query_param = 'page'
    
    # Custom page size options for frontend
    page_size_options = [10, 20, 50, 100, 500, 1000, "All"]
    
    def get_paginated_response(self, data):
        """Enhanced pagination response with modern UI support"""
        current_page = self.page.number
        total_pages = self.page.paginator.num_pages
        total_count = self.page.paginator.count
        
        # Calculate pagination window (show 5 pages around current)
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)
        
        # Adjust window if we're near the beginning or end
        if end_page - start_page < 4:
            if start_page == 1:
                end_page = min(total_pages, start_page + 4)
            else:
                start_page = max(1, end_page - 4)
        
        # Generate page numbers for frontend
        page_numbers = list(range(start_page, end_page + 1))
        
        # Calculate range display
        start_item = (current_page - 1) * self.page_size + 1
        end_item = min(current_page * self.page_size, total_count)
        
        return Response({
            'count': total_count,
            'total_pages': total_pages,
            'current_page': current_page,
            'page_size': self.page_size,
            'page_size_options': self.page_size_options,
            'has_next': self.page.has_next(),
            'has_previous': self.page.has_previous(),
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'page_numbers': page_numbers,  # For modern pagination UI
            'start_page': start_page,
            'end_page': end_page,
            'show_first': start_page > 1,
            'show_last': end_page < total_pages,
            'range_display': f"Showing {start_item}-{end_item} of {total_count}",
            'pagination_used': True,  # NEW: Indicates pagination was used
            'results': data
        })

class FileUploadAutoSchema(SwaggerAutoSchema):
    """Custom schema for file upload endpoints"""
    
    def get_consumes(self):
        """Force multipart/form-data for file uploads"""
        if self.method.lower() in ['post', 'put', 'patch']:
            return ['multipart/form-data']
        return super().get_consumes()
    
    def get_request_body_schema(self, serializer):
        """Override request body schema for file uploads"""
        if self.method.lower() in ['post', 'put', 'patch']:
            return None  # Don't generate request body schema, use manual parameters
        return super().get_request_body_schema(serializer)

@swagger_auto_schema(
    method='post',
    operation_description="Authenticate with Microsoft Azure AD",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['id_token'],
        properties={
            'id_token': openapi.Schema(type=openapi.TYPE_STRING, description='Microsoft ID Token'),
            'graph_access_token': openapi.Schema(type=openapi.TYPE_STRING, description='Microsoft Graph Access Token'),
        },
    ),
    responses={
        200: openapi.Response(description='Successful authentication'),
        400: 'Bad Request',
        401: 'Unauthorized'
    },
    tags=['Authentication']
)
@api_view(['POST'])
@permission_classes([AllowAny])
def authenticate_microsoft(request):
    """
    ✅ FIXED: Authenticate with Microsoft and return YOUR JWT tokens
    
    Flow:
    1. Validate Microsoft ID token
    2. Store Microsoft Graph token (for Graph API calls)
    3. Generate YOUR OWN JWT tokens (for YOUR API access)
    4. Return YOUR JWT tokens to frontend
    """
 
    
    try:
        id_token = request.data.get('id_token')
        graph_access_token = request.data.get('graph_access_token')
        
        if not id_token:
            logger.error('No id_token provided')
            return Response(
                {'error': 'id_token is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logger.info(f'Token length: {len(id_token)}, Graph token: {bool(graph_access_token)}')
        
        # ✅ Validate Microsoft token AND get YOUR JWT tokens
        user, access_token, refresh_token = MicrosoftTokenValidator.validate_token_and_create_jwt(
            id_token, 
            graph_access_token
        )
        
        logger.info(f'✅ Authentication successful for user: {user.username}')
        
        # ✅ Verify Graph token was stored
        if graph_access_token:
            stored_token = UserGraphToken.get_valid_token(user)
            if stored_token:
                logger.info(f'✅ Graph token verified in database for {user.username}')
            else:
                logger.warning(f'⚠️ Graph token not found in database for {user.username}')
        
        # ✅ Enhanced user data
        user_data = {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'has_graph_token': bool(graph_access_token),
        }
        
        # ✅ CRITICAL: Return YOUR JWT tokens, not Microsoft tokens
        logger.info(f'✅ Returning JWT tokens for {user.username}')
        
        return Response({
            'success': True,
            'access': access_token,      # ← YOUR JWT access token for API calls
            'refresh': refresh_token,     # ← YOUR JWT refresh token
            'user': user_data,
            'graph_token_stored': bool(graph_access_token),
            'token_info': {
                'type': 'JWT',
                'access_token_length': len(access_token),
                'use_for_api_calls': True,
                'microsoft_graph_available': bool(graph_access_token)
            }
        })
        
    except AuthenticationFailed as e:
        logger.error(f'❌ Authentication failed: {str(e)}')
        return Response(
            {'error': str(e)},
            status=status.HTTP_401_UNAUTHORIZED
        )
    except Exception as e:
        logger.error(f'❌ Unexpected error: {str(e)}')
        import traceback
        logger.error(f'Traceback: {traceback.format_exc()}')
        return Response(
            {'error': f'Authentication failed: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

@swagger_auto_schema(
    method='get',
    operation_description="Get current user information",
    responses={
        200: openapi.Response(
            description="User information retrieved successfully",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                    'user': openapi.Schema(type=openapi.TYPE_OBJECT),
                    'employee': openapi.Schema(type=openapi.TYPE_OBJECT),
                }
            )
        ),
        401: openapi.Response(description="Unauthorized - Invalid or missing token")
    },
    security=[{'Bearer': []}]
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_info(request):
    """Get current user info"""
    try:
       
        serializer = UserSerializer(request.user)
   
        try:
            employee = Employee.objects.select_related(
                'user', 'business_function', 'department', 'unit', 
                'job_function', 'position_group', 'status', 'line_manager'
            ).prefetch_related('tags').get(user=request.user)
            
            employee_data = EmployeeDetailSerializer(employee).data
            
            
        except Employee.DoesNotExist:
            logger.info(f'[{request.user.username}] No employee profile found')
            employee_data = None
        except Exception as e:
            logger.error(f'[{request.user.username}] Error during employee profile processing for user ID {request.user.id}: {str(e)}')
            logger.error(f'[{request.user.username}] Employee processing traceback: {traceback.format_exc()}')
            employee_data = None
        
        response_data = {
            'success': True,
            'user': serializer.data,
            'employee': employee_data
        }
        

        return Response(response_data, status=status.HTTP_200_OK)
    
    except Exception as e:
        logger.error(f'[{request.user.username}] Unhandled error in user_info: {str(e)}')
        logger.error(f'[{request.user.username}] Full traceback: {traceback.format_exc()}')
        return Response({
            "error": f"Failed to get user info: {str(e)}",
            "success": False
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ComprehensiveEmployeeFilter:
    """
    COMPLETELY FIXED: Frontend component-lərinə uyğun tam filter sistemi
    Comma-separated values-ları düzgün parse edir və backend-də işləyir
    """
    
    def __init__(self, queryset, params):
        self.queryset = queryset
        self.params = params
        # ✅ NEW: Cache parsed values to avoid re-parsing
        self._parsed_cache = {}
    
    def parse_comma_separated(self, param_value):
        """Parse comma-separated string into list of cleaned values"""
        if not param_value:
            return []
        
        if isinstance(param_value, list):
            # Already a list - flatten and clean
            result = []
            for item in param_value:
                if isinstance(item, str) and ',' in item:
                    # Split comma-separated items in list
                    result.extend([val.strip() for val in item.split(',') if val.strip()])
                elif item:
                    result.append(str(item).strip())
            return result
        elif isinstance(param_value, str):
            # Split comma-separated string
            return [val.strip() for val in param_value.split(',') if val.strip()]
        else:
            return [str(param_value).strip()] if param_value else []
    
    def parse_int_list(self, param_value):
        """Parse comma-separated string into list of integers"""
        string_values = self.parse_comma_separated(param_value)
        int_values = []
        for val in string_values:
            try:
                int_values.append(int(val))
            except (ValueError, TypeError):
                continue
        return int_values
    
    def get_filter_values(self, param_name):
        """Get filter values, handling both getlist() and comma-separated strings"""
        # ✅ Check cache first
        if param_name in self._parsed_cache:
            return self._parsed_cache[param_name]
        
        # Try getlist first (for Django QueryDict)
        if hasattr(self.params, 'getlist'):
            values = self.params.getlist(param_name)
            if values:
                # Process each value in case it contains comma-separated items
                all_values = []
                for value in values:
                    all_values.extend(self.parse_comma_separated(value))
                # ✅ Cache the result
                self._parsed_cache[param_name] = all_values
                return all_values
        
        # Fallback to get() for single value (might be comma-separated)
        single_value = self.params.get(param_name)
        if single_value:
            result = self.parse_comma_separated(single_value)
            # ✅ Cache the result
            self._parsed_cache[param_name] = result
            return result
        
        return []
    
    def get_int_filter_values(self, param_name):
        """Get integer filter values"""
        # ✅ Check cache first
        cache_key = f"{param_name}_int"
        if cache_key in self._parsed_cache:
            return self._parsed_cache[cache_key]
        
        string_values = self.get_filter_values(param_name)
        int_values = []
        for val in string_values:
            try:
                int_values.append(int(val))
            except (ValueError, TypeError):
                continue
        
        # ✅ Cache the result
        self._parsed_cache[cache_key] = int_values
        return int_values
    
    def filter(self):
        queryset = self.queryset
        
        print(f"🔍 FILTER DEBUG: Raw params = {dict(self.params)}")
        
        # ===========================================
        # 1. SEARCH FILTERS (Text-based)
        # ===========================================
        
        # General search - multiple fields
        search = self.params.get('search')
        if search:
            print(f"🔍 Applying general search: {search}")
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(employee_id__icontains=search) |
                Q(user__email__icontains=search) |
                Q(job_title__icontains=search) |
                Q(business_function__name__icontains=search) |
                Q(department__name__icontains=search) |
                Q(father_name__icontains=search) |
                Q(job_function__name__icontains=search) | 
                Q(phone__icontains=search)
            )
        
        # FIXED: Specific employee search (from employee_search field)
        employee_search_values = self.get_filter_values('employee_search')
        if employee_search_values:
            print(f"🔍 Applying employee search: {employee_search_values}")
            # Try to find by ID first, then by other fields
            employee_q = Q()
            for search_val in employee_search_values:
                try:
                    # Try as integer ID first
                    emp_id = int(search_val)
                    employee_q |= Q(id=emp_id)
                except (ValueError, TypeError):
                    pass
                
                # Also search by string fields
                employee_q |= (
                    Q(employee_id__icontains=search_val) |
                    Q(full_name__icontains=search_val) |
                    Q(user__first_name__icontains=search_val) |
                    Q(user__last_name__icontains=search_val) |
                    Q(user__email__icontains=search_val)
                )
            
            if employee_q:
                queryset = queryset.filter(employee_q)
        
        # Line manager search
        line_manager_search = self.params.get('line_manager_search')
        if line_manager_search:
            print(f"🔍 Applying line manager search: {line_manager_search}")
            queryset = queryset.filter(
                Q(line_manager__id=line_manager_search) |
                Q(line_manager__employee_id__icontains=line_manager_search) |
                Q(line_manager__full_name__icontains=line_manager_search)
            )
        
        # Job title search
        job_title_search = self.params.get('job_title_search')
        if job_title_search:
            print(f"🔍 Applying job title search: {job_title_search}")
            queryset = queryset.filter(job_title__icontains=job_title_search)
        
        # ===========================================
        # 2. MULTI-SELECT FILTERS (Arrays) - COMPLETELY FIXED
        # ===========================================
        
        # FIXED: Business Functions (array)
        business_function_ids = self.get_int_filter_values('business_function')
        if business_function_ids:
            print(f"🏭 Applying business function filter: {business_function_ids}")
            queryset = queryset.filter(business_function__id__in=business_function_ids)
        
        # FIXED: Departments (array)
        department_ids = self.get_int_filter_values('department')
        if department_ids:
            print(f"🏢 Applying department filter: {department_ids}")
            queryset = queryset.filter(department__id__in=department_ids)
        
        # FIXED: Units (array)
        unit_ids = self.get_int_filter_values('unit')
        if unit_ids:
            print(f"🏢 Applying unit filter: {unit_ids}")
            queryset = queryset.filter(unit__id__in=unit_ids)
        
        # FIXED: Job Functions (array)
        job_function_ids = self.get_int_filter_values('job_function')
        if job_function_ids:
 
            queryset = queryset.filter(job_function__id__in=job_function_ids)
        
        # FIXED: Position Groups (array)
        position_group_ids = self.get_int_filter_values('position_group')
        if position_group_ids:
         
            queryset = queryset.filter(position_group__id__in=position_group_ids)
        

           
                
        grading_levels = self.get_filter_values('grading_level')
        if grading_levels:
         
            queryset = queryset.filter(grading_level__in=grading_levels)
        
        # FIXED: Contract Duration (array)
        contract_durations = self.get_filter_values('contract_duration')
        if contract_durations:
        
            queryset = queryset.filter(contract_duration__in=contract_durations)
        
        # FIXED: Line Managers (array)
        line_manager_ids = self.get_int_filter_values('line_manager')
        if line_manager_ids:
        
            queryset = queryset.filter(line_manager__id__in=line_manager_ids)
        
        # FIXED: Tags (array)
        tag_ids = self.get_int_filter_values('tags')
        if tag_ids:
           
            queryset = queryset.filter(tags__id__in=tag_ids).distinct()
        
        # FIXED: Gender (array)
        genders = self.get_filter_values('gender')
        if genders:
           
            queryset = queryset.filter(gender__in=genders)
        
        # ===========================================
        # 3. DATE RANGE FILTERS
        # ===========================================
        
        # Start Date Range
        start_date_from = self.params.get('start_date_from')
        start_date_to = self.params.get('start_date_to')
        if start_date_from:
            try:
                start_date_from_parsed = parse_date(start_date_from)
                if start_date_from_parsed:
               
                    queryset = queryset.filter(start_date__gte=start_date_from_parsed)
            except:
                pass
        if start_date_to:
            try:
                start_date_to_parsed = parse_date(start_date_to)
                if start_date_to_parsed:
              
                    queryset = queryset.filter(start_date__lte=start_date_to_parsed)
            except:
                pass
        
        # Contract End Date Range
        contract_end_date_from = self.params.get('contract_end_date_from')
        contract_end_date_to = self.params.get('contract_end_date_to')
        if contract_end_date_from:
            try:
                contract_end_from_parsed = parse_date(contract_end_date_from)
                if contract_end_from_parsed:
                 
                    queryset = queryset.filter(contract_end_date__gte=contract_end_from_parsed)
            except:
                pass
        if contract_end_date_to:
            try:
                contract_end_to_parsed = parse_date(contract_end_date_to)
                if contract_end_to_parsed:
                   
                    queryset = queryset.filter(contract_end_date__lte=contract_end_to_parsed)
            except:
                pass
        
        # ===========================================
        # 4. NUMERIC/RANGE FILTERS
        # ===========================================
        
        # Years of Service Range
        years_of_service_min = self.params.get('years_of_service_min')
        years_of_service_max = self.params.get('years_of_service_max')
        
        if years_of_service_min or years_of_service_max:
            today = date.today()
            
            if years_of_service_min:
                try:
                    min_years = float(years_of_service_min)
                    # Employee should have started at least min_years ago
                    min_date = today - timedelta(days=int(min_years * 365.25))
                   
                    queryset = queryset.filter(start_date__lte=min_date)
                except:
                    pass
            
            if years_of_service_max:
                try:
                    max_years = float(years_of_service_max)
                    # Employee should have started at most max_years ago
                    max_date = today - timedelta(days=int(max_years * 365.25))
                 
                    queryset = queryset.filter(start_date__gte=max_date)
                except:
                    pass
        
        # ===========================================
        # 5. BOOLEAN FILTERS
        # ===========================================
        
        # Is Active
        is_active = self.params.get('is_active')
        if is_active:
            if is_active.lower() == 'true':
               
                queryset = queryset.filter(status__affects_headcount=True)
            elif is_active.lower() == 'false':
                print(f"❌ Applying is_active: False")
                queryset = queryset.filter(status__affects_headcount=False)
        
        # Org Chart Visible
        is_visible_in_org_chart = self.params.get('is_visible_in_org_chart')
        if is_visible_in_org_chart:
            visible = is_visible_in_org_chart.lower() == 'true'
      
            queryset = queryset.filter(is_visible_in_org_chart=visible)
        
        # Is Deleted (for admin purposes)
        is_deleted = self.params.get('is_deleted')
        if is_deleted:
            if is_deleted.lower() == 'true':
           
                from .models import Employee
                queryset = Employee.all_objects.filter(
                    pk__in=queryset.values_list('pk', flat=True),
                    is_deleted=True
                )
            elif is_deleted.lower() == 'false':
             
                queryset = queryset.filter(is_deleted=False)
            elif is_deleted.lower() == 'all':
             
                from .models import Employee
                queryset = Employee.all_objects.filter(
                    pk__in=queryset.values_list('pk', flat=True)
                )
        
        # ===========================================
        # 6. SPECIAL CALCULATED FILTERS
        # ===========================================
        
        # Status needs update (handled in view after filtering)
        status_needs_update = self.params.get('status_needs_update')
        if status_needs_update and status_needs_update.lower() == 'true':
        
            pass
        
        # Contract expiring soon
        contract_expiring_days = self.params.get('contract_expiring_days')
        if contract_expiring_days:
            try:
                days = int(contract_expiring_days)
                expiry_date = date.today() + timedelta(days=days)
          
                queryset = queryset.filter(
                    contract_end_date__lte=expiry_date,
                    contract_end_date__gte=date.today()
                )
            except:
                pass
        
        final_count = queryset.count()
       
        
        return queryset

class AdvancedEmployeeSorter:
    """
    MultipleSortingSystem.jsx component-inə uyğun sorting sistemi
    Frontend-dən gələn sorting array-ini düzgün işləyir
    """
    
    # Frontend component-də istifadə olunan sortable fields
    SORTABLE_FIELDS = {
        # Basic Information
        'name': 'full_name',
        'employee_name': 'full_name',
        'full_name': 'full_name',
        'first_name': 'user__first_name',
        'last_name': 'user__last_name',
        'employee_id': 'employee_id',
        'email': 'user__email',
        'phone': 'phone',
        'father_name': 'father_name',
        
        # Job Information
        'job_title': 'job_title',
        
        # Organizational Structure
        'business_function_name': 'business_function__name',
        'business_function_code': 'business_function__code',
        'department_name': 'department__name',
        'unit_name': 'unit__name',
        'job_function_name': 'job_function__name',
        
        # Position & Grading
        'position_group_name': 'position_group__name',
        'position_group_level': 'position_group__hierarchy_level',
        'grading_level': 'grading_level',
  
        
        # Management
        'line_manager_name': 'line_manager__full_name',
        'line_manager_hc_number': 'line_manager__employee_id',
        
        # Employment Dates
        'start_date': 'start_date',
        'end_date': 'end_date',
        'contract_start_date': 'contract_start_date',
        'contract_end_date': 'contract_end_date',
        
        # Contract Information
        'contract_duration': 'contract_duration',
        'contract_duration_display': 'contract_duration',
        
        # Status
        'status_name': 'status__name',
        'status_color': 'status__color',
        'current_status_display': 'status__name',
        
        # Personal Information
        'date_of_birth': 'date_of_birth',
        'gender': 'gender',
        
        # Calculated Fields (special handling)
        'years_of_service': 'start_date',  # Sort by start_date, reverse order
        'direct_reports_count': 'direct_reports_count',  # Need annotation
        
        # Metadata
        'created_at': 'created_at',
        'updated_at': 'updated_at',
        'is_visible_in_org_chart': 'is_visible_in_org_chart',
        'is_deleted': 'is_deleted',
    }
    
    # ✅ NEW: Date fields for special handling
    DATE_FIELDS = {
        'start_date', 'end_date', 'contract_start_date', 'contract_end_date',
        'date_of_birth', 'created_at', 'updated_at'
    }
    
    def __init__(self, queryset, sorting_params):
        self.queryset = queryset
        self.sorting_params = sorting_params or []
    
    def sort(self):
        """
        Frontend MultipleSortingSystem component-indən gələn sorting parametrlərini işləyir
        Format: [{'field': 'employee_name', 'direction': 'asc'}, ...]
        """
        
        if not self.sorting_params:
            # Default sorting
            return self.queryset.order_by('full_name')
        
        order_fields = []
        needs_annotation = False
        
        # Process each sorting parameter
        for sort_param in self.sorting_params:
            if isinstance(sort_param, dict):
                field_name = sort_param.get('field', '')
                direction = sort_param.get('direction', 'asc')
            else:
                # Fallback for string format like "-employee_name"
                if sort_param.startswith('-'):
                    field_name = sort_param[1:]
                    direction = 'desc'
                else:
                    field_name = sort_param
                    direction = 'asc'
            
            if not field_name or field_name not in self.SORTABLE_FIELDS:
                continue
            
            db_field = self.SORTABLE_FIELDS[field_name]
            
            # Special handling for calculated fields
            if field_name == 'years_of_service':
                # For years of service, reverse the direction since we sort by start_date
                direction = 'desc' if direction == 'asc' else 'asc'
            elif field_name == 'direct_reports_count':
                # Need to annotate with direct reports count
                needs_annotation = True
                db_field = 'direct_reports_count'
            
            # Apply direction
            if direction == 'desc':
                db_field = f'-{db_field}'
            
            order_fields.append(db_field)
        
        # Apply annotations if needed
        if needs_annotation:
            self.queryset = self.queryset.annotate(
                direct_reports_count=Count(
                    'direct_reports',
                    filter=Q(direct_reports__status__affects_headcount=True, direct_reports__is_deleted=False)
                )
            )
        
        if order_fields:
            # Add secondary sort for consistency
            if 'employee_id' not in [f.lstrip('-') for f in order_fields]:
                order_fields.append('employee_id')
            
            return self.queryset.order_by(*order_fields)
        
        return self.queryset.order_by('employee_id')
class BusinessFunctionViewSet(viewsets.ModelViewSet):
    queryset = BusinessFunction.objects.all().order_by('name')
    serializer_class = BusinessFunctionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'code']
    ordering = ['code']

class DepartmentViewSet(viewsets.ModelViewSet):
    """
    ENHANCED: Department ViewSet with bulk creation for multiple business functions
    """
    queryset = Department.objects.select_related('business_function').all().order_by('name')
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['business_function', 'is_active']
    search_fields = ['name']
    ordering = ['business_function__code']
    
    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['name', 'business_function_id'],
            properties={
                'name': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='Department name',
                    example='IT Department'
                ),
                'business_function_id': openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Schema(type=openapi.TYPE_INTEGER),
                    description='Business function ID(s) - Single integer or array of integers for bulk creation',
                    example=[1, 2, 3]
                ),
                'is_active': openapi.Schema(
                    type=openapi.TYPE_BOOLEAN,
                    default=True,
                    description='Whether this department is active'
                )
            }
        ),
        responses={
            201: openapi.Response(
                description="Department(s) created successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'department': openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            description='Single department (for single creation)'
                        ),
                        'created_departments': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(type=openapi.TYPE_OBJECT),
                            description='Multiple departments (for bulk creation)'
                        ),
                        'success_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'error_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'errors': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_STRING)),
                        'bulk_operation': openapi.Schema(type=openapi.TYPE_BOOLEAN)
                    }
                )
            ),
            400: openapi.Response(description="Bad request - validation errors")
        }
    )
    def create(self, request, *args, **kwargs):
        """
        FIXED: Enhanced create to support both single and bulk creation
        """
        try:
            # CRITICAL FIX: DON'T transform the data here, let serializer handle it
            data = request.data.copy()
            
            # Just pass the data as-is to the serializer
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            
            # Check if this is a bulk operation
            self.perform_create(serializer)
            
            # Get bulk result from serializer context
            bulk_result = serializer.context.get('bulk_result', {})
            
            if bulk_result:
                # Bulk creation
                created_departments_data = DepartmentSerializer(
                    bulk_result.get('created_departments', []),
                    many=True,
                    context={'request': request}
                ).data
                
                
                
                return Response({
                    'success': True,
                    'message': f"Created {bulk_result['success_count']} departments, {bulk_result['error_count']} failed",
                    'created_departments': created_departments_data,
                    'success_count': bulk_result['success_count'],
                    'error_count': bulk_result['error_count'],
                    'errors': bulk_result.get('errors', []),
                    'bulk_operation': True
                }, status=status.HTTP_201_CREATED if bulk_result['success_count'] > 0 else status.HTTP_400_BAD_REQUEST)
            
            # Single creation
            headers = self.get_success_headers(serializer.data)
            
            return Response({
                'success': True,
                'message': 'Department created successfully',
                'department': serializer.data,
                'bulk_operation': False
            }, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            logger.error(f"Department creation failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to create department(s): {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class UnitViewSet(viewsets.ModelViewSet):
    """
    ENHANCED: Unit ViewSet with bulk creation for multiple departments
    """
    queryset = Unit.objects.select_related('department__business_function').all().order_by('name')
    serializer_class = UnitSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['department', 'is_active']
    search_fields = ['name']
    ordering = ['department__business_function__code']
    
    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['name', 'department_id'],
            properties={
                'name': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='Unit name',
                    example='Backend Team'
                ),
                'department_id': openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Schema(type=openapi.TYPE_INTEGER),
                    description='Department ID(s) - Single integer or array of integers for bulk creation',
                    example=[1, 2, 3]
                ),
                'is_active': openapi.Schema(
                    type=openapi.TYPE_BOOLEAN,
                    default=True,
                    description='Whether this unit is active'
                )
            }
        ),
        responses={
            201: openapi.Response(
                description="Unit(s) created successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'unit': openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            description='Single unit (for single creation)'
                        ),
                        'created_units': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(type=openapi.TYPE_OBJECT),
                            description='Multiple units (for bulk creation)'
                        ),
                        'success_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'error_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'errors': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_STRING)),
                        'bulk_operation': openapi.Schema(type=openapi.TYPE_BOOLEAN)
                    }
                )
            ),
            400: openapi.Response(description="Bad request - validation errors")
        }
    )
    def create(self, request, *args, **kwargs):
        """
        FIXED: Enhanced create to support both single and bulk creation
        """
        try:
            # CRITICAL FIX: DON'T transform the data here, let serializer handle it
            data = request.data.copy()
            
            # Just pass the data as-is to the serializer
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            
            # Check if this is a bulk operation by checking if serializer has bulk_result in context after save
            self.perform_create(serializer)
            
            # Get bulk result from serializer context
            bulk_result = serializer.context.get('bulk_result', {})
            
            if bulk_result:
                # Bulk creation
                created_units_data = UnitSerializer(
                    bulk_result.get('created_units', []),
                    many=True,
                    context={'request': request}
                ).data
                
               
                
                return Response({
                    'success': True,
                    'message': f"Created {bulk_result['success_count']} units, {bulk_result['error_count']} failed",
                    'created_units': created_units_data,
                    'success_count': bulk_result['success_count'],
                    'error_count': bulk_result['error_count'],
                    'errors': bulk_result.get('errors', []),
                    'bulk_operation': True
                }, status=status.HTTP_201_CREATED if bulk_result['success_count'] > 0 else status.HTTP_400_BAD_REQUEST)
            
            # Single creation
            headers = self.get_success_headers(serializer.data)
            
            return Response({
                'success': True,
                'message': 'Unit created successfully',
                'unit': serializer.data,
                'bulk_operation': False
            }, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            logger.error(f"Unit creation failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to create unit(s): {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class JobTitleViewSet(viewsets.ModelViewSet):

    queryset = JobTitle.objects.all().order_by('name')
    serializer_class = JobTitleSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = ModernPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [ 'is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['name']
    
    @swagger_auto_schema(
        operation_description="""
        Create a new job title.
        
        Job titles must be linked to a job function and should be unique.
        """,
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['name'],
            properties={
                'name': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='Job title name (must be unique)',
                    example='Senior Software Engineer'
                ),
                'description': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='Description of the job title',
                    example='Senior level software engineering position with 5+ years experience'
                ),
               
                'is_active': openapi.Schema(
                    type=openapi.TYPE_BOOLEAN,
                    default=True,
                    description='Whether this job title is active'
                )
            }
        ),
        responses={
            201: openapi.Response(
                description="Job title created successfully",
                schema=JobTitleSerializer
            ),
            400: "Bad request - validation errors"
        }
    )
    def create(self, request, *args, **kwargs):
        """Create a new job title"""
        try:
            
            
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            
            
            
            headers = self.get_success_headers(serializer.data)
            return Response({
                'success': True,
                'message': f"Job title '{serializer.data['name']}' created successfully",
                'job_title': serializer.data
            }, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            logger.error(f"Job title creation failed: {str(e)}")
            return Response(
                {'error': f'Failed to create job title: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        operation_description="""
        Update an existing job title.
        
        You can update the name, description or active status.
        """,
        request_body=JobTitleSerializer,
        responses={
            200: openapi.Response(
                description="Job title updated successfully",
                schema=JobTitleSerializer
            ),
            400: "Bad request - validation errors",
            404: "Job title not found"
        }
    )
    def update(self, request, *args, **kwargs):
        """Update a job title"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            old_name = instance.name
            
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            
            
            return Response({
                'success': True,
                'message': f"Job title updated successfully",
                'job_title': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Job title update failed: {str(e)}")
            return Response(
                {'error': f'Failed to update job title: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        operation_description="""
        Delete a job title (soft delete).
        
        Note: Job titles that are currently in use by employees cannot be deleted.
        """,
        responses={
            204: "Job title deleted successfully",
            400: "Cannot delete - job title is in use",
            404: "Job title not found"
        }
    )
    def destroy(self, request, *args, **kwargs):
        """Soft delete a job title"""
        try:
            instance = self.get_object()
            
            # Check if any employees are using this job title
            employees_using = Employee.objects.filter(
                job_title=instance.name,
                is_deleted=False
            ).count()
            
            if employees_using > 0:
                return Response({
                    'error': f'Cannot delete job title. {employees_using} employee(s) are currently using this job title.',
                    'employees_count': employees_using
                }, status=status.HTTP_400_BAD_REQUEST)
            
            job_title_name = instance.name
            instance.soft_delete(user=request.user)
            
            
            
            return Response({
                'success': True,
                'message': f"Job title '{job_title_name}' deleted successfully"
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            logger.error(f"Job title deletion failed: {str(e)}")
            return Response(
                {'error': f'Failed to delete job title: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class JobFunctionViewSet(viewsets.ModelViewSet):
    """UPDATED: Employee count əlavə olundu"""
    queryset = JobFunction.objects.all().order_by('name')
    serializer_class = JobFunctionSerializer
    pagination_class = ModernPagination
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active']
    search_fields = ['name']
    ordering = ['name']

class PositionGroupViewSet(viewsets.ModelViewSet):
    queryset = PositionGroup.objects.all().order_by('hierarchy_level')  # Bu yol
    serializer_class = PositionGroupSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]  # OrderingFilter əlavə et
    filterset_fields = ['is_active']
    search_fields = ['name']
    ordering_fields = ['hierarchy_level', 'name']  # İstifadəçi hansı fieldlara görə sort edə bilər
    ordering = ['hierarchy_level']  # Default ordering
    
    @action(detail=True, methods=['get'])
    def grading_levels(self, request, pk=None):
        """Get available grading levels for this position group"""
        position_group = self.get_object()
        levels = position_group.get_grading_levels()
        return Response({
            'position_group': position_group.get_name_display(),
            'shorthand': position_group.grading_shorthand,
            'levels': levels
        })

class EmployeeTagViewSet(viewsets.ModelViewSet):
    queryset = EmployeeTag.objects.all().order_by('name')
    serializer_class = EmployeeTagSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [ 'is_active']
    search_fields = ['name']
    ordering = [ 'name']

class EmployeeStatusViewSet(viewsets.ModelViewSet):
    queryset = EmployeeStatus.objects.all()
    serializer_class = EmployeeStatusSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status_type', 'affects_headcount', 'allows_org_chart', 'is_active']
    search_fields = ['name']
    ordering = ['order', 'name']

class ContractTypeConfigViewSet(viewsets.ModelViewSet):
    queryset = ContractTypeConfig.objects.all()
    serializer_class = ContractTypeConfigSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['contract_type', 'enable_auto_transitions', 'is_active']
    search_fields = ['contract_type', 'display_name']
    ordering = ['contract_type']
    

class VacantPositionViewSet(viewsets.ModelViewSet):
    """FIXED: Vacant Position ViewSet with proper field validation"""
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'business_function', 'department', 'position_group', 'is_filled', 'include_in_headcount'
    ]
    search_fields = ['job_title', 'position_id']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = VacantPosition.objects.select_related(
            'business_function', 'department', 'unit', 'job_function',
            'position_group', 'reporting_to', 'filled_by_employee', 'created_by', 'vacancy_status'
        ).all()
        
        # Default olaraq filled olanları gizlət
        show_filled = self.request.query_params.get('show_filled', 'false').lower() == 'true'
        if not show_filled:
            queryset = queryset.filter(is_filled=False)
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return VacantPositionListSerializer
        elif self.action in ['create', 'update', 'partial_update']:  # UPDATE ƏLAVƏ ET
            return VacantPositionCreateSerializer  # Eyni serializer istifadə et
        elif self.action == 'convert_to_employee':
            return VacancyToEmployeeConversionSerializer
        else:
            return VacantPositionDetailSerializer
    
    @swagger_auto_schema(
    method='post',
    operation_description="Convert vacancy to employee with required fields only",
    # ƏLAVƏ ET:
    request_body=VacancyToEmployeeConversionSerializer,  # Dəqiq serializer göstər
    manual_parameters=[
        # Yalnız file field-lər burada olsun
        openapi.Parameter(
            'document',
            openapi.IN_FORM,
            description="Employee document file",
            type=openapi.TYPE_FILE,
            required=False
        ),
        openapi.Parameter(
            'profile_photo',
            openapi.IN_FORM,
            description="Profile photo",
            type=openapi.TYPE_FILE,
            required=False
        ),
    ],
    consumes=['multipart/form-data'],
    responses={
        201: openapi.Response(
            description="Employee created successfully from vacancy",
            schema=EmployeeDetailSerializer
        ),
        400: openapi.Response(description="Bad request - validation errors"),
        404: openapi.Response(description="Vacancy not found")
    }
)
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def convert_to_employee(self, request, pk=None):
        """Convert vacancy to employee with all required and optional fields"""
        vacancy = self.get_object()
        
        if vacancy.is_filled:
            return Response(
                {'error': 'Vacancy is already filled'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check required fields
        required_fields = ['first_name', 'last_name', 'email', 'start_date', 'contract_duration']
        missing_fields = []
        
        for field in required_fields:
            if not request.data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return Response({
                'error': 'Missing required fields',
                'missing_fields': missing_fields,
                'message': f'The following fields are required: {", ".join(missing_fields)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate email uniqueness
        email = request.data.get('email')
        if User.objects.filter(email=email).exists():
            return Response({
                'error': 'Email already exists',
                'email': email,
                'message': f'An account with email {email} already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Prepare data for serializer
        data = request.data.copy()
        
        # Handle tag_ids if provided as comma-separated string or array
        if 'tag_ids' in data:
            if isinstance(data['tag_ids'], str):
                try:
                    tag_ids = [int(id.strip()) for id in data['tag_ids'].split(',') if id.strip()]
                    data['tag_ids'] = tag_ids
                except ValueError:
                    return Response({
                        'error': 'Invalid tag_ids format',
                        'message': 'Use comma-separated integers (e.g., "1,2,3")'
                    }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create serializer with context that includes vacancy
        serializer = VacancyToEmployeeConversionSerializer(
            data=data, 
            context={'request': request, 'vacancy': vacancy}  # Pass vacancy in context
        )
        
        if not serializer.is_valid():
            return Response({
                'error': 'Validation failed',
                'details': serializer.errors,
                'message': 'Please check the provided data and try again'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            employee = serializer.save()
            
            return Response({
                'success': True,
                'message': f'Vacancy {vacancy.position_id} successfully converted to employee with ID {employee.employee_id}',
                'employee': EmployeeDetailSerializer(employee, context={'request': request}).data,
                'generated_employee_id': employee.employee_id,
                'conversion_details': {
                    'vacancy_position_id': vacancy.position_id,
                    'employee_auto_id': employee.employee_id,
                    'business_function': vacancy.business_function.code if vacancy.business_function else None,
                    'files_uploaded': {
                        'has_document': bool(request.FILES.get('document')),
                        'has_profile_photo': bool(request.FILES.get('profile_photo'))
                    }
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error converting vacancy to employee: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response({
                'error': 'Failed to convert vacancy to employee',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmployeeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = ModernPagination  # Use modern pagination
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_queryset(self):
        """✅ UPDATED: Filter based on user access"""
        from .models import Employee
        
        base_queryset = Employee.objects.select_related(
            'user', 'business_function', 'department', 'unit', 'job_function',
            'position_group', 'status', 'line_manager', 'original_vacancy'
        ).prefetch_related(
            'tags', 'documents', 'activities'
        ).all().order_by('full_name')
        
        # Apply access control
        return filter_headcount_queryset(self.request.user, base_queryset)    

    def _clean_form_data(self, data):
        """Comprehensive data cleaning for form data"""
        cleaned_data = {}
        
        # Handle each field with appropriate conversion
        for key, value in data.items():
            # Skip files and empty values
            if hasattr(value, 'read') or value in [None, '']:
                cleaned_data[key] = value
                continue
            
            # Convert list values (like from getlist())
            if isinstance(value, list):
                if len(value) == 1:
                    value = value[0]
                elif len(value) == 0:
                    cleaned_data[key] = None
                    continue
            
            # Boolean fields
            if key in ['is_visible_in_org_chart']:
                if isinstance(value, str):
                    cleaned_data[key] = value.lower() in ['true', '1', 'yes', 'on']
                else:
                    cleaned_data[key] = bool(value)
            
            # Integer fields (foreign keys and IDs)
            elif key in ['business_function', 'department', 'unit', 'job_function', 
                         'position_group', 'line_manager', 'vacancy_id', 'original_employee_pk']:
                try:
                    cleaned_data[key] = int(value) if value else None
                except (ValueError, TypeError):
                    cleaned_data[key] = None
            
            # Date fields
            elif key in ['date_of_birth', 'start_date', 'end_date', 'contract_start_date']:
                if isinstance(value, str) and value.strip():
                    # Validate date format
                    try:
                        from datetime import datetime
                        datetime.strptime(value.strip(), '%Y-%m-%d')
                        cleaned_data[key] = value.strip()
                    except ValueError:
                        # Invalid date format, skip this field
                        continue
                else:
                    cleaned_data[key] = value
            
            # Choice fields that need string values
            elif key in ['gender', 'contract_duration', 'document_type']:
                if isinstance(value, list):
                    cleaned_data[key] = value[0] if value else None
                else:
                    cleaned_data[key] = str(value).strip() if value else None
            
            # Array fields (tag_ids)
            elif key == 'tag_ids':
                if isinstance(value, str):
                    # Convert comma-separated string to list of integers
                    try:
                        cleaned_data[key] = [int(id.strip()) for id in value.split(',') if id.strip()]
                    except ValueError:
                        cleaned_data[key] = []
                elif isinstance(value, list):
                    try:
                        cleaned_data[key] = [int(id) for id in value if id]
                    except ValueError:
                        cleaned_data[key] = []
                else:
                    cleaned_data[key] = []
            
            # String fields
            else:
                if isinstance(value, list):
                    cleaned_data[key] = value[0] if value else ''
                else:
                    cleaned_data[key] = str(value).strip() if value else ''
        
        return cleaned_data
    
    def get_serializer_class(self):
        from .serializers import (
            EmployeeListSerializer, EmployeeDetailSerializer, EmployeeCreateUpdateSerializer
        )
        if self.action == 'list':
            return EmployeeListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return EmployeeCreateUpdateSerializer
        else:
            return EmployeeDetailSerializer
    
    def list(self, request, *args, **kwargs):
        """✅ UPDATED: Add access info to response with full filtering and pagination"""
        access = get_headcount_access(request.user)
        
        # ✅ NEW: Check if user is viewing their own profile
        current_user_employee = None
        try:
            current_user_employee = Employee.objects.get(user=request.user)
        except Employee.DoesNotExist:
            pass
        
        # Regular employee - REDIRECT to own profile if exists
        if not access['is_manager'] and not access['can_view_all']:
            if current_user_employee:
                
                
                serializer = self.get_serializer(current_user_employee)
                
                return Response({
                    'count': 1,
                    'pagination_used': False,
                    'results': [serializer.data],
                    'access_info': {
                        'can_view_all': False,
                        'is_manager': False,
                        'viewing_own_profile': True,
                        'accessible_count': 1
                    },
                    'summary': {
                        'total_records': 1,
                        'employee_records': 1,
                        'vacancy_records': 0,
                        'includes_vacancies': False,
                        'unified_view': False,
                        'own_profile_only': True
                    }
                })
            else:
                # User has no employee profile and no access
                return Response({
                    'error': 'Access Denied',
                    'message': 'You do not have permission to view headcount data.',
                    'detail': 'Only managers and administrators can access the headcount table.',
                    'can_view_all': False,
                    'is_manager': False,
                    'has_employee_profile': False,
                    'suggestion': 'Contact HR to create your employee profile.'
                }, status=status.HTTP_403_FORBIDDEN)
            
        try:
            include_vacancies = request.query_params.get('include_vacancies', 'true').lower() == 'true'
            
            # ✅ FIX: Pagination parametrlərini yoxla
            page_param = request.query_params.get('page')
            page_size_param = request.query_params.get('page_size')
            use_pagination = request.query_params.get('use_pagination', '').lower() == 'true'
            
            # ✅ CRITICAL: Filter parametrlərini detect et
            filter_params = [
                'search', 'employee_search', 'business_function', 'department', 
                'unit', 'job_function', 'position_group', 'status', 'grading_level',
                'line_manager', 'tags', 'gender', 'start_date_from', 'start_date_to',
                'contract_end_date_from', 'contract_end_date_to', 'is_active', 
                'status_needs_update', 'job_title_search', 'contract_duration',
                'is_visible_in_org_chart', 'contract_expiring_days',
                'years_of_service_min', 'years_of_service_max'
            ]
            
            has_filters = any(request.query_params.get(param) for param in filter_params)
            
            # ✅ NEW: Log active filters for debugging
            if has_filters:
                active_filters = {k: request.query_params.get(k) for k in filter_params if request.query_params.get(k)}
                
            
            # ✅ Manager can only select their accessible business functions
            if not access['can_view_all'] and access['accessible_business_functions']:
                bf_filter = request.query_params.get('business_function')
                if bf_filter:
                    bf_ids = [int(id) for id in bf_filter.split(',')]
                    # Check if requested BFs are accessible
                    invalid_bfs = set(bf_ids) - set(access['accessible_business_functions'])
                    if invalid_bfs:
                        return Response({
                            'error': 'Access Denied',
                            'message': 'You cannot access the selected company(ies)',
                            'invalid_business_functions': list(invalid_bfs)
                        }, status=status.HTTP_403_FORBIDDEN)
            
            should_paginate = bool(page_param or page_size_param or use_pagination)
            
            if include_vacancies:
                response = self._get_unified_employee_vacancy_list(request, should_paginate)
            else:
                response = self._get_employee_only_list(request, should_paginate)
            
            # ✅ Add access info to response
            if isinstance(response.data, dict):
                response.data['access_info'] = {
                    'can_view_all': access['can_view_all'],
                    'is_manager': access['is_manager'],
                    'accessible_count': len(access['accessible_employee_ids']) if access['accessible_employee_ids'] else 'all',
                    'accessible_business_functions': access['accessible_business_functions'] if not access['can_view_all'] else None
                }
            
            return response
                
        except Exception as e:
            logger.error(f"Error in employee list view: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve employees: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='post',
        operation_description="Upload document for specific employee by employee_id",
        manual_parameters=[
            openapi.Parameter(
                'employee_id',
                openapi.IN_FORM,
                description='Employee ID (e.g., GEO1, HC001)',
                type=openapi.TYPE_STRING,
                required=True
            ),
            openapi.Parameter(
                'document',
                openapi.IN_FORM,
                description='Document file to upload',
                type=openapi.TYPE_FILE,
                required=True
            ),
            openapi.Parameter(
                'document_name',
                openapi.IN_FORM,
                description='Document name/title',
                type=openapi.TYPE_STRING,
                required=True
            ),
            openapi.Parameter(
                'document_type',
                openapi.IN_FORM,
                description='Document type',
                type=openapi.TYPE_STRING,
                enum=['CONTRACT', 'ID', 'CERTIFICATE', 'CV', 'PERFORMANCE', 'MEDICAL', 'TRAINING', 'OTHER'],
                required=True
            ),
            openapi.Parameter(
                'description',
                openapi.IN_FORM,
                description='Document description (optional)',
                type=openapi.TYPE_STRING,
                required=False
            ),
            openapi.Parameter(
                'expiry_date',
                openapi.IN_FORM,
                description='Document expiry date (YYYY-MM-DD) (optional)',
                type=openapi.TYPE_STRING,
                required=False
            ),
            openapi.Parameter(
                'is_confidential',
                openapi.IN_FORM,
                description='Mark as confidential',
                type=openapi.TYPE_BOOLEAN,
                required=False
            ),
        ],
        consumes=['multipart/form-data'],
        responses={
            201: openapi.Response(
                description="Document uploaded successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'document': openapi.Schema(type=openapi.TYPE_OBJECT),
                    }
                )
            ),
            400: "Bad request",
            404: "Employee not found"
        }
    )
    @action(detail=False, methods=['post'], url_path='upload-document', parser_classes=[MultiPartParser, FormParser])
    def upload_employee_document(self, request):
        """✅ Upload document for employee by employee_id"""
        try:
            # Get parameters
            employee_id = request.data.get('employee_id')
            document_file = request.FILES.get('document')
            document_name = request.data.get('document_name')
            document_type = request.data.get('document_type', 'OTHER')
            description = request.data.get('description', '')
            expiry_date = request.data.get('expiry_date')
            is_confidential = request.data.get('is_confidential', 'false').lower() == 'true'
            
            # Validation
            if not employee_id:
                return Response(
                    {'error': 'employee_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not document_file:
                return Response(
                    {'error': 'document file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not document_name:
                return Response(
                    {'error': 'document_name is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find employee by employee_id
            try:
                employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                return Response(
                    {'error': f'Employee with ID {employee_id} not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # ✅ Check access permissions
            access = get_headcount_access(request.user)
            
            can_upload = False
            if access['can_view_all']:
                can_upload = True
            elif employee.user and employee.user.id == request.user.id:
                can_upload = True
            elif access['is_manager'] and access['accessible_employee_ids'] and employee.id in access['accessible_employee_ids']:
                can_upload = True
            
            if not can_upload:
                return Response(
                    {
                        'error': 'Access Denied',
                        'message': 'You do not have permission to upload documents for this employee'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Parse expiry date if provided
            expiry_date_obj = None
            if expiry_date:
                try:
                    from datetime import datetime
                    expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                except ValueError:
                    return Response(
                        {'error': 'Invalid expiry_date format. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Validate file size (10MB max)
            if document_file.size > 10 * 1024 * 1024:
                return Response(
                    {'error': 'File size exceeds 10MB limit'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate document type
            valid_types = ['CONTRACT', 'ID', 'CERTIFICATE', 'CV', 'PERFORMANCE', 'MEDICAL', 'TRAINING', 'OTHER']
            if document_type not in valid_types:
                return Response(
                    {'error': f'Invalid document_type. Must be one of: {", ".join(valid_types)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create document
            document = EmployeeDocument.objects.create(
                employee=employee,
                name=document_name,
                document_type=document_type,
                document_file=document_file,
                description=description,
                expiry_date=expiry_date_obj,
                is_confidential=is_confidential,
                uploaded_by=request.user,
                document_status='ACTIVE',
                version=1,
                is_current_version=True
            )
            
            # Log activity
            EmployeeActivity.objects.create(
                employee=employee,
                activity_type='DOCUMENT_UPLOADED',
                description=f"Document '{document_name}' ({document_type}) uploaded",
                performed_by=request.user,
                metadata={
                    'document_id': document.id,
                    'document_name': document_name,
                    'document_type': document_type,
                    'file_size': document.file_size,
                    'is_confidential': is_confidential,
                    'uploaded_via': 'direct_api'
                }
            )
            
            logger.info(f"✅ Document uploaded for employee {employee_id} by {request.user.username}")
            
            # Serialize response
            serializer = EmployeeDocumentSerializer(document, context={'request': request})
            
            return Response({
                'success': True,
                'message': f'Document uploaded successfully for employee {employee.full_name}',
                'document': serializer.data,
                'employee': {
                    'id': employee.id,
                    'employee_id': employee.employee_id,
                    'name': employee.full_name
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Document upload failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Document upload failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='delete',
        operation_description="Delete employee document by document ID",
        manual_parameters=[
            openapi.Parameter(
                'document_id',
                openapi.IN_QUERY,
                description='Document ID to delete',
                type=openapi.TYPE_INTEGER,
                required=True
            ),
        ],
        responses={
            200: openapi.Response(
                description="Document deleted successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                    }
                )
            ),
            404: "Document not found"
        }
    )
    @action(detail=False, methods=['delete'], url_path='delete-document')
    def delete_employee_document(self, request):
        """✅ Delete employee document"""
        try:
            document_id = request.query_params.get('document_id')
            
            if not document_id:
                return Response(
                    {'error': 'document_id parameter is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find document
            try:
                document = EmployeeDocument.objects.get(id=document_id, is_deleted=False)
            except EmployeeDocument.DoesNotExist:
                return Response(
                    {'error': f'Document with ID {document_id} not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            employee = document.employee
            
            # ✅ Check access permissions
            access = get_headcount_access(request.user)
            
            can_delete = False
            if access['can_view_all']:
                can_delete = True
            elif employee.user and employee.user.id == request.user.id:
                can_delete = True
            elif access['is_manager'] and access['accessible_employee_ids'] and employee.id in access['accessible_employee_ids']:
                can_delete = True
            
            if not can_delete:
                return Response(
                    {
                        'error': 'Access Denied',
                        'message': 'You do not have permission to delete this document'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Store info before deletion
            document_info = {
                'id': document.id,
                'name': document.name,
                'type': document.document_type,
                'employee_id': employee.employee_id,
                'employee_name': employee.full_name
            }
            
            # Delete document file from storage
            if document.document_file:
                try:
                    if hasattr(document.document_file, 'path'):
                        import os
                        file_path = document.document_file.path
                        if os.path.exists(file_path):
                            os.remove(file_path)
                            logger.info(f"✅ Document file deleted: {file_path}")
                except Exception as e:
                    logger.warning(f"Could not delete document file: {e}")
            
            # Soft delete document
            document.soft_delete(user=request.user)
            
            # Log activity
            EmployeeActivity.objects.create(
                employee=employee,
                activity_type='UPDATED',
                description=f"Document '{document_info['name']}' ({document_info['type']}) deleted",
                performed_by=request.user,
                metadata={
                    'document_id': document_info['id'],
                    'document_name': document_info['name'],
                    'document_type': document_info['type'],
                    'deletion_method': 'direct_api'
                }
            )
            
            logger.info(f"✅ Document {document_id} deleted for employee {employee.employee_id}")
            
            return Response({
                'success': True,
                'message': f"Document '{document_info['name']}' deleted successfully",
                'deleted_document': document_info
            })
            
        except Exception as e:
            logger.error(f"Document deletion failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Document deletion failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='get',
        operation_description="Get all documents for employee by employee_id",
        manual_parameters=[
            openapi.Parameter(
                'employee_id',
                openapi.IN_QUERY,
                description='Employee ID (e.g., GEO1, HC001)',
                type=openapi.TYPE_STRING,
                required=True
            ),
        ],
        responses={
            200: openapi.Response(
                description="Documents retrieved successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'employee': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'documents': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                        'total_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                    }
                )
            ),
            404: "Employee not found"
        }
    )
    @action(detail=False, methods=['get'], url_path='get-documents')
    def get_employee_documents(self, request):
        """✅ Get all documents for employee"""
        try:
            employee_id = request.query_params.get('employee_id')
            
            if not employee_id:
                return Response(
                    {'error': 'employee_id parameter is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find employee
            try:
                employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                return Response(
                    {'error': f'Employee with ID {employee_id} not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # ✅ Check access permissions
            access = get_headcount_access(request.user)
            
            can_view = False
            if access['can_view_all']:
                can_view = True
            elif employee.user and employee.user.id == request.user.id:
                can_view = True
            elif access['is_manager'] and access['accessible_employee_ids'] and employee.id in access['accessible_employee_ids']:
                can_view = True
            
            if not can_view:
                return Response(
                    {
                        'error': 'Access Denied',
                        'message': 'You do not have permission to view this employee\'s documents'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Get documents
            documents = EmployeeDocument.objects.filter(
                employee=employee,
                is_deleted=False
            ).order_by('-uploaded_at')
            
            # Serialize
            serializer = EmployeeDocumentSerializer(documents, many=True, context={'request': request})
            
            # Summary by type
            summary_by_type = {}
            for doc in documents:
                doc_type = doc.get_document_type_display()
                summary_by_type[doc_type] = summary_by_type.get(doc_type, 0) + 1
            
            return Response({
                'success': True,
                'employee': {
                    'id': employee.id,
                    'employee_id': employee.employee_id,
                    'name': employee.full_name,
                    'job_title': employee.job_title
                },
                'documents': serializer.data,
                'total_count': documents.count(),
                'summary_by_type': summary_by_type
            })
            
        except Exception as e:
            logger.error(f"Get documents failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve documents: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        operation_description="Get detailed information for a specific employee",
        responses={
            200: openapi.Response(
                description="Employee details",
                schema=EmployeeDetailSerializer
            ),
            403: "Access denied - not allowed to view this employee",
            404: "Employee not found"
        }
    )
    def retrieve(self, request, *args, **kwargs):
        """
        ✅ UPDATED: Allow users to view their own profile even without headcount access
        """
        try:
            employee_id = kwargs.get('pk')
            
            # Try to get employee
            try:
                employee = Employee.objects.select_related(
                    'user', 'business_function', 'department', 'unit', 'job_function',
                    'position_group', 'status', 'line_manager', 'original_vacancy'
                ).prefetch_related(
                    'tags', 'documents', 'activities'
                ).get(pk=employee_id)
            except Employee.DoesNotExist:
                return Response(
                    {'detail': 'Employee not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Check access
            access = get_headcount_access(request.user)
            
            # ✅ CRITICAL: Allow users to view their own profile
            is_own_profile = employee.user and employee.user.id == request.user.id
            
            if is_own_profile:
                # User viewing their own profile - always allow
                logger.info(f"✅ User {request.user.username} viewing own profile: {employee.employee_id}")
            elif access['can_view_all']:
                # Admin can view all
                logger.info(f"✅ Admin {request.user.username} viewing employee: {employee.employee_id}")
            elif access['is_manager'] and access['accessible_employee_ids']:
                # Manager can only view accessible employees
                if employee.id not in access['accessible_employee_ids']:
                    logger.warning(f"⚠️ Manager {request.user.username} attempted to view unauthorized employee: {employee.employee_id}")
                    return Response(
                        {
                            'error': 'Access Denied',
                            'message': 'You do not have permission to view this employee.',
                            'detail': 'You can only view employees in your team or company.',
                        },
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                # Regular employee without access
                logger.warning(f"⚠️ User {request.user.username} attempted to view employee without access: {employee.employee_id}")
                return Response(
                    {
                        'error': 'Access Denied',
                        'message': 'You do not have permission to view employee profiles.',
                        'detail': 'Only managers and administrators can view employee profiles.',
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Serialize and return
            serializer = self.get_serializer(employee)
            
            # Add context information
            response_data = serializer.data
            response_data['access_context'] = {
                'is_own_profile': is_own_profile,
                'can_edit': is_own_profile or access['can_view_all'],
                'can_view_team': access['is_manager'] or access['can_view_all'],
                'viewer_is_admin': access['can_view_all']
            }
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"❌ Error retrieving employee: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve employee: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @swagger_auto_schema(
        method='get',
        operation_description="Get current logged-in user's employee profile",
        responses={
            200: openapi.Response(
                description="Current user's employee profile",
                schema=EmployeeDetailSerializer
            ),
            404: "Employee profile not found for current user"
        }
    )
    @action(detail=False, methods=['get'])
    def get_my_profile(self, request):
        """
        ✅ NEW: Get current user's own employee profile
        Accessible to ALL authenticated users regardless of headcount access
        """
        try:
            # Get current user's employee profile
            try:
                employee = Employee.objects.select_related(
                    'user', 'business_function', 'department', 'unit', 'job_function',
                    'position_group', 'status', 'line_manager', 'original_vacancy'
                ).prefetch_related(
                    'tags', 'documents', 'activities'
                ).get(user=request.user)
                
             
                
            except Employee.DoesNotExist:
                logger.warning(f"⚠️ No employee profile found for user: {request.user.username}")
                return Response(
                    {
                        'error': 'Employee profile not found',
                        'message': 'No employee record is associated with your user account.',
                        'user_email': request.user.email,
                        'contact_hr': 'Please contact HR to create your employee profile.'
                    },
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Serialize employee data
            serializer = EmployeeDetailSerializer(employee, context={'request': request})
            
            # Get additional context
            direct_reports_count = employee.direct_reports.filter(
                status__affects_headcount=True,
                is_deleted=False
            ).count()
            
            # Get team members if user is a manager
            team_members = []
            if direct_reports_count > 0:
                team = employee.direct_reports.filter(
                    status__affects_headcount=True,
                    is_deleted=False
                ).select_related('position_group', 'department')[:10]
                
                team_members = [{
                    'id': member.id,
                    'employee_id': member.employee_id,
                    'name': member.full_name,
                    'job_title': member.job_title,
                    'department': member.department.name if member.department else None
                } for member in team]
            
            # Get manager chain
            manager_chain = []
            current_manager = employee.line_manager
            visited = set()
            
            while current_manager and current_manager.id not in visited and len(manager_chain) < 5:
                visited.add(current_manager.id)
                manager_chain.append({
                    'id': current_manager.id,
                    'employee_id': current_manager.employee_id,
                    'name': current_manager.full_name,
                    'job_title': current_manager.job_title
                })
                current_manager = current_manager.line_manager
            
            return Response({
                'employee': serializer.data,
                'profile_context': {
                    'is_manager': direct_reports_count > 0,
                    'direct_reports_count': direct_reports_count,
                    'team_members': team_members,
                    'manager_chain': manager_chain,
                    'levels_to_top': len(manager_chain)
                },
                'access_info': {
                    'can_view_headcount': False,  # Will be updated by middleware if needed
                    'can_edit_own_profile': True,
              
                }
            })
            
        except Exception as e:
            logger.error(f"❌ Error getting my profile: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve profile: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _get_employee_only_list(self, request, should_paginate):
        """FIXED: Employee-only list logic WITH PROPER SORTING"""
        
        # ✅ 1. GET BASE QUERYSET
        queryset = self.get_queryset()
        
        # ✅ 2. APPLY FILTERS FIRST (BEFORE SORTING!)
        employee_filter = ComprehensiveEmployeeFilter(queryset, request.query_params)
        queryset = employee_filter.filter()
        
     
        
        # ✅ 3. APPLY SORTING TO FILTERED QUERYSET
        sorting_data = request.query_params.get('sorting')
        if sorting_data:
            try:
                import json
                sorting_params = json.loads(sorting_data)
            except:
                ordering = request.query_params.get('ordering', '')
                sort_params = [param.strip() for param in ordering.split(',') if param.strip()]
                sorting_params = []
                for param in sort_params:
                    if param.startswith('-'):
                        sorting_params.append({'field': param[1:], 'direction': 'desc'})
                    else:
                        sorting_params.append({'field': param, 'direction': 'asc'})
        else:
            sorting_params = []
        
        # ✅ CRITICAL: Apply sorting to FILTERED queryset
        if not sorting_params:
            queryset = queryset.order_by('full_name')
        else:
            employee_sorter = AdvancedEmployeeSorter(queryset, sorting_params)
            queryset = employee_sorter.sort()
         
        
        total_count = queryset.count()

        
        if not should_paginate:
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                'count': total_count,
                'pagination_used': False,
                'results': serializer.data,
                'summary': {
                    'total_records': total_count,
                    'employee_records': total_count,
                    'vacancy_records': 0,
                    'includes_vacancies': False,
                    'unified_view': False
                }
            })
        else:
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                paginated_response = self.get_paginated_response(serializer.data)
                paginated_response.data['summary'] = {
                    'total_records': total_count,
                    'employee_records': total_count,
                    'vacancy_records': 0,
                    'includes_vacancies': False,
                    'unified_view': False
                }
                return paginated_response
    
    def _get_unified_employee_vacancy_list(self, request, should_paginate):
        """COMPLETELY FIXED: Get unified list with proper vacancy access control"""
        
        # ✅ GET ACCESS INFO FIRST
        access = get_headcount_access(request.user)
        
        # Parse status filter
        status_param = request.query_params.get('status', '')
        status_values = [s.strip() for s in status_param.split(',') if s.strip()]
        
  
        
        # Check if status is VACANT
        has_vacant = False
        has_employee_status = False
        employee_status_ids = []
        
        try:
            vacant_statuses = EmployeeStatus.objects.filter(
                Q(name__iexact='VACANT') | Q(name__iexact='VACANCY')
            )
            vacant_status_ids = list(vacant_statuses.values_list('id', flat=True))
    
        except Exception as e:
            logger.error(f"[STATUS] Error getting VACANT statuses: {e}")
            vacant_status_ids = []
   
        
        for status_val in status_values:
            is_vacant_status = False
            
            if status_val.upper() in ['VACANT', 'VACANCY']:
                is_vacant_status = True
              
            
            try:
                status_id = int(status_val)
                if status_id in vacant_status_ids:
                    is_vacant_status = True
                  
                else:
                    employee_status_ids.append(status_id)
                 
            except (ValueError, TypeError):
                try:
                    status_obj = EmployeeStatus.objects.get(name__iexact=status_val)
                    if status_obj.id in vacant_status_ids:
                        is_vacant_status = True
              
                    else:
                        employee_status_ids.append(status_obj.id)
                        
                except EmployeeStatus.DoesNotExist:
                    logger.warning(f"[STATUS] '{status_val}' not found in database")
            
            if is_vacant_status:
                has_vacant = True
            else:
                has_employee_status = True
        
    
        
        # Decide what to include
        if not status_values:
            include_employees = True
            include_vacancies = True
        
        elif has_vacant and not has_employee_status:
            include_employees = False
            include_vacancies = True
       
        elif has_employee_status and not has_vacant:
            include_employees = True
            include_vacancies = False

        else:
            include_employees = True
            include_vacancies = True
         
        
        # ====== GET EMPLOYEES WITH ACCESS CONTROL ======
        if include_employees:
            employee_queryset = self.get_queryset()  # Already filtered by access
       
            
            if employee_status_ids:
                employee_queryset = employee_queryset.filter(status__id__in=employee_status_ids)
          
            
            filter_params = request.query_params.copy()
            if 'status' in filter_params:
                del filter_params['status']
            
            employee_filter = ComprehensiveEmployeeFilter(employee_queryset, filter_params)
            filtered_employees = employee_filter.filter()
      
        else:
            filtered_employees = Employee.objects.none()
       
        
        # ====== GET VACANCIES WITH ACCESS CONTROL ======
        if include_vacancies:
            vacancy_queryset = VacantPosition.objects.filter(
                is_filled=False,
                is_deleted=False,
                include_in_headcount=True
            ).select_related(
                'business_function', 'department', 'unit', 'job_function',
                'position_group', 'vacancy_status', 'reporting_to'
            )
            
     
            
            # ✅ CRITICAL: Apply access control to vacancies
            if not access['can_view_all']:
                if access['is_manager'] and access['accessible_business_functions']:
                    # Manager can only see vacancies in their business functions
                    vacancy_queryset = vacancy_queryset.filter(
                        business_function_id__in=access['accessible_business_functions']
                    )
                    
                else:
                    # Regular employee - NO ACCESS to vacancies
                    vacancy_queryset = VacantPosition.objects.none()
                    
            
            # Apply organizational filters
            filter_params = request.query_params.copy()
            if 'status' in filter_params:
                del filter_params['status']
            
            vacancy_filter_q = self._get_vacancy_filter_from_employee_params(filter_params)
            if vacancy_filter_q:
                vacancy_queryset = vacancy_queryset.filter(vacancy_filter_q)
   
            
            filtered_vacancies = vacancy_queryset
          
        else:
            filtered_vacancies = VacantPosition.objects.none()
           
        
        # ====== BUILD UNIFIED DATA ======
        unified_data = []
        
        # Add employees
        if filtered_employees.exists():
            employee_serializer = EmployeeListSerializer(filtered_employees, many=True, context={'request': request})
            for emp_data in employee_serializer.data:
                emp_data['is_vacancy'] = False
                emp_data['record_type'] = 'employee'
                unified_data.append(emp_data)
            
        
        # Add vacancies
        if filtered_vacancies.exists():
            for vacancy in filtered_vacancies:
                vacancy_data = self._convert_vacancy_to_employee_format(vacancy, request)
                unified_data.append(vacancy_data)
            
        
        
        
        # Apply sorting
        sorting_params = self._get_sorting_params_from_request(request)
        if sorting_params:
            unified_data = self._sort_unified_data(unified_data, sorting_params)
            
        else:
            unified_data.sort(key=lambda x: x.get('name', ''))
            
        
        # Return response
        if should_paginate:
            return self._paginate_unified_data(unified_data, request)
        else:
            return Response({
                'count': len(unified_data),
                'pagination_used': False,
                'results': unified_data,
                'summary': {
                    'total_records': len(unified_data),
                    'employee_records': filtered_employees.count() if include_employees else 0,
                    'vacancy_records': filtered_vacancies.count() if include_vacancies else 0,
                    'includes_vacancies': include_vacancies,
                    'includes_employees': include_employees,
                    'status_filter': status_values,
                    'has_vacant_filter': has_vacant,
                    'has_employee_status_filter': has_employee_status,
                    'mode': 'only_vacancies' if (not include_employees and include_vacancies) else
                            'only_employees' if (include_employees and not include_vacancies) else
                            'both' if (include_employees and include_vacancies) else 'none',
                },
                'access_info': {
                    'can_view_all': access['can_view_all'],
                    'is_manager': access['is_manager'],
                    'accessible_business_functions': access['accessible_business_functions'] if not access['can_view_all'] else None,
                    'vacancy_access_applied': not access['can_view_all']
                }
            })
    def _get_vacancy_filter_from_employee_params(self, params):
        """Convert employee filter parameters to vacancy filters where applicable"""
        filters = Q()
        
        
        
        # Business function filter
        business_function_ids = self._get_int_list_param(params, 'business_function')
        if business_function_ids:
            filters &= Q(business_function__id__in=business_function_ids)
            
        
        # Department filter
        department_ids = self._get_int_list_param(params, 'department')
        if department_ids:
            filters &= Q(department__id__in=department_ids)
           
        
        # Unit filter
        unit_ids = self._get_int_list_param(params, 'unit')
        if unit_ids:
            filters &= Q(unit__id__in=unit_ids)
            
        
        # Job function filter
        job_function_ids = self._get_int_list_param(params, 'job_function')
        if job_function_ids:
            filters &= Q(job_function__id__in=job_function_ids)
            
        
        # Position group filter
        position_group_ids = self._get_int_list_param(params, 'position_group')
        if position_group_ids:
            filters &= Q(position_group__id__in=position_group_ids)
            
        
        # Grading level filter
        grading_levels = self._get_string_list_param(params, 'grading_level')
        if grading_levels:
            filters &= Q(grading_level__in=grading_levels)
            
        
        # Line manager filter (reporting_to in vacancy)
        line_manager_ids = self._get_int_list_param(params, 'line_manager')
        if line_manager_ids:
            filters &= Q(reporting_to__id__in=line_manager_ids)
          
        
        # General search
        search = params.get('search')
        if search:
            filters &= (
                Q(job_title__icontains=search) |
                Q(position_id__icontains=search) |
                Q(business_function__name__icontains=search) |
                Q(department__name__icontains=search) |
                Q(notes__icontains=search)
            )
           
        
        # Job title search
        job_title_search = params.get('job_title_search')
        if job_title_search:
            filters &= Q(job_title__icontains=job_title_search)
           
        
        # Department search
        department_search = params.get('department_search')
        if department_search:
            filters &= Q(department__name__icontains=department_search)
            
        
        # Org chart visibility
        is_visible_in_org_chart = params.get('is_visible_in_org_chart')
        if is_visible_in_org_chart:
            visible = is_visible_in_org_chart.lower() == 'true'
            filters &= Q(is_visible_in_org_chart=visible)
            
        
        if filters.children:
            logger.info(f"[VAC FILTER] Total filters applied: {len(filters.children)}")
        else:
            logger.info("[VAC FILTER] No filters applied")
        
        return filters if filters.children else None
    
    def _get_int_list_param(self, params, param_name):
        """Helper to get integer list from parameters"""
        values = []
        if hasattr(params, 'getlist'):
            param_values = params.getlist(param_name)
        else:
            param_values = [params.get(param_name)] if params.get(param_name) else []
        
        for value in param_values:
            if value:
                if ',' in str(value):
                    values.extend([int(v.strip()) for v in str(value).split(',') if v.strip().isdigit()])
                elif str(value).isdigit():
                    values.append(int(value))
        
        return values
    
    def _get_string_list_param(self, params, param_name):
        """Helper to get string list from parameters"""
        values = []
        if hasattr(params, 'getlist'):
            param_values = params.getlist(param_name)
        else:
            param_values = [params.get(param_name)] if params.get(param_name) else []
        
        for value in param_values:
            if value:
                if ',' in str(value):
                    values.extend([v.strip() for v in str(value).split(',') if v.strip()])
                else:
                    values.append(str(value).strip())
        
        return values
    
    def _convert_vacancy_to_employee_format(self, vacancy, request):
        """Convert vacancy to employee-like format for unified display"""
        
        # ✅ CRITICAL FIX: Ensure vacancy status is properly set
        vacancy_status_name = 'VACANT'
        vacancy_status_color = '#F97316'  # Orange color for vacancies
        
        if vacancy.vacancy_status:
            vacancy_status_name = vacancy.vacancy_status.name
            vacancy_status_color = vacancy.vacancy_status.color
        
        return {
            'id': vacancy.original_employee_pk or vacancy.id,  
            'employee_id': vacancy.position_id,
            'name': "VACANT",
            'email': None,
            'father_name': None,
            'date_of_birth': None,
            'gender': None,
            'phone': None,
            'business_function_name': vacancy.business_function.name if vacancy.business_function else 'N/A',
            'business_function_code': vacancy.business_function.code if vacancy.business_function else 'N/A',
            'business_function_id': vacancy.business_function.id if vacancy.business_function else None,
            'department_name': vacancy.department.name if vacancy.department else 'N/A',
            'department_id': vacancy.department.id if vacancy.department else None,
            'unit_name': vacancy.unit.name if vacancy.unit else None,
            'unit_id': vacancy.unit.id if vacancy.unit else None,
            'job_function_name': vacancy.job_function.name if vacancy.job_function else 'N/A',
            'job_function_id': vacancy.job_function.id if vacancy.job_function else None,
            'job_title': vacancy.job_title,
            'position_group_name': vacancy.position_group.get_name_display() if vacancy.position_group else 'N/A',
            'position_group_level': vacancy.position_group.hierarchy_level if vacancy.position_group else 0,
            'position_group_id': vacancy.position_group.id if vacancy.position_group else None,
            'grading_level': vacancy.grading_level,
            'start_date': None,
            'end_date': None,
            'contract_duration': 'VACANT',
            'contract_duration_display': 'Vacant Position',
            'contract_start_date': None,
            'contract_end_date': None,
            'contract_extensions': 0,
            'last_extension_date': None,
            'line_manager_name': vacancy.reporting_to.full_name if vacancy.reporting_to else None,
            'line_manager_hc_number': vacancy.reporting_to.employee_id if vacancy.reporting_to else None,
            'line_manager_email': vacancy.reporting_to.user.email if (vacancy.reporting_to and vacancy.reporting_to.user) else None,
            
            # ✅ CRITICAL: Status fields that frontend checks
            'status_name': vacancy_status_name,
            'status_color': vacancy_status_color,
            'current_status_display': vacancy_status_name,
            
            'tag_names': [],
            'years_of_service': 0,
            'is_visible_in_org_chart': vacancy.is_visible_in_org_chart,
            'direct_reports_count': 0,
            'status_needs_update': False,
            'created_at': vacancy.created_at,
            'updated_at': vacancy.updated_at,
            'profile_image_url': None,
            'is_deleted': False,
            
            # ✅ CRITICAL: Mark as vacancy for filtering
            'is_vacancy': True,
            'record_type': 'vacancy',
            
            'vacancy_details': {
                'internal_id': vacancy.id,
                'position_id': vacancy.position_id,
                'include_in_headcount': vacancy.include_in_headcount,
                'is_filled': vacancy.is_filled,
                'filled_date': vacancy.filled_date,
                'notes': vacancy.notes,
                'original_employee_pk': vacancy.original_employee_pk
            }
        }
    
    def _get_sorting_params_from_request(self, request):
        """Extract sorting parameters from request"""
        sorting_data = request.query_params.get('sorting')
        if sorting_data:
            try:
                import json
                return json.loads(sorting_data)
            except:
                pass
        
        # Fallback to ordering parameter
        ordering = request.query_params.get('ordering', '')
        if ordering:
            sort_params = [param.strip() for param in ordering.split(',') if param.strip()]
            sorting_params = []
            for param in sort_params:
                if param.startswith('-'):
                    sorting_params.append({'field': param[1:], 'direction': 'desc'})
                else:
                    sorting_params.append({'field': param, 'direction': 'asc'})
            return sorting_params
        
        return []
    
    def _sort_unified_data(self, data, sorting_params):
        """Sort unified employee and vacancy data - FIXED for date fields"""
        from datetime import datetime, date
        
        def parse_date_value(value):
            """Parse date value safely"""
            if value is None:
                return None
            
            # Already a date/datetime object
            if isinstance(value, (datetime, date)):
                return value
            
            # String date
            if isinstance(value, str):
                try:
                    # Try ISO format
                    return datetime.fromisoformat(value.replace('Z', '+00:00'))
                except:
                    try:
                        # Try date-only format
                        return datetime.strptime(value, '%Y-%m-%d')
                    except:
                        return None
            
            return None
            
        def get_sort_key(item, field, direction):
            value = item.get(field, '')
            
            # ✅ CRITICAL: Handle date fields specially
            if field in ['start_date', 'end_date', 'created_at', 'updated_at', 
                         'contract_start_date', 'contract_end_date', 'date_of_birth']:
                parsed_date = parse_date_value(value)
                if parsed_date:
                    return parsed_date
                # Return min/max datetime for None values
                return datetime.min if direction == 'asc' else datetime.max
            
            # Handle None values for other fields
            if value is None:
                return '' if direction == 'asc' else 'z' * 100
            
            # Handle numbers
            if field in ['years_of_service', 'direct_reports_count', 'position_group_level']:
                try:
                    return float(value) if value else 0
                except:
                    return 0
            
            # Handle strings (case-insensitive)
            return str(value).lower()
        
        # Apply multi-level sorting
        for sort_param in reversed(sorting_params):
            field = sort_param.get('field', '')
            direction = sort_param.get('direction', 'asc')
            
            if field:
                try:
                    data.sort(
                        key=lambda x: get_sort_key(x, field, direction),
                        reverse=(direction == 'desc')
                    )
                   
                except Exception as e:
                    logger.error(f"Error sorting by {field}: {e}")
                    # Skip this sort field if it fails
                    continue
        
        return data
    
    def _paginate_unified_data(self, data, request):
        """Apply pagination to unified data"""
        page_size = int(request.query_params.get('page_size', 20))
        page = int(request.query_params.get('page', 1))
        
        # Calculate pagination
        total_count = len(data)
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        
        paginated_data = data[start_index:end_index]
        
        # Calculate pagination info
        total_pages = (total_count + page_size - 1) // page_size
        has_next = page < total_pages
        has_previous = page > 1
        
        # Calculate page numbers for display
        start_page = max(1, page - 2)
        end_page = min(total_pages, page + 2)
        
        if end_page - start_page < 4:
            if start_page == 1:
                end_page = min(total_pages, start_page + 4)
            else:
                start_page = max(1, end_page - 4)
        
        page_numbers = list(range(start_page, end_page + 1))
        
        # Calculate range display
        start_item = start_index + 1 if paginated_data else 0
        end_item = min(end_index, total_count)
        
        # Count records by type
        employee_count = len([item for item in data if not item.get('is_vacancy', False)])
        vacancy_count = len([item for item in data if item.get('is_vacancy', False)])
        
        return Response({
            'count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'page_size': page_size,
            'page_size_options': [10, 20, 50, 100, 500, 1000, "All"],
            'has_next': has_next,
            'has_previous': has_previous,
            'next': None,
            'previous': None,
            'page_numbers': page_numbers,
            'start_page': start_page,
            'end_page': end_page,
            'show_first': start_page > 1,
            'show_last': end_page < total_pages,
            'range_display': f"Showing {start_item}-{end_item} of {total_count}",
            'pagination_used': True,
            'results': paginated_data,
            'summary': {
                'total_records': total_count,
                'employee_records': employee_count,
                'vacancy_records': vacancy_count,
                'includes_vacancies': True,
                'unified_view': True,
                'current_page_employees': len([item for item in paginated_data if not item.get('is_vacancy', False)]),
                'current_page_vacancies': len([item for item in paginated_data if item.get('is_vacancy', False)])
            }
        })
    @swagger_auto_schema(
        auto_schema=FileUploadAutoSchema,
        operation_description="Create a new employee with optional document and profile photo",
        manual_parameters=[
            # Required fields
            openapi.Parameter('first_name', openapi.IN_FORM, description="First name", type=openapi.TYPE_STRING, required=True),
            openapi.Parameter('last_name', openapi.IN_FORM, description="Last name", type=openapi.TYPE_STRING, required=True),
            openapi.Parameter('email', openapi.IN_FORM, description="Email", type=openapi.TYPE_STRING, required=True),
           
            openapi.Parameter('job_title', openapi.IN_FORM, description="Job title", type=openapi.TYPE_STRING, required=True),
            openapi.Parameter('start_date', openapi.IN_FORM, description="Start date (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=True),
            openapi.Parameter('business_function', openapi.IN_FORM, description="Business function ID", type=openapi.TYPE_INTEGER, required=True),
            openapi.Parameter('department', openapi.IN_FORM, description="Department ID", type=openapi.TYPE_INTEGER, required=True),
            openapi.Parameter('job_function', openapi.IN_FORM, description="Job function ID", type=openapi.TYPE_INTEGER, required=True),
            openapi.Parameter('position_group', openapi.IN_FORM, description="Position group ID", type=openapi.TYPE_INTEGER, required=True),
            openapi.Parameter('contract_duration', openapi.IN_FORM, description="Contract duration", type=openapi.TYPE_STRING, required=True),
            
            # Optional basic fields
            openapi.Parameter('father_name', openapi.IN_FORM, description="Father name", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('date_of_birth', openapi.IN_FORM, description="Date of birth (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('gender', openapi.IN_FORM, description="Gender", type=openapi.TYPE_STRING, enum=['MALE', 'FEMALE'], required=False),
            openapi.Parameter('phone', openapi.IN_FORM, description="Phone number", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('address', openapi.IN_FORM, description="Address", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('emergency_contact', openapi.IN_FORM, description="Emergency contact", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('unit', openapi.IN_FORM, description="Unit ID", type=openapi.TYPE_INTEGER, required=False),
            openapi.Parameter('grading_level', openapi.IN_FORM, description="Grading level", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('contract_start_date', openapi.IN_FORM, description="Contract start date (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('line_manager', openapi.IN_FORM, description="Line manager ID", type=openapi.TYPE_INTEGER, required=False),
            openapi.Parameter('is_visible_in_org_chart', openapi.IN_FORM, description="Visible in org chart", type=openapi.TYPE_BOOLEAN, required=False),
            openapi.Parameter('notes', openapi.IN_FORM, description="Notes", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('end_date', openapi.IN_FORM, description="End date (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=False),
            
            # File fields
            openapi.Parameter(
                'document',
                openapi.IN_FORM,
                description="Employee document file (PDF, DOC, DOCX, TXT, XLS, XLSX)",
                type=openapi.TYPE_FILE,
                required=False
            ),
            openapi.Parameter(
                'profile_photo',
                openapi.IN_FORM,
                description="Profile photo (JPG, PNG, GIF, BMP)",
                type=openapi.TYPE_FILE,
                required=False
            ),
            openapi.Parameter(
                'document_type',
                openapi.IN_FORM,
                description="Document type",
                type=openapi.TYPE_STRING,
                enum=['CONTRACT', 'ID', 'CERTIFICATE', 'CV', 'PERFORMANCE', 'MEDICAL', 'TRAINING', 'OTHER'],
                required=False
            ),
            openapi.Parameter(
                'document_name',
                openapi.IN_FORM,
                description="Document name (optional, will use filename if not provided)",
                type=openapi.TYPE_STRING,
                required=False
            ),
        ],
        responses={
            201: openapi.Response(description="Employee created successfully", schema=EmployeeDetailSerializer),
            400: openapi.Response(description="Bad request - validation errors"),
        }
    )
    def create(self, request, *args, **kwargs):
        """Create a new employee with optional document and profile photo"""
        return super().create(request, *args, **kwargs)
    
    @swagger_auto_schema(
    auto_schema=FileUploadAutoSchema,
    operation_description="Update an existing employee with optional document and profile photo",
    manual_parameters=[
        # All fields are optional for update
        openapi.Parameter('first_name', openapi.IN_FORM, description="First name", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('last_name', openapi.IN_FORM, description="Last name", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('email', openapi.IN_FORM, description="Email", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('job_title', openapi.IN_FORM, description="Job title", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('start_date', openapi.IN_FORM, description="Start date (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('business_function', openapi.IN_FORM, description="Business function ID", type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('department', openapi.IN_FORM, description="Department ID", type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('job_function', openapi.IN_FORM, description="Job function ID", type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('position_group', openapi.IN_FORM, description="Position group ID", type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('contract_duration', openapi.IN_FORM, description="Contract duration", type=openapi.TYPE_STRING, required=False),
        
        # Optional fields
        openapi.Parameter('father_name', openapi.IN_FORM, description="Father name", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('date_of_birth', openapi.IN_FORM, description="Date of birth (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('gender', openapi.IN_FORM, description="Gender", type=openapi.TYPE_STRING, enum=['MALE', 'FEMALE'], required=False),
        openapi.Parameter('phone', openapi.IN_FORM, description="Phone number", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('address', openapi.IN_FORM, description="Address", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('emergency_contact', openapi.IN_FORM, description="Emergency contact", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('unit', openapi.IN_FORM, description="Unit ID", type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('grading_level', openapi.IN_FORM, description="Grading level", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('contract_start_date', openapi.IN_FORM, description="Contract start date (YYYY-MM-DD)", type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('line_manager', openapi.IN_FORM, description="Line manager ID", type=openapi.TYPE_INTEGER, required=False),
        openapi.Parameter('is_visible_in_org_chart', openapi.IN_FORM, description="Visible in org chart", type=openapi.TYPE_BOOLEAN, required=False),
        openapi.Parameter('notes', openapi.IN_FORM, description="Notes", type=openapi.TYPE_STRING, required=False),
        
        # File fields
        openapi.Parameter(
            'document',
            openapi.IN_FORM,
            description="Employee document file (PDF, DOC, DOCX, TXT, XLS, XLSX)",
            type=openapi.TYPE_FILE,
            required=False
        ),
        openapi.Parameter(
            'profile_photo',
            openapi.IN_FORM,
            description="Profile photo (JPG, PNG, GIF, BMP)",
            type=openapi.TYPE_FILE,
            required=False
        ),
        openapi.Parameter(
            'document_type',
            openapi.IN_FORM,
            description="Document type",
            type=openapi.TYPE_STRING,
            enum=['CONTRACT', 'ID', 'CERTIFICATE', 'CV', 'PERFORMANCE', 'MEDICAL', 'TRAINING', 'OTHER'],
            required=False
        ),
        openapi.Parameter(
            'document_name',
            openapi.IN_FORM,
            description="Document name (optional, will use filename if not provided)",
            type=openapi.TYPE_STRING,
            required=False
        ),
    ],
    responses={
        200: openapi.Response(description="Employee updated successfully", schema=EmployeeDetailSerializer),
        400: openapi.Response(description="Bad request - validation errors"),
        404: openapi.Response(description="Employee not found"),
    }
)
    def update(self, request, *args, **kwargs):
        """✅ SIMPLIFIED: Let serializer handle everything"""
        try:
            # Get the employee
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            

            
            # ✅ FIX: Use serializer's update method - it handles everything
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            
            # Save via serializer (calls serializer's update method)
            self.perform_update(serializer)
            
    
            
            # Return updated data
            return Response(serializer.data)
            
        except Exception as e:
        
            return Response(
                {'error': f'Employee update failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    @swagger_auto_schema(
        method='post',
        operation_description="Toggle org chart visibility for single employee",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['employee_id'],
            properties={
                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER, description='Employee ID')
            }
        ),
        responses={
            200: openapi.Response(
                description="Org chart visibility toggled successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'employee_name': openapi.Schema(type=openapi.TYPE_STRING),
                        'is_visible_in_org_chart': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                    }
                )
            ),
            400: openapi.Response(description="Bad request"),
            404: openapi.Response(description="Employee not found")
        }
    )
    @action(detail=False, methods=['post'], url_path='toggle-org-chart-visibility')
    def toggle_org_chart_visibility(self, request):
        """Toggle org chart visibility for single employee"""
        try:
            employee_id = request.data.get('employee_id')
            
            if not employee_id:
                return Response(
                    {'error': 'employee_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            try:
                employee = Employee.objects.get(id=employee_id)
            except Employee.DoesNotExist:
                return Response(
                    {'error': 'Employee not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Toggle the visibility
            old_visibility = employee.is_visible_in_org_chart
            employee.is_visible_in_org_chart = not old_visibility
            employee.updated_by = request.user
            employee.save()
            
            # Log activity
            EmployeeActivity.objects.create(
                employee=employee,
                activity_type='UPDATED',
                description=f"Org chart visibility changed from {old_visibility} to {employee.is_visible_in_org_chart}",
                performed_by=request.user,
                metadata={
                    'field_changed': 'is_visible_in_org_chart',
                    'old_value': old_visibility,
                    'new_value': employee.is_visible_in_org_chart,
                    'action': 'toggle_org_chart_visibility'
                }
            )
            
            visibility_text = "visible" if employee.is_visible_in_org_chart else "hidden"
            
            return Response({
                'success': True,
                'message': f'{employee.full_name} is now {visibility_text} in org chart',
                'employee_id': employee.id,
                'employee_name': employee.full_name,
                'is_visible_in_org_chart': employee.is_visible_in_org_chart,
                'previous_visibility': old_visibility
            })
            
        except Exception as e:
            logger.error(f"Toggle org chart visibility failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to toggle org chart visibility: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='post',
        operation_description="Bulk toggle org chart visibility for multiple employees",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['employee_ids'],
            properties={
                'employee_ids': openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Schema(type=openapi.TYPE_INTEGER),
                    description='List of employee IDs to toggle'
                ),
                'set_visible': openapi.Schema(
                    type=openapi.TYPE_BOOLEAN,
                    description='Optional: Set specific visibility (true/false). If not provided, will toggle each employee individually.'
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="Bulk org chart visibility update completed",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'total_employees': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'updated_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'set_visible_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'set_hidden_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'results': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                    'employee_name': openapi.Schema(type=openapi.TYPE_STRING),
                                    'old_visibility': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    'new_visibility': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    'action': openapi.Schema(type=openapi.TYPE_STRING)
                                }
                            )
                        )
                    }
                )
            ),
            400: openapi.Response(description="Bad request")
        }
    )
    @action(detail=False, methods=['post'], url_path='bulk-toggle-org-chart-visibility')
    def bulk_toggle_org_chart_visibility(self, request):
        """Bulk toggle org chart visibility for multiple employees"""
        try:
            employee_ids = request.data.get('employee_ids', [])
            set_visible = request.data.get('set_visible')  # Optional: force set to specific value
            
            if not employee_ids:
                return Response(
                    {'error': 'employee_ids list is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not isinstance(employee_ids, list):
                return Response(
                    {'error': 'employee_ids must be a list'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            employees = Employee.objects.filter(id__in=employee_ids)
            
            if employees.count() != len(employee_ids):
                return Response(
                    {'error': 'Some employee IDs were not found'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            updated_count = 0
            set_visible_count = 0
            set_hidden_count = 0
            results = []
            
            with transaction.atomic():
                for employee in employees:
                    old_visibility = employee.is_visible_in_org_chart
                    
                    # Determine new visibility
                    if set_visible is not None:
                        # Force set to specific value
                        new_visibility = bool(set_visible)
                        action = 'set_visible' if new_visibility else 'set_hidden'
                    else:
                        # Toggle current value
                        new_visibility = not old_visibility
                        action = 'toggled'
                    
                    # Only update if visibility actually changes
                    if old_visibility != new_visibility:
                        employee.is_visible_in_org_chart = new_visibility
                        employee.updated_by = request.user
                        employee.save()
                        
                        # Log activity
                        EmployeeActivity.objects.create(
                            employee=employee,
                            activity_type='UPDATED',
                            description=f"Org chart visibility bulk changed from {old_visibility} to {new_visibility}",
                            performed_by=request.user,
                            metadata={
                                'field_changed': 'is_visible_in_org_chart',
                                'old_value': old_visibility,
                                'new_value': new_visibility,
                                'action': 'bulk_toggle_org_chart_visibility',
                                'bulk_operation': True
                            }
                        )
                        
                        updated_count += 1
                        
                        if new_visibility:
                            set_visible_count += 1
                        else:
                            set_hidden_count += 1
                    
                    results.append({
                        'employee_id': employee.id,
                        'employee_name': employee.full_name,
                        'old_visibility': old_visibility,
                        'new_visibility': new_visibility,
                        'action': action,
                        'changed': old_visibility != new_visibility
                    })
            
            return Response({
                'success': True,
                'message': f'Org chart visibility update completed: {updated_count} employees updated',
                'total_employees': len(employee_ids),
                'updated_count': updated_count,
                'set_visible_count': set_visible_count,
                'set_hidden_count': set_hidden_count,
                'results': results
            })
            
        except Exception as e:
            logger.error(f"Bulk toggle org chart visibility failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to toggle org chart visibility: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """Override destroy to use soft delete"""
        instance = self.get_object()
        instance.soft_delete(user=request.user)
        
        # Log activity
        EmployeeActivity.objects.create(
            employee=instance,
            activity_type='SOFT_DELETED',
            description=f"Employee {instance.full_name} was soft deleted",
            performed_by=request.user
        )
        
        return Response(status=status.HTTP_204_NO_CONTENT)

    # views.py - EmployeeViewSet ASSET ENDPOINTS UPDATE

    @swagger_auto_schema(
        method='get',
        operation_description="Get employee's assigned assets with detailed status and clarification info",
        responses={
            200: openapi.Response(
                description="Assets retrieved successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'employee': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'assets': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                        'summary': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'pending_actions': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                    }
                )
            ),
            403: "Access denied",
            404: "Employee not found"
        }
    )
    @action(detail=True, methods=['get'], url_path='assets')
    def get_employee_assets(self, request, pk=None):
        
        try:
            employee = self.get_object()
            
            # Check access permissions
            from .asset_permissions import get_asset_access_level
            access = get_asset_access_level(request.user)
            
            # Verify access
            can_view = False
            if access['can_view_all_assets']:
                can_view = True
            elif employee.user and employee.user.id == request.user.id:
                can_view = True
            elif access['accessible_employee_ids'] and employee.id in access['accessible_employee_ids']:
                can_view = True
            
            if not can_view:
                return Response(
                    {
                        'error': 'Access Denied',
                        'message': 'You do not have permission to view this employee\'s assets'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Get assets
            from .asset_models import Asset
            assets = Asset.objects.filter(assigned_to=employee).select_related(
                'batch', 'category', 'batch__category'
            ).prefetch_related('assignments').order_by('-updated_at')
            
            # Categorize by status
            asset_data = []
            pending_actions = []
            
            for asset in assets:
                # Get current assignment
                current_assignment = asset.assignments.filter(check_in_date__isnull=True).first()
                
                asset_info = {
                    'id': str(asset.id),
                    'asset_number': asset.asset_number,
                    'asset_name': asset.asset_name,
                    'serial_number': asset.serial_number,
                    'category': asset.category.name if asset.category else None,
                    'batch_number': asset.batch.batch_number if asset.batch else None,
                    'batch_info': {
                        'id': asset.batch.id if asset.batch else None,
                        'batch_number': asset.batch.batch_number if asset.batch else None,
                        'purchase_date': asset.batch.purchase_date if asset.batch else None,
                        'unit_price': float(asset.batch.unit_price) if asset.batch else None,
                    } if asset.batch else None,
                    'status': asset.status,
                    'status_display': asset.get_status_display(),
                    
                    # ✅ Action permissions
                    'can_accept': asset.can_be_approved(),
                    'can_request_clarification': asset.can_request_clarification(),
                    'can_cancel': asset.status in ['ASSIGNED', 'NEED_CLARIFICATION'] and access['can_manage_all_assets'],
                    
                    # Assignment details
                    'assignment': {
                        'check_out_date': current_assignment.check_out_date.isoformat() if current_assignment else None,
                        'check_out_notes': current_assignment.check_out_notes if current_assignment else None,
                        'condition': current_assignment.condition_on_checkout if current_assignment else None,
                        'days_assigned': current_assignment.get_duration_days() if current_assignment else 0,
                        'assigned_by': (
                            current_assignment.assigned_by.get_full_name() 
                            if current_assignment and current_assignment.assigned_by else None
                        )
                    } if current_assignment else None,
                    
                    # ✅ Clarification info
                    'clarification': self._get_asset_clarification_detail(asset)
                }
                
                asset_data.append(asset_info)
                
                # Build pending actions list
                if asset.status == 'ASSIGNED':
                    pending_actions.append({
                        'asset_id': str(asset.id),
                        'asset_name': asset.asset_name,
                        'action_type': 'ACCEPT',
                        'description': 'Approve this asset assignment',
                        'priority': 'high',
                        'days_waiting': current_assignment.get_duration_days() if current_assignment else 0
                    })
                elif asset.status == 'NEED_CLARIFICATION' and not asset.clarification_response:
                    pending_actions.append({
                        'asset_id': str(asset.id),
                        'asset_name': asset.asset_name,
                        'action_type': 'CLARIFICATION_PENDING',
                        'description': 'Waiting for clarification response',
                        'priority': 'medium',
                        'requested_reason': asset.clarification_requested_reason
                    })
            
            # Summary by status
            summary = {
                'total_assets': assets.count(),
                'by_status': {},
                'by_category': {},
                'pending_actions_count': len(pending_actions)
            }
            
            # Count by status
            for status_choice in Asset.STATUS_CHOICES:
                status_code = status_choice[0]
                count = assets.filter(status=status_code).count()
                if count > 0:
                    summary['by_status'][status_code] = {
                        'label': status_choice[1],
                        'count': count
                    }
            
            # Count by category
            from django.db.models import Count
            category_counts = assets.values('category__name').annotate(count=Count('id'))
            for cat in category_counts:
                if cat['category__name']:
                    summary['by_category'][cat['category__name']] = cat['count']
            
            return Response({
                'employee': {
                    'id': employee.id,
                    'employee_id': employee.employee_id,
                    'name': employee.full_name,
                    'job_title': employee.job_title,
                    'department': employee.department.name if employee.department else None
                },
                'assets': asset_data,
                'summary': summary,
                'pending_actions': pending_actions,
                'access_info': {
                    'can_accept': employee.user and employee.user.id == request.user.id,
                    'can_manage': access['can_manage_all_assets'],
                    'access_level': access['access_level']
                }
            })
            
        except Exception as e:
            logger.error(f"Error getting employee assets: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve assets: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _get_asset_clarification_detail(self, asset):
        """Get detailed clarification information"""
        if asset.status == 'NEED_CLARIFICATION' or (
            hasattr(asset, 'clarification_requested_reason') and asset.clarification_requested_reason
        ):
            return {
                'has_clarification': True,
                'requested': {
                    'reason': asset.clarification_requested_reason,
                    'requested_at': asset.clarification_requested_at.isoformat() if asset.clarification_requested_at else None,
                    'requested_by': (
                        asset.clarification_requested_by.get_full_name() 
                        if asset.clarification_requested_by else None
                    )
                },
                'response': {
                    'text': asset.clarification_response,
                    'provided_at': asset.clarification_provided_at.isoformat() if asset.clarification_provided_at else None,
                    'provided_by': (
                        asset.clarification_provided_by.get_full_name() 
                        if asset.clarification_provided_by else None
                    )
                } if asset.clarification_response else None,
                'status': 'pending' if not asset.clarification_response else 'resolved',
                'is_pending': not bool(asset.clarification_response)
            }
        return None
    
    @swagger_auto_schema(
        method='post',
        operation_description="Accept assigned asset (Employee approval)",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['asset_id'],
            properties={
                'asset_id': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    format='uuid',
                    description='Asset ID to accept'
                ),
                'comments': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='Optional acceptance comments'
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="Asset accepted successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'asset': openapi.Schema(type=openapi.TYPE_OBJECT)
                    }
                )
            ),
            400: "Bad request",
            403: "Access denied",
            404: "Asset not found"
        }
    )
    @action(detail=True, methods=['post'], url_path='accept-asset')
    def accept_assigned_asset(self, request, pk=None):
        """✅ Employee accepts an assigned asset"""
        try:
            employee = self.get_object()
            
            # Verify this is the employee's own action
            if not employee.user or employee.user.id != request.user.id:
                return Response(
                    {'error': 'You can only accept assets assigned to you'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            asset_id = request.data.get('asset_id')
            comments = request.data.get('comments', '')
            
            if not asset_id:
                return Response(
                    {'error': 'asset_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from .asset_models import Asset, AssetActivity
            
            try:
                asset = Asset.objects.get(id=asset_id, assigned_to=employee)
            except Asset.DoesNotExist:
                return Response(
                    {'error': 'Asset not found or not assigned to this employee'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Check if asset can be approved
            if not asset.can_be_approved():
                return Response(
                    {'error': f'Asset cannot be accepted. Current status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                # Update asset status from ASSIGNED to IN_USE
                asset.status = 'IN_USE'
                asset.updated_by = request.user
                asset.save()
                
                # Log asset activity
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='ACCEPTED',
                    description=f"Asset accepted by {employee.full_name}",
                    performed_by=request.user,
                    metadata={
                        'comments': comments,
                        'employee_id': employee.employee_id,
                        'employee_name': employee.full_name,
                        'acceptance_date': timezone.now().isoformat(),
                        'previous_status': 'ASSIGNED',
                        'new_status': 'IN_USE'
                    }
                )
                
                # Log employee activity
                from .models import EmployeeActivity
                EmployeeActivity.objects.create(
                    employee=employee,
                    activity_type='ASSET_ACCEPTED',
                    description=f"Accepted asset {asset.asset_name} - {asset.serial_number}",
                    performed_by=request.user,
                    metadata={
                        'asset_id': str(asset.id),
                        'asset_number': asset.asset_number,
                        'asset_name': asset.asset_name,
                        'serial_number': asset.serial_number,
                        'comments': comments
                    }
                )
            
            logger.info(f"✅ Asset {asset.asset_number} accepted by {employee.full_name}")
            
            return Response({
                'success': True,
                'message': f'Asset {asset.asset_name} accepted successfully',
                'asset': {
                    'id': str(asset.id),
                    'asset_number': asset.asset_number,
                    'name': asset.asset_name,
                    'serial_number': asset.serial_number,
                    'status': asset.status,
                    'status_display': asset.get_status_display()
                }
            })
            
        except Exception as e:
            logger.error(f"Error accepting asset: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to accept asset: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='post',
        operation_description="Request clarification for assigned asset",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['asset_id', 'clarification_reason'],
            properties={
                'asset_id': openapi.Schema(type=openapi.TYPE_STRING, format='uuid'),
                'clarification_reason': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='Reason for requesting clarification',
                    example='The serial number does not match the label on the device'
                )
            }
        ),
        responses={
            200: openapi.Response(description="Clarification requested successfully"),
            400: "Bad request",
            403: "Access denied",
            404: "Asset not found"
        }
    )
    @action(detail=True, methods=['post'], url_path='request-asset-clarification')
    def request_asset_clarification(self, request, pk=None):
        """✅ Employee requests clarification about an assigned asset"""
        try:
            employee = self.get_object()
            
            # Verify this is the employee's own action
            if not employee.user or employee.user.id != request.user.id:
                return Response(
                    {'error': 'You can only request clarification for assets assigned to you'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            asset_id = request.data.get('asset_id')
            clarification_reason = request.data.get('clarification_reason')
            
            if not asset_id or not clarification_reason:
                return Response(
                    {'error': 'asset_id and clarification_reason are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from .asset_models import Asset, AssetActivity
            
            try:
                asset = Asset.objects.get(id=asset_id, assigned_to=employee)
            except Asset.DoesNotExist:
                return Response(
                    {'error': 'Asset not found or not assigned to this employee'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Check if asset can request clarification
            if not asset.can_request_clarification():
                return Response(
                    {'error': f'Cannot request clarification for asset with status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                # Update asset status to NEED_CLARIFICATION
                old_status = asset.status
                asset.status = 'NEED_CLARIFICATION'
                asset.clarification_requested_reason = clarification_reason
                asset.clarification_requested_at = timezone.now()
                asset.clarification_requested_by = request.user
                # Clear previous clarification response
                asset.clarification_response = None
                asset.clarification_provided_at = None
                asset.clarification_provided_by = None
                asset.updated_by = request.user
                asset.save()
                
                # Log asset activity
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='CLARIFICATION_REQUESTED',
                    description=f"Clarification requested by {employee.full_name}: {clarification_reason}",
                    performed_by=request.user,
                    metadata={
                        'clarification_reason': clarification_reason,
                        'employee_id': employee.employee_id,
                        'employee_name': employee.full_name,
                        'request_date': timezone.now().isoformat(),
                        'previous_status': old_status,
                        'new_status': 'NEED_CLARIFICATION'
                    }
                )
                
                # Log employee activity
                from .models import EmployeeActivity
                EmployeeActivity.objects.create(
                    employee=employee,
                    activity_type='ASSET_CLARIFICATION_REQUESTED',
                    description=f"Requested clarification for asset {asset.asset_name} - {asset.serial_number}",
                    performed_by=request.user,
                    metadata={
                        'asset_id': str(asset.id),
                        'asset_number': asset.asset_number,
                        'asset_name': asset.asset_name,
                        'serial_number': asset.serial_number,
                        'clarification_reason': clarification_reason
                    }
                )
            
            logger.info(f"✅ Clarification requested for asset {asset.asset_number} by {employee.full_name}")
            
            return Response({
                'success': True,
                'message': f'Clarification requested for asset {asset.asset_name}',
                'asset': {
                    'id': str(asset.id),
                    'asset_number': asset.asset_number,
                    'name': asset.asset_name,
                    'serial_number': asset.serial_number,
                    'status': asset.status,
                    'status_display': asset.get_status_display(),
                    'clarification_reason': clarification_reason
                }
            })
            
        except Exception as e:
            logger.error(f"Error requesting clarification: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to request clarification: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    def _generate_bulk_template(self):
        """Generate Excel template with dropdowns and validation"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Template"
        
        # Define headers with validation requirements (ENHANCED with father_name)
        headers = [
             'Employee ID (optional - auto-generated)', 'First Name*', 'Last Name*', 'Email*',
            'Date of Birth', 'Gender', 'Father Name', 'Address', 'Phone', 'Emergency Contact',
            'Business Function*', 'Department*', 'Unit', 'Job Function*',
            'Job Title*', 'Position Group*', 'Grading Level',
            'Start Date*', 'Contract Duration*', 'Contract Start Date',
            'Line Manager Employee ID', 'Is Visible in Org Chart',
            'Tag Names (comma separated)', 'Notes'
        ]
        # Create reference sheets for dropdowns
        self._create_reference_sheets(wb)
        
        # Add data validations
        self._add_data_validations(ws)
        # Write headers
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample data row (ENHANCED with father_name)
        sample_data = [
            'HC001', 'John', 'Doe', 'john.doe@company.com',
            '1990-01-15', 'MALE', 'Robert Doe', '123 Main St, City', '+994501234567', 'Jane Doe +994501234568',
            'IT', 'Software Development', 'Backend Team', 'Software Engineer',
            'Senior Software Engineer', 'SENIOR SPECIALIST', 'SS_M',
            '2024-01-15', 'PERMANENT', '2024-01-15',
            'HC002', 'TRUE', 'SKILL:Python,STATUS:New Hire', 'New team member'
        ]
        ws.append(sample_data)
        
    
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        self._add_instructions_sheet(wb)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employee_bulk_template_{date.today()}.xlsx"'
        
        return response
    
    def _create_reference_sheets(self, workbook):
        """Create reference sheets with lookup data"""
        from openpyxl.styles import Font, PatternFill
        
        # Business Functions sheet
        bf_sheet = workbook.create_sheet(title="Business Functions")
        bf_sheet.append(['Business Function'])
        for bf in BusinessFunction.objects.filter(is_active=True).order_by('name'):
            bf_sheet.append([bf.name])
        
        # Departments sheet
        dept_sheet = workbook.create_sheet(title="Departments")
        dept_sheet.append(['Business Function', 'Department'])
        for dept in Department.objects.select_related('business_function').filter(is_active=True).order_by('business_function__name', 'name'):
            dept_sheet.append([dept.business_function.name, dept.name])
        
        # Units sheet
        unit_sheet = workbook.create_sheet(title="Units")
        unit_sheet.append(['Department', 'Unit'])
        for unit in Unit.objects.select_related('department').filter(is_active=True).order_by('department__name', 'name'):
            unit_sheet.append([unit.department.name, unit.name])
        
        # Job Functions sheet
        jf_sheet = workbook.create_sheet(title="Job Functions")
        jf_sheet.append(['Job Function'])
        for jf in JobFunction.objects.filter(is_active=True).order_by('name'):
            jf_sheet.append([jf.name])
        
        # Position Groups sheet
        pg_sheet = workbook.create_sheet(title="Position Groups")
        pg_sheet.append(['Position Group', 'Available Grading Levels'])
        for pg in PositionGroup.objects.filter(is_active=True).order_by('hierarchy_level'):
            levels = ', '.join([level['code'] for level in pg.get_grading_levels()])
            pg_sheet.append([pg.get_name_display(), levels])
        
        # Line Managers sheet
        lm_sheet = workbook.create_sheet(title="Line Managers")
        lm_sheet.append(['Employee ID', 'Name', 'Position'])
        for manager in Employee.objects.filter(
            status__affects_headcount=True,
            position_group__hierarchy_level__lte=4,
            is_deleted=False
        ).order_by('employee_id'):
            lm_sheet.append([manager.employee_id, manager.full_name, manager.job_title])
        
        # Other options sheet
        options_sheet = workbook.create_sheet(title="Options")
        options_sheet.append(['Gender Options'])
        options_sheet.append(['MALE'])
        options_sheet.append(['FEMALE'])
        options_sheet.append([''])
        options_sheet.append(['Contract Duration Options'])
        
        # Get contract duration choices properly
        try:
            contract_configs = ContractTypeConfig.objects.filter(is_active=True).order_by('contract_type')
            if contract_configs.exists():
                for config in contract_configs:
                    options_sheet.append([config.contract_type])
            else:
                default_durations = ['3_MONTHS', '6_MONTHS', '1_YEAR', '2_YEARS', '3_YEARS', 'PERMANENT']
                for duration in default_durations:
                    options_sheet.append([duration])
        except Exception as e:
            logger.error(f"Error getting contract durations: {e}")
            default_durations = ['3_MONTHS', '6_MONTHS', '1_YEAR', '2_YEARS', '3_YEARS', 'PERMANENT']
            for duration in default_durations:
                options_sheet.append([duration])
        
        options_sheet.append([''])
        options_sheet.append(['Boolean Options'])
        options_sheet.append(['TRUE'])
        options_sheet.append(['FALSE'])
    
    
    def _add_data_validations(self, worksheet):
        """Add data validation to template"""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Gender validation (column F)
        gender_validation = DataValidation(
            type="list",
            formula1='"MALE,FEMALE"',
            showDropDown=True
        )
        gender_validation.add("F3:F1000")
        worksheet.add_data_validation(gender_validation)
        
        # Business Function validation (column K)
        bf_validation = DataValidation(
            type="list",
            formula1="'Business Functions'!A2:A100",
            showDropDown=True
        )
        bf_validation.add("K3:K1000")
        worksheet.add_data_validation(bf_validation)
        
        # Job Function validation (column N)
        jf_validation = DataValidation(
            type="list",
            formula1="'Job Functions'!A2:A100",
            showDropDown=True
        )
        jf_validation.add("N3:N1000")
        worksheet.add_data_validation(jf_validation)
        
        # Position Group validation (column P)
        pg_validation = DataValidation(
            type="list",
            formula1="'Position Groups'!A2:A100",
            showDropDown=True
        )
        pg_validation.add("P3:P1000")
        worksheet.add_data_validation(pg_validation)
        
        # Contract Duration validation (column S)
        contract_validation = DataValidation(
            type="list",
            formula1='"3_MONTHS,6_MONTHS,1_YEAR,2_YEARS,3_YEARS,PERMANENT"',
            showDropDown=True
        )
        contract_validation.add("S3:S1000")
        worksheet.add_data_validation(contract_validation)
        
        # Boolean validation for Org Chart visibility (column V)
        bool_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            showDropDown=True
        )
        bool_validation.add("V3:V1000")
        worksheet.add_data_validation(bool_validation)
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet to the workbook"""
        from openpyxl.styles import Font, PatternFill
        
        instructions_sheet = workbook.create_sheet(title="Instructions")
        
        instructions = [
            ["BULK EMPLOYEE CREATION TEMPLATE INSTRUCTIONS"],
            [""],
            ["REQUIRED FIELDS (marked with *)"],
            ["• Employee ID: Unique identifier (e.g., HC001)"],
            ["• First Name: Employee's first name"],
            ["• Last Name: Employee's last name"],
            ["• Email: Unique email address"],
            ["• Business Function: Must match exactly from dropdown"],
            ["• Department: Must exist under selected Business Function"],
            ["• Job Function: Must match exactly from dropdown"],
            ["• Job Title: Position title"],
            ["• Position Group: Must match exactly from dropdown"],
            ["• Start Date: Format YYYY-MM-DD (e.g., 2024-01-15)"],
            ["• Contract Duration: Select from dropdown"],
            [""],
            ["OPTIONAL FIELDS"],
            ["• Date of Birth: Format YYYY-MM-DD"],
            ["• Gender: MALE or FEMALE"],
            ["• Father Name: Father's name (optional)"],
            ["• Unit: Must exist under selected Department"],
            ["• Grading Level: Must be valid for Position Group (see Position Groups sheet)"],
            ["• Contract Start Date: If different from Start Date"],
            ["• Line Manager Employee ID: Must be existing employee ID (see Line Managers sheet)"],
            ["• Is Visible in Org Chart: TRUE or FALSE (default: TRUE)"],
            ["• Tag Names: Comma separated, format TYPE:Name (e.g., SKILL:Python,STATUS:New)"],
            [""],
            ["VALIDATION RULES"],
            ["• Employee IDs must be unique"],
            ["• Email addresses must be unique"],
            ["• Departments must belong to selected Business Function"],
            ["• Units must belong to selected Department"],
            ["• Grading Levels must be valid for Position Group"],
            ["• Line Manager must be existing employee"],
            ["• Dates must be in YYYY-MM-DD format"],
            [""],
            ["NOTES"],
            ["• Remove the sample data row before uploading"],
            ["• Check the reference sheets for valid values"],
            ["• Ensure all required fields are filled"],
            ["• Date format must be YYYY-MM-DD"],
            ["• Maximum 1000 employees per upload"],
            ["• Father Name is optional but can be useful for identification"]
        ]
        
        for row in instructions:
            instructions_sheet.append(row)
        
        # Style the title
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        instructions_sheet['A1'].font = title_font
        instructions_sheet['A1'].fill = title_fill
        
        # Auto-adjust column width
        instructions_sheet.column_dimensions['A'].width = 80
    def _validate_and_prepare_employee_data(self, row, business_functions, departments, 
                                      units, job_functions, position_groups, 
                                      employee_lookup, tags_lookup, default_status, row_num):
        """Validate and prepare employee data from Excel row"""
        data = {}
        errors = []
        
        # Required fields validation
        required_fields = {
            'Employee ID*': 'employee_id',
            'First Name*': 'first_name',
            'Last Name*': 'last_name',
            'Email*': 'email',
            'Business Function*': 'business_function_name',
            'Department*': 'department_name',
            'Job Function*': 'job_function_name',
            'Job Title*': 'job_title',
            'Position Group*': 'position_group_name',
            'Start Date*': 'start_date',
            'Contract Duration*': 'contract_duration'
        }
        
        for excel_field, data_field in required_fields.items():
            value = row.get(excel_field)
            if pd.isna(value) or not str(value).strip():
                errors.append(f"Missing required field: {excel_field}")
            else:
                data[data_field] = str(value).strip()
        
        if errors:
            return {'error': f"Row {row_num}: {'; '.join(errors)}"}
        
        # Validate unique fields
        if Employee.objects.filter(employee_id=data['employee_id']).exists():
            errors.append(f"Employee ID {data['employee_id']} already exists")
        
        if User.objects.filter(email=data['email']).exists():
            errors.append(f"Email {data['email']} already exists")
        
        # Validate business structure
        business_function = business_functions.get(data['business_function_name'])
        if not business_function:
            errors.append(f"Invalid Business Function: {data['business_function_name']}")
        else:
            data['business_function'] = business_function
            
            # Validate department
            dept_key = f"{data['business_function_name']}|{data['department_name']}"
            department = departments.get(dept_key)
            if not department:
                errors.append(f"Invalid Department: {data['department_name']} for Business Function: {data['business_function_name']}")
            else:
                data['department'] = department
                
                # Validate unit (optional)
                unit_name = row.get('Unit')
                if not pd.isna(unit_name) and str(unit_name).strip():
                    unit_key = f"{data['department_name']}|{str(unit_name).strip()}"
                    unit = units.get(unit_key)
                    if not unit:
                        errors.append(f"Invalid Unit: {unit_name} for Department: {data['department_name']}")
                    else:
                        data['unit'] = unit
        
        # Validate job function
        job_function = job_functions.get(data['job_function_name'])
        if not job_function:
            errors.append(f"Invalid Job Function: {data['job_function_name']}")
        else:
            data['job_function'] = job_function
        
        # Validate position group
        position_group = position_groups.get(data['position_group_name'])
        if not position_group:
            errors.append(f"Invalid Position Group: {data['position_group_name']}")
        else:
            data['position_group'] = position_group
            
            # Validate grading level
            grading_level = row.get('Grading Level')
            if not pd.isna(grading_level) and str(grading_level).strip():
                grading_level = str(grading_level).strip()
                valid_levels = [level['code'] for level in position_group.get_grading_levels()]
                if grading_level not in valid_levels:
                    errors.append(f"Invalid Grading Level: {grading_level} for Position Group: {data['position_group_name']}")
                else:
                    data['grading_level'] = grading_level
            else:
                # Default to median
                data['grading_level'] = f"{position_group.grading_shorthand}_M"
        
        # Validate dates
        try:
            start_date = pd.to_datetime(data['start_date']).date()
            data['start_date'] = start_date
        except:
            errors.append(f"Invalid Start Date format: {data['start_date']} (use YYYY-MM-DD)")
        
        contract_start_date = row.get('Contract Start Date')
        if not pd.isna(contract_start_date):
            try:
                data['contract_start_date'] = pd.to_datetime(contract_start_date).date()
            except:
                errors.append(f"Invalid Contract Start Date format: {contract_start_date} (use YYYY-MM-DD)")
        else:
            data['contract_start_date'] = data.get('start_date')
        
        # FIXED: Validate contract duration properly
        contract_duration = data['contract_duration']
        try:
            # Check if it exists in ContractTypeConfig
            if not ContractTypeConfig.objects.filter(contract_type=contract_duration, is_active=True).exists():
                # Get available options for error message
                available_durations = list(ContractTypeConfig.objects.filter(is_active=True).values_list('contract_type', flat=True))
                if not available_durations:
                    # Fallback to default options
                    available_durations = ['3_MONTHS', '6_MONTHS', '1_YEAR', '2_YEARS', '3_YEARS', 'PERMANENT']
                errors.append(f"Invalid Contract Duration: {contract_duration}. Available options: {', '.join(available_durations)}")
        except Exception as e:
            logger.error(f"Error validating contract duration: {e}")
            # Fallback validation
            default_durations = ['3_MONTHS', '6_MONTHS', '1_YEAR', '2_YEARS', '3_YEARS', 'PERMANENT']
            if contract_duration not in default_durations:
                errors.append(f"Invalid Contract Duration: {contract_duration}. Available options: {', '.join(default_durations)}")
        
        # Validate line manager (optional) - ENHANCED
        line_manager_id = row.get('Line Manager Employee ID')
        if not pd.isna(line_manager_id) and str(line_manager_id).strip():
            line_manager = employee_lookup.get(str(line_manager_id).strip())
            if not line_manager:
                errors.append(f"Line Manager not found: {line_manager_id}")
            else:
                data['line_manager'] = line_manager
        
        # Process optional fields
        optional_fields = {
            'Date of Birth': 'date_of_birth',
            'Gender': 'gender',
            'Father Name': 'father_name',  # NEW FIELD
            'Address': 'address',
            'Phone': 'phone',
            'Emergency Contact': 'emergency_contact',
            'Notes': 'notes'
        }
        
        for excel_field, data_field in optional_fields.items():
            value = row.get(excel_field)
            if not pd.isna(value) and str(value).strip():
                if data_field == 'date_of_birth':
                    try:
                        data[data_field] = pd.to_datetime(value).date()
                    except:
                        errors.append(f"Invalid Date of Birth format: {value} (use YYYY-MM-DD)")
                elif data_field == 'gender':
                    gender_value = str(value).strip().upper()
                    if gender_value in ['MALE', 'FEMALE']:
                        data[data_field] = gender_value
                    else:
                        errors.append(f"Invalid Gender: {value} (use MALE or FEMALE)")
                else:
                    data[data_field] = str(value).strip()
        
        # Process org chart visibility
        org_chart_visible = row.get('Is Visible in Org Chart')
        if not pd.isna(org_chart_visible):
            org_chart_str = str(org_chart_visible).strip().upper()
            data['is_visible_in_org_chart'] = org_chart_str == 'TRUE'
        else:
            data['is_visible_in_org_chart'] = True
        
        # Process tags
        tag_names = row.get('Tag Names (comma separated)')
        if not pd.isna(tag_names) and str(tag_names).strip():
            tags = []
            for tag_spec in str(tag_names).split(','):
                tag_spec = tag_spec.strip()
                if ':' in tag_spec:
                    tag_name = tag_spec.split(':', 1)
                 
                    tag_name = tag_name.strip()
                    
                    # Get or create tag
                    tag, created = EmployeeTag.objects.get_or_create(
                        name=tag_name,
                        defaults={
                          
                            'is_active': True
                        }
                    )
                    tags.append(tag)
                else:
                    # Simple tag name without type
                    tag, created = EmployeeTag.objects.get_or_create(
                        name=tag_spec,
                        defaults={ 'is_active': True}
                    )
                    tags.append(tag)
            data['tags'] = tags
        
        # Set default status
        data['status'] = default_status
        
        if errors:
            return {'error': f"Row {row_num}: {'; '.join(errors)}"}
        
        return data
    
    @action(detail=False, methods=['post'])
    def export_selected(self, request):
        """COMPLETELY FIXED: Export selected employees to Excel or CSV with proper field handling"""
        try:
            # Extract data from request
            export_format = request.data.get('export_format', 'excel')
            employee_ids = request.data.get('employee_ids', [])
            include_fields = request.data.get('include_fields', None)
            
            
            # Build queryset
            if employee_ids:
                # Selected employees export
                queryset = Employee.objects.filter(id__in=employee_ids)
                
            else:
                # Filtered or all employees export
                queryset = self.get_queryset()
                
                # Apply filtering from query parameters
                employee_filter = ComprehensiveEmployeeFilter(queryset, request.query_params)
                queryset = employee_filter.filter()
                
            
            # Apply sorting
            sort_params = request.query_params.get('ordering', '').split(',')
            sort_params = [param.strip() for param in sort_params if param.strip()]
            if sort_params:
                employee_sorter = AdvancedEmployeeSorter(queryset, sort_params)
                queryset = employee_sorter.sort()
            
            # COMPLETELY FIXED: Enhanced field mapping for export
            complete_field_mappings = {
                # Basic Information
                'employee_id': 'Employee ID',
                'name': 'Full Name',
                'email': 'Email',
                'father_name': 'Father Name',
                'date_of_birth': 'Date of Birth',
                'gender': 'Gender',
                'phone': 'Phone',
                'address': 'Address',
                'emergency_contact': 'Emergency Contact',
                
                # Job Information
                'job_title': 'Job Title',
                'business_function_name': 'Business Function',
                'business_function_code': 'Business Function Code',
                'business_function_id': 'Business Function ID',
                'department_name': 'Department',
                'department_id': 'Department ID',
                'unit_name': 'Unit',
                'unit_id': 'Unit ID',
                'job_function_name': 'Job Function',
                'job_function_id': 'Job Function ID',
                
                # Position & Grading
                'position_group_name': 'Position Group',
                'position_group_level': 'Position Level',
                'position_group_id': 'Position Group ID',
                'grading_level': 'Grade Level',
                
                # Management
                'line_manager_name': 'Line Manager',
                'line_manager_hc_number': 'Manager Employee ID',
                'direct_reports_count': 'Direct Reports Count',
                
                # Contract & Employment
                'contract_duration': 'Contract Duration',
                'contract_duration_display': 'Contract Duration Display',
                'contract_start_date': 'Contract Start Date',
                'contract_end_date': 'Contract End Date',
                'contract_extensions': 'Contract Extensions',
                'last_extension_date': 'Last Extension Date',
                'start_date': 'Start Date',
                'end_date': 'End Date',
                'years_of_service': 'Years of Service',
                
                # Status
                'status_name': 'Employment Status',
                'status_color': 'Status Color',
                'current_status_display': 'Current Status Display',
                'status_needs_update': 'Status Needs Update',
                'is_visible_in_org_chart': 'Visible in Org Chart',
                
                # Tags
                'tag_names': 'Tags',
                
                # Dates & Metadata
                'created_at': 'Created Date',
                'updated_at': 'Last Updated',
                'is_deleted': 'Is Deleted',
                
                # Additional Fields
                'documents_count': 'Documents Count',
                'activities_count': 'Activities Count',
                'profile_image_url': 'Profile Image URL'
            }
            
            # Determine fields to export
            if include_fields and isinstance(include_fields, list) and len(include_fields) > 0:
                # Use specified fields
                fields_to_include = include_fields
                
            else:
                # Use default essential fields
                fields_to_include = [
                    'employee_id', 'name', 'email', 'job_title', 'business_function_name',
                    'department_name', 'unit_name', 'position_group_name', 'grading_level',
                    'status_name', 'line_manager_name', 'start_date', 'contract_duration_display',
                    'phone', 'father_name', 'years_of_service'
                ]
               
            
            # Filter out invalid fields and log which ones are valid
            valid_fields = []
            invalid_fields = []
            
            for field in fields_to_include:
                if field in complete_field_mappings:
                    valid_fields.append(field)
                else:
                    invalid_fields.append(field)
            
            if invalid_fields:
                logger.warning(f"⚠️ FIXED: Invalid fields ignored: {invalid_fields}")
            
            if not valid_fields:
                # Fallback to basic fields if no valid fields
                valid_fields = ['employee_id', 'name', 'email', 'job_title', 'department_name']
                logger.warning("⚠️ FIXED: No valid fields, using fallback basic fields")
            
            
            
            # Export based on format
            if export_format == 'csv':
                return self._export_to_csv_fixed(queryset, valid_fields, complete_field_mappings)
            else:
                return self._export_to_excel_fixed(queryset, valid_fields, complete_field_mappings)
                
        except Exception as e:
            logger.error(f"❌ FIXED Export failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _export_to_excel_fixed(self, queryset, fields, field_mappings):
        """COMPLETELY FIXED: Export employees to Excel with proper field handling"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees Export"
        
        # Write headers using field mappings
        headers = [field_mappings.get(field, field.replace('_', ' ').title()) for field in fields]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        serializer = EmployeeListSerializer(queryset, many=True, context={'request': self.request})
        
        for employee_data in serializer.data:
            row_data = []
            for field in fields:
                value = employee_data.get(field, '')
                
                # Handle special field processing
                if field == 'tag_names':
                    # Convert tag objects to comma-separated names
                    if isinstance(value, list):
                        tag_names = []
                        for tag in value:
                            if isinstance(tag, dict) and 'name' in tag:
                                tag_names.append(tag['name'])
                            elif isinstance(tag, str):
                                tag_names.append(tag)
                        value = ', '.join(tag_names)
                    elif not value:
                        value = ''
                elif field == 'status_needs_update':
                    # Convert boolean to Yes/No
                    value = 'Yes' if value else 'No'
                elif field == 'is_visible_in_org_chart':
                    # Convert boolean to Yes/No
                    value = 'Yes' if value else 'No'
                elif field == 'is_deleted':
                    # Convert boolean to Yes/No
                    value = 'Yes' if value else 'No'
                elif value is None:
                    value = ''
                
                row_data.append(str(value))
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Prepare response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employees_export_{date.today()}.xlsx"'
        
       
        return response
    
    def _export_to_csv_fixed(self, queryset, fields, field_mappings):
        """COMPLETELY FIXED: Export employees to CSV with proper field handling"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="employees_export_{date.today()}.csv"'
        
        # Add BOM for proper UTF-8 handling in Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Write headers using field mappings
        headers = [field_mappings.get(field, field.replace('_', ' ').title()) for field in fields]
        writer.writerow(headers)
        
        # Write data
        serializer = EmployeeListSerializer(queryset, many=True, context={'request': self.request})
        
        for employee_data in serializer.data:
            row_data = []
            for field in fields:
                value = employee_data.get(field, '')
                
                # Handle special field processing (same as Excel)
                if field == 'tag_names':
                    if isinstance(value, list):
                        tag_names = []
                        for tag in value:
                            if isinstance(tag, dict) and 'name' in tag:
                                tag_names.append(tag['name'])
                            elif isinstance(tag, str):
                                tag_names.append(tag)
                        value = ', '.join(tag_names)
                    elif not value:
                        value = ''
                elif field == 'status_needs_update':
                    value = 'Yes' if value else 'No'
                elif field == 'is_visible_in_org_chart':
                    value = 'Yes' if value else 'No'
                elif field == 'is_deleted':
                    value = 'Yes' if value else 'No'
                elif value is None:
                    value = ''
                
                row_data.append(str(value))
            writer.writerow(row_data)
        
        
        return response
    
    def _process_bulk_employee_data_from_excel(self, df, user):
        """Excel data-sını process et və employee-lar yarat"""
        results = {
            'total_rows': len(df),
            'successful': 0,
            'failed': 0,
            'errors': [],
            'created_employees': []
        }
        
        try:
           
            
            # Convert all columns to string to avoid Series object issues
            df_str = df.astype(str)
            
            # Enhanced column mappings - more flexible matching
            column_mappings = {
                'employee_id': ['Employee ID (optional - auto-generated)', 'Employee ID', 'employee_id'],  # ✅ UPDATED
                'first_name': ['First Name*', 'First Name', 'first_name'],
                'last_name': ['Last Name*', 'Last Name', 'last_name'],
                'email': ['Email*', 'Email', 'email'],
                'date_of_birth': ['Date of Birth', 'date_of_birth'],
                'gender': ['Gender', 'gender'],
                'father_name': ['Father Name', 'father_name'],
                'phone': ['Phone', 'phone'],
                'address': ['Address', 'address'],
                'emergency_contact': ['Emergency Contact', 'emergency_contact'],
                'business_function': ['Business Function*', 'Business Function', 'business_function'],
                'department': ['Department*', 'Department', 'department'],
                'unit': ['Unit', 'unit'],
                'job_function': ['Job Function*', 'Job Function', 'job_function'],
                'job_title': ['Job Title*', 'Job Title', 'job_title'],
                'position_group': ['Position Group*', 'Position Group', 'position_group'],
                'grading_level': ['Grading Level', 'grading_level'],
                'start_date': ['Start Date*', 'Start Date', 'start_date'],
                'contract_duration': ['Contract Duration*', 'Contract Duration', 'contract_duration'],
                'contract_start_date': ['Contract Start Date', 'contract_start_date'],
                'line_manager_id': ['Line Manager Employee ID', 'Line Manager ID', 'line_manager_id'],
                'is_visible_in_org_chart': ['Is Visible in Org Chart', 'Org Chart Visible'],
                'tags': ['Tag Names (comma separated)', 'Tags', 'tags'],
                'notes': ['Notes', 'notes']
            }
            
            # Find actual column names - exact matching first
            actual_columns = {}
            df_columns = [str(col).strip() for col in df_str.columns]
            
            
            
            # Map columns with exact matching
            for field, possible_names in column_mappings.items():
                found_column = None
                for possible_name in possible_names:
                    for df_col in df_columns:
                        if df_col.strip() == possible_name.strip():
                            found_column = df_col
                            break
                    if found_column:
                        break
                
                if found_column:
                    actual_columns[field] = found_column
               
            
            
            
            # ✅ UPDATED: Check required fields - employee_id ARTIQ REQUIRED DEYIL!
            required_fields = ['first_name', 'last_name', 'email', 
                              'business_function', 'department', 'job_function', 
                              'job_title', 'position_group', 'start_date', 'contract_duration']
            
            missing_required = []
            for req_field in required_fields:
                if req_field not in actual_columns:
                    missing_required.append(req_field)
            
            if missing_required:
                error_msg = f"Missing required columns: {', '.join(missing_required)}"
                logger.error(error_msg)
                results['errors'].append(error_msg)
                results['failed'] = len(df_str)
                return results
            
            # ✅ UPDATED: Remove sample data row (daha flexible)
            df_clean = df_str.copy()
            
            # Remove rows with sample data if employee_id column exists
            if 'employee_id' in actual_columns:
                employee_id_col = actual_columns['employee_id']
                sample_ids = ['HC001', 'HC002', 'EMP001', 'TEST001']
                
                for sample_id in sample_ids:
                    df_clean = df_clean[df_clean[employee_id_col].str.strip() != sample_id]
            
            # Remove rows with sample names
            first_name_col = actual_columns['first_name']
            sample_names = ['John', 'Jane', 'Test', 'Sample']
            for sample_name in sample_names:
                df_clean = df_clean[df_clean[first_name_col].str.strip() != sample_name]
            
            # Remove completely empty rows
            df_clean = df_clean.dropna(how='all')
            
            # ✅ CRITICAL FIX: Employee ID validation - boş ola bilər
            # YALNIZ first_name və email-i yoxlayırıq
            df_clean = df_clean[df_clean[first_name_col].notna()]
            df_clean = df_clean[df_clean[first_name_col].str.strip() != '']
            df_clean = df_clean[df_clean[first_name_col].str.strip() != 'nan']
            
            
            
            if df_clean.empty:
                results['errors'].append("No valid data rows found. Please add employee data after removing sample rows.")
                results['failed'] = len(df_str)
                return results
            
            # Update total_rows
            results['total_rows'] = len(df_clean)
            
            # Prepare lookup dictionaries
            business_functions = {}
            for bf in BusinessFunction.objects.filter(is_active=True):
                business_functions[bf.name.lower()] = bf
            
            departments = {}
            for dept in Department.objects.select_related('business_function').filter(is_active=True):
                departments[dept.name.lower()] = dept
            
            job_functions = {}
            for jf in JobFunction.objects.filter(is_active=True):
                job_functions[jf.name.lower()] = jf
            
            position_groups = {}
            for pg in PositionGroup.objects.filter(is_active=True):
                position_groups[pg.get_name_display().lower()] = pg
            
            employee_lookup = {}
            for emp in Employee.objects.all():
                employee_lookup[emp.employee_id] = emp
            
            # Get default status
            default_status = EmployeeStatus.objects.filter(is_default_for_new_employees=True).first()
            if not default_status:
                default_status = EmployeeStatus.objects.filter(is_active=True).first()
            
            if not default_status:
                results['errors'].append("No employee status found. Please create default status first.")
                results['failed'] = len(df_clean)
                return results
            
            # ✅ HƏR ROW ÜÇÜN AYRI TRANSACTION - BÖYÜK TRANSACTION YOX!
            for index, row in df_clean.iterrows():
                try:
                    # ✅ HƏR EMPLOYEE ÜÇÜN AYRI TRANSACTION
                    with transaction.atomic():
                        # Extract required fields with safe string conversion
                        def safe_get(col_name, default=''):
                            """Safely get value from row, handling lists and various data types"""
                            if col_name not in actual_columns:
                                return default
                            
                            value = row.get(actual_columns[col_name], default)
                            
                            # Handle None values
                            if value is None:
                                return default
                            
                            # Handle list values (pandas sometimes returns lists for cells)
                            if isinstance(value, list):
                                if len(value) > 0:
                                    value = value[0]  # Take first element
                                else:
                                    return default
                            
                            # Handle pandas Series (can happen with duplicate columns)
                            if hasattr(value, 'iloc'):
                                try:
                                    value = value.iloc[0] if len(value) > 0 else default
                                except:
                                    value = default
                            
                            # Convert to string and check for empty/nan values
                            str_value = str(value).strip()
                            
                            # Check for various empty representations
                            if str_value.lower() in ['nan', 'none', 'nat', '', 'null']:
                                return default
                            
                            return str_value
                        
                        # ✅ CRITICAL CHANGE: Employee ID is now OPTIONAL
                        employee_id_from_excel = safe_get('employee_id')
                        
                        first_name = safe_get('first_name')
                        last_name = safe_get('last_name')
                        email = safe_get('email')
                        business_function_name = safe_get('business_function')
                        department_name = safe_get('department')
                        job_function_name = safe_get('job_function')
                        job_title = safe_get('job_title')
                        position_group_name = safe_get('position_group')
                        start_date_str = safe_get('start_date')
                        contract_duration = safe_get('contract_duration', 'PERMANENT')
                        
                       
                        
                        # ✅ UPDATED: Validate required fields - employee_id ARTIQ YOXLANMIR
                        if not all([first_name, last_name, email, business_function_name, 
                                   department_name, job_function_name, job_title, position_group_name, start_date_str]):
                            results['errors'].append(f"Row {index + 2}: Missing required data")
                            results['failed'] += 1
                            continue
                        
                        # ✅ CRITICAL: Business function validation - ALWAYS needed
                        business_function = business_functions.get(business_function_name.lower())
                        if not business_function:
                            results['errors'].append(f"Row {index + 2}: Business Function '{business_function_name}' not found")
                            results['failed'] += 1
                            continue
                        
                        # ✅ UPDATED: Employee ID handling
                        validated_employee_id = None
                        if employee_id_from_excel:
                            # User provided employee_id - check for duplicates
                            if Employee.objects.filter(employee_id=employee_id_from_excel).exists():
                                results['errors'].append(f"Row {index + 2}: Employee ID {employee_id_from_excel} already exists")
                                results['failed'] += 1
                                continue
                            validated_employee_id = employee_id_from_excel
                            
                        else:
                            # No employee_id provided - will be auto-generated by Employee.save()
                            logger.info(f"Row {index + 2}: No employee_id provided, will be auto-generated for {first_name} {last_name} in {business_function.name}")
                        
                        # Check email duplicates
                        if User.objects.filter(email=email).exists():
                            results['errors'].append(f"Row {index + 2}: Email {email} already exists")
                            results['failed'] += 1
                            continue
                        
                        # Validate department
                        department = departments.get(department_name.lower())
                        if not department:
                            results['errors'].append(f"Row {index + 2}: Department '{department_name}' not found")
                            results['failed'] += 1
                            continue
                        
                        # Validate job function
                        job_function = job_functions.get(job_function_name.lower())
                        if not job_function:
                            results['errors'].append(f"Row {index + 2}: Job Function '{job_function_name}' not found")
                            results['failed'] += 1
                            continue
                        
                        # Validate position group
                        position_group = position_groups.get(position_group_name.lower())
                        if not position_group:
                            results['errors'].append(f"Row {index + 2}: Position Group '{position_group_name}' not found")
                            results['failed'] += 1
                            continue
                        
                        # Parse start date
                        try:
                            start_date = pd.to_datetime(start_date_str).date()
                        except:
                            results['errors'].append(f"Row {index + 2}: Invalid start date '{start_date_str}'")
                            results['failed'] += 1
                            continue
                        
                        # Validate contract duration
                        try:
                            if not ContractTypeConfig.objects.filter(contract_type=contract_duration, is_active=True).exists():
                                available_durations = list(ContractTypeConfig.objects.filter(is_active=True).values_list('contract_type', flat=True))
              
                                
                                if not available_durations:
                                    available_durations = ['3_MONTHS', '6_MONTHS', '1_YEAR', '2_YEARS', '3_YEARS', 'PERMANENT']
                                
                                if contract_duration not in available_durations:
                                    results['errors'].append(f"Row {index + 2}: Invalid contract duration '{contract_duration}'. Available: {', '.join(available_durations)}")
                                    results['failed'] += 1
                                    continue
                        except Exception as e:
                            logger.error(f"Error validating contract duration for row {index + 2}: {e}")
                            default_durations = ['3_MONTHS', '6_MONTHS', '1_YEAR', '2_YEARS', '3_YEARS', 'PERMANENT']
                            if contract_duration not in default_durations:
                                contract_duration = 'PERMANENT'
                        
                        # Optional fields
                        date_of_birth = None
                        dob_str = safe_get('date_of_birth')
                        if dob_str:
                            try:
                                date_of_birth = pd.to_datetime(dob_str).date()
                            except:
                                pass
                        
                        gender = safe_get('gender').upper()
                        if gender not in ['MALE', 'FEMALE']:
                            gender = None
                        
                        father_name = safe_get('father_name')
                        
                        # ✅ PHONE FIELD VALIDATION - TRUNCATE IF TOO LONG
                        phone = safe_get('phone')
                        if phone and len(phone) > 15:
                            phone = phone[:15]
                            logger.warning(f"Row {index + 2}: Phone number truncated to 15 characters")
                        
                        address = safe_get('address')
                        emergency_contact = safe_get('emergency_contact')
                        
                        # Unit (optional)
                        unit = None
                        unit_name = safe_get('unit')
                        if unit_name:
                            unit = Unit.objects.filter(name__iexact=unit_name, department=department).first()
                        
                        # Grading level
                        grading_level = safe_get('grading_level')
                        if not grading_level:
                            grading_level = f"{position_group.grading_shorthand}_M"
                        
                        # Contract start date
                        contract_start_date = start_date
                        csd_str = safe_get('contract_start_date')
                        if csd_str:
                            try:
                                contract_start_date = pd.to_datetime(csd_str).date()
                            except:
                                pass
                        
                        # Line manager
                        line_manager = None
                        line_manager_id = safe_get('line_manager_id')
                        if line_manager_id:
                            line_manager = employee_lookup.get(line_manager_id)
                        
                        # Org chart visibility
                        is_visible_str = safe_get('is_visible_in_org_chart', 'TRUE').upper()
                        is_visible_in_org_chart = is_visible_str in ['TRUE', '1', 'YES']
                        
                        notes = safe_get('notes')
                        
                        # ✅ CRITICAL: Create employee - employee_id conditional
                        employee_data = {
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'date_of_birth': date_of_birth,
                            'gender': gender,
                            'father_name': father_name,
                            'address': address,
                            'phone': phone,
                            'emergency_contact': emergency_contact,
                            'business_function': business_function,
                            'department': department,
                            'unit': unit,
                            'job_function': job_function,
                            'job_title': job_title,
                            'position_group': position_group,
                            'grading_level': grading_level,
                            'start_date': start_date,
                            'contract_duration': contract_duration,
                            'contract_start_date': contract_start_date,
                            'line_manager': line_manager,
                            'status': default_status,
                            'is_visible_in_org_chart': is_visible_in_org_chart,
                            'notes': notes,
                            'created_by': user
                        }
                        
                        # ✅ ONLY add employee_id if user provided it
                        if validated_employee_id:
                            employee_data['employee_id'] = validated_employee_id
                        
                        # Create employee - save() will auto-generate employee_id if not provided
                        employee = Employee.objects.create(**employee_data)
                        
                        # Process tags
                        tags_str = safe_get('tags')
                        if tags_str:
                            tags = []
                            for tag_spec in tags_str.split(','):
                                tag_spec = tag_spec.strip()
                                if ':' in tag_spec:
                                    tag_parts = tag_spec.split(':', 1)
                                    if len(tag_parts) >= 2:
                                        tag_name = tag_parts[1].strip()
                                    else:
                                        tag_name = tag_spec.strip()
                                else:
                                    tag_name = tag_spec.strip()
                                
                                if tag_name:
                                    tag, created = EmployeeTag.objects.get_or_create(
                                        name=tag_name,
                                        defaults={'is_active': True}
                                    )
                                    tags.append(tag)
                            
                            if tags:
                                employee.tags.set(tags)
                        
                        # Log activity
                        EmployeeActivity.objects.create(
                            employee=employee,
                            activity_type='BULK_CREATED',
                            description=f"Employee {employee.full_name} created via bulk upload" + 
                                      (f" with provided ID {validated_employee_id}" if validated_employee_id else " with auto-generated ID"),
                            performed_by=user,
                            metadata={
                                'bulk_creation': True, 
                                'row_number': index + 2,
                                'employee_id_auto_generated': not bool(validated_employee_id)
                            }
                        )
                        
                        results['successful'] += 1
                        results['created_employees'].append({
                            'employee_id': employee.employee_id,
                            'name': employee.full_name,
                            'email': employee.email,
                            'id_auto_generated': not bool(validated_employee_id)
                        })
                        
                        
                    
                except Exception as e:
                    # ✅ Bu row fail oldu, amma digərləri davam edər
                    error_msg = f"Row {index + 2}: {str(e)}"
                    results['errors'].append(error_msg)
                    results['failed'] += 1
                    logger.error(f"❌ Error creating employee from row {index + 2}: {e}")
                    logger.error(f"Traceback: {traceback.format_exc()}")
                    continue  # Növbəti row-a keç
            
            
            return results
            
        except Exception as e:
            logger.error(f"❌ Bulk processing failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            results['errors'].append(f"Processing failed: {str(e)}")
            results['failed'] = results['total_rows']
            return results
    
    # views.py - EmployeeViewSet içində job description endpointləri

    @swagger_auto_schema(
        method='get',
        operation_description="Get job descriptions assigned to this employee",
        responses={
            200: openapi.Response(
                description="Job descriptions retrieved successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'employee': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'job_descriptions': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                        'pending_approval_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'total_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                    }
                )
            ),
            403: "Access denied - not authorized to view this employee's job descriptions",
            404: "Employee not found"
        }
    )
    @action(detail=True, methods=['get'])
    def job_descriptions(self, request, pk=None):
        """✅ FIXED: Get job descriptions with proper access control"""
        
        # Get the employee whose job descriptions we want to view
        try:
            employee = Employee.objects.select_related('line_manager', 'user').get(pk=pk)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # ✅ CRITICAL: Check access permissions
        access = get_headcount_access(request.user)
        
        # Check if user can view this employee's job descriptions
        can_view = False
        access_reason = None
        
        # 1. Admin can view all
        if access['can_view_all']:
            can_view = True
            access_reason = 'admin'
            
        
        # 2. User viewing their own job descriptions
        elif employee.user and employee.user.id == request.user.id:
            can_view = True
            access_reason = 'own_profile'
            
        
        # 3. Manager viewing direct report's job descriptions
        elif access['is_manager'] and access['accessible_employee_ids']:
            if employee.id in access['accessible_employee_ids']:
                can_view = True
                access_reason = 'manager'
                
        # ❌ Access denied
        if not can_view:
            logger.warning(f"⚠️ User {request.user.username} attempted unauthorized access to employee {employee.employee_id}'s job descriptions")
            return Response(
                {
                    'error': 'Access Denied',
                    'message': 'You do not have permission to view this employee\'s job descriptions.',
                    'detail': 'You can only view your own job descriptions or those of your direct reports.',
                },
                status=status.HTTP_403_FORBIDDEN
            )
        
        # ✅ Get job descriptions through assignments
        job_description_assignments = JobDescriptionAssignment.objects.filter(
            employee=employee,
            is_active=True
        ).select_related(
            'job_description__business_function',
            'job_description__department', 
            'job_description__job_function',
            'job_description__position_group',
            'reports_to'
        ).order_by('-created_at')
        
        # Serialize assignments with job description details
        serializer = EmployeeJobDescriptionSerializer(
            job_description_assignments,
            many=True,
            context={'request': request}
        )
        
        return Response({
            'employee': {
                'id': employee.id,
                'name': employee.full_name,
                'employee_id': employee.employee_id,
                'job_title': employee.job_title
            },
            'job_descriptions': serializer.data,
            'pending_approval_count': job_description_assignments.filter(
                status__in=['PENDING_EMPLOYEE', 'PENDING_LINE_MANAGER']
            ).count(),
            'total_count': job_description_assignments.count(),
            'access_context': {
                'accessed_as': access_reason,
                'can_edit': access_reason in ['admin', 'own_profile'],
                'is_manager_view': access_reason == 'manager'
            }
        })
    
    @action(detail=True, methods=['get'])
    def team_job_descriptions(self, request, pk=None):
        """✅ FIXED: Get job descriptions for manager's direct reports with access control"""
        
        # Get the manager
        try:
            manager = Employee.objects.select_related('user').get(pk=pk)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Manager not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # ✅ CRITICAL: Check access permissions
        access = get_headcount_access(request.user)
        
        # Check if user can view this manager's team job descriptions
        can_view = False
        access_reason = None
        
        # 1. Admin can view all
        if access['can_view_all']:
            can_view = True
            access_reason = 'admin'
            
        
        # 2. Manager viewing their own team's job descriptions
        elif manager.user and manager.user.id == request.user.id:
            can_view = True
            access_reason = 'own_team'
            
        
        # 3. Upper manager viewing lower manager's team (manager's manager)
        elif access['is_manager'] and access['accessible_employee_ids']:
            if manager.id in access['accessible_employee_ids']:
                can_view = True
                access_reason = 'upper_manager'
                
        # ❌ Access denied
        if not can_view:
            logger.warning(f"⚠️ User {request.user.username} attempted unauthorized access to manager {manager.employee_id}'s team job descriptions")
            return Response(
                {
                    'error': 'Access Denied',
                    'message': 'You do not have permission to view this team\'s job descriptions.',
                    'detail': 'You can only view job descriptions for your own team or teams you manage.',
                },
                status=status.HTTP_403_FORBIDDEN
            )
        
        # ✅ Get job description assignments for team members
        team_job_description_assignments = JobDescriptionAssignment.objects.filter(
            employee__line_manager=manager,
            employee__status__affects_headcount=True,
            employee__is_deleted=False,
            is_active=True
        ).select_related(
            'employee',
            'job_description__business_function',
            'job_description__department',
            'job_description__job_function',
            'job_description__position_group',
            'reports_to'
        ).order_by('-created_at')
        
        serializer = ManagerJobDescriptionSerializer(
            team_job_description_assignments, 
            many=True, 
            context={'request': request}
        )
        
        # Get team member count
        team_members_count = Employee.objects.filter(
            line_manager=manager, 
            status__affects_headcount=True,
            is_deleted=False
        ).count()
        
        return Response({
            'manager': {
                'id': manager.id,
                'name': manager.full_name,
                'employee_id': manager.employee_id,
                'job_title': manager.job_title
            },
            'team_job_descriptions': serializer.data,
            'pending_approval_count': team_job_description_assignments.filter(
                status='PENDING_LINE_MANAGER'
            ).count(),
            'total_count': team_job_description_assignments.count(),
            'total_team_members': team_members_count,
            'access_context': {
                'accessed_as': access_reason,
                'can_approve': access_reason in ['admin', 'own_team', 'upper_manager'],
                'is_admin_view': access_reason == 'admin'
            }
        })
   
    @swagger_auto_schema(
        method='post',
        operation_description="Add tag to single employee",
        request_body=SingleEmployeeTagUpdateSerializer,
        responses={200: "Tag added successfully", 400: "Bad request"}
    )
    @action(detail=False, methods=['post'], url_path='add-tag')
    def add_tag_to_employee(self, request):
        """Add tag to single employee using employee ID"""
        serializer = SingleEmployeeTagUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_id = serializer.validated_data['employee_id']
        tag_id = serializer.validated_data['tag_id']
        
        try:
            employee = Employee.objects.get(id=employee_id)
            tag = EmployeeTag.objects.get(id=tag_id)
            
            if employee.add_tag(tag, request.user):
                return Response({
                    'success': True,
                    'message': f'Tag "{tag.name}" added to {employee.full_name}',
                    'employee_id': employee.id,
                    'employee_name': employee.full_name,
                    'tag_name': tag.name
                })
            else:
                return Response({
                    'success': False,
                    'message': f'Tag "{tag.name}" already exists on {employee.full_name}',
                    'employee_id': employee.id,
                    'employee_name': employee.full_name,
                    'tag_name': tag.name
                })
        except (Employee.DoesNotExist, EmployeeTag.DoesNotExist) as e:
            return Response({'error': str(e)}, status=status.HTTP_404_NOT_FOUND)
       
    @swagger_auto_schema(
        method='post',
        operation_description="Remove tag from single employee",
        request_body=SingleEmployeeTagUpdateSerializer,
        responses={200: "Tag removed successfully", 400: "Bad request"}
    )
    @action(detail=False, methods=['post'], url_path='remove-tag')
    def remove_tag_from_employee(self, request):
        """Remove tag from single employee using employee ID"""
        serializer = SingleEmployeeTagUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_id = serializer.validated_data['employee_id']
        tag_id = serializer.validated_data['tag_id']
        
        try:
            employee = Employee.objects.get(id=employee_id)
            tag = EmployeeTag.objects.get(id=tag_id)
            
            if employee.remove_tag(tag, request.user):
                return Response({
                    'success': True,
                    'message': f'Tag "{tag.name}" removed from {employee.full_name}',
                    'employee_id': employee.id,
                    'employee_name': employee.full_name,
                    'tag_name': tag.name
                })
            else:
                return Response({
                    'success': False,
                    'message': f'Tag "{tag.name}" was not found on {employee.full_name}',
                    'employee_id': employee.id,
                    'employee_name': employee.full_name,
                    'tag_name': tag.name
                })
        except (Employee.DoesNotExist, EmployeeTag.DoesNotExist) as e:
            return Response({'error': str(e)}, status=status.HTTP_404_NOT_FOUND)
    

    @swagger_auto_schema(
        method='post',
        operation_description="Add tag to multiple employees and set status to INACTIVE",
        request_body=BulkEmployeeTagUpdateSerializer,
        responses={
            200: openapi.Response(
                description="Tags added and statuses updated successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'total_employees': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'added_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'already_had_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'status_changed_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'results': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                    'employee_name': openapi.Schema(type=openapi.TYPE_STRING),
                                    'status': openapi.Schema(type=openapi.TYPE_STRING),
                                    'tag_added': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    'status_changed': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    'old_status': openapi.Schema(type=openapi.TYPE_STRING),
                                    'new_status': openapi.Schema(type=openapi.TYPE_STRING)
                                }
                            )
                        )
                    }
                )
            ),
            400: "Bad request"
        }
    )
    @action(detail=False, methods=['post'], url_path='bulk-add-tag')
    def bulk_add_tag(self, request):
        """
        ✅ UPDATED: Add tag to multiple employees and automatically set status to INACTIVE
        """
        serializer = BulkEmployeeTagUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_ids = serializer.validated_data['employee_ids']
        tag_id = serializer.validated_data['tag_id']
        
        try:
            tag = EmployeeTag.objects.get(id=tag_id)
            employees = Employee.objects.filter(id__in=employee_ids)
            
            # Get INACTIVE status
            inactive_status = EmployeeStatus.objects.filter(
                status_type='INACTIVE',
                is_active=True
            ).first()
            
            if not inactive_status:
                logger.warning("INACTIVE status not found in system")
            
            added_count = 0
            already_had_count = 0
            status_changed_count = 0
            results = []
            
            with transaction.atomic():
                for employee in employees:
                    result_data = {
                        'employee_id': employee.id,
                        'employee_name': employee.full_name,
                        'employee_hc_id': employee.employee_id,
                        'tag_added': False,
                        'status_changed': False,
                        'old_status': employee.status.name if employee.status else None,
                        'new_status': None
                    }
                    
                    # Check if tag already exists
                    tag_existed = employee.tags.filter(id=tag.id).exists()
                    
                    if not tag_existed:
                        # Add tag using the updated add_tag method
                        # This will automatically change status to INACTIVE
                        tag_added = employee.add_tag(tag, request.user)
                        
                        if tag_added:
                            added_count += 1
                            result_data['tag_added'] = True
                            result_data['status'] = 'added'
                            
                            # Refresh to get updated status
                            employee.refresh_from_db()
                            result_data['new_status'] = employee.status.name if employee.status else None
                            
                            # Check if status was changed
                            if result_data['old_status'] != result_data['new_status']:
                                status_changed_count += 1
                                result_data['status_changed'] = True
                                
                                
                    else:
                        already_had_count += 1
                        result_data['status'] = 'already_had'
                        result_data['new_status'] = result_data['old_status']
                    
                    results.append(result_data)
            
            return Response({
                'success': True,
                'message': f"Tag '{tag.name}' processed for {len(employee_ids)} employees. {status_changed_count} statuses changed to INACTIVE.",
                'tag_name': tag.name,
                'total_employees': len(employee_ids),
                'added_count': added_count,
                'already_had_count': already_had_count,
                'status_changed_count': status_changed_count,
                'inactive_status_available': bool(inactive_status),
                'results': results
            })
            
        except EmployeeTag.DoesNotExist:
            return Response({'error': 'Tag not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Bulk add tag failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Bulk add tag failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 
    
    
    @action(detail=False, methods=['get'], url_path='archived-employees')
    def get_archived_employees(self, request):
        """ENHANCED: Get list of ALL archived employees with deletion type filtering"""
        try:
            archives = EmployeeArchive.objects.all().order_by('-deleted_at')
            
            # Apply filtering
            search = request.query_params.get('search')
            if search:
                archives = archives.filter(
                    Q(original_employee_id__icontains=search) |
                    Q(full_name__icontains=search) |
                    Q(email__icontains=search) |
                    Q(business_function_name__icontains=search) |
                    Q(department_name__icontains=search)
                )
            
            # NEW: Deletion type filtering
            deletion_type = request.query_params.get('deletion_type')
            if deletion_type:
                if deletion_type.lower() == 'soft':
                    archives = archives.filter(employee_still_exists=True)
                elif deletion_type.lower() == 'hard':
                    archives = archives.filter(employee_still_exists=False)
            
            # Employee still exists filter (legacy support)
            still_exists = request.query_params.get('employee_still_exists')
            if still_exists:
                if still_exists.lower() == 'true':
                    archives = archives.filter(employee_still_exists=True)
                elif still_exists.lower() == 'false':
                    archives = archives.filter(employee_still_exists=False)
            
            # Date filtering
            deleted_after = request.query_params.get('deleted_after')
            if deleted_after:
                try:
                    date_after = datetime.strptime(deleted_after, '%Y-%m-%d').date()
                    archives = archives.filter(deleted_at__date__gte=date_after)
                except ValueError:
                    pass
            
            deleted_before = request.query_params.get('deleted_before')
            if deleted_before:
                try:
                    date_before = datetime.strptime(deleted_before, '%Y-%m-%d').date()
                    archives = archives.filter(deleted_at__date__lte=date_before)
                except ValueError:
                    pass
            
            # Pagination
            page_size = int(request.query_params.get('page_size', 20))
            page = int(request.query_params.get('page', 1))
            start = (page - 1) * page_size
            end = start + page_size
            
            total_count = archives.count()
            paginated_archives = archives[start:end]
            
            archive_data = []
            for archive in paginated_archives:
                archive_info = archive.get_enhanced_deletion_summary()
                
                archive_data.append({
                    'id': archive.id,
                    'reference': archive.get_archive_reference(),
                    'original_employee_id': archive.original_employee_id,
                    'original_employee_pk': archive.original_employee_pk,
                    'full_name': archive.full_name,
                    'email': archive.email,
                    'job_title': archive.job_title,
                    'business_function_name': archive.business_function_name,
                    'department_name': archive.department_name,
                    'unit_name': archive.unit_name,
                    'start_date': archive.start_date,
                    'end_date': archive.end_date,
                    'contract_duration': archive.contract_duration,
                    'line_manager_name': archive.line_manager_name,
                    'deletion_notes': archive.deletion_notes,
                    'deleted_by': archive.deleted_by.get_full_name() if archive.deleted_by else 'System',
                    'deleted_at': archive.deleted_at,
                    'updated_at': archive.updated_at,
                    'has_complete_data': bool(archive.original_data),
                    'data_quality': archive.get_data_quality_display(),
                    'archive_version': archive.archive_version,
                    'days_since_deletion': (timezone.now() - archive.deleted_at).days if archive.deleted_at else 0,
                    
                    # NEW: Enhanced deletion type information
                    'deletion_type': archive_info['deletion_type'],
                    'deletion_type_display': archive_info['deletion_type_display'],
                    'employee_still_exists': archive.employee_still_exists,
                    'can_be_restored': archive_info['can_be_restored'],
                    'is_restorable': archive_info['is_restorable'],
                    'restoration_available': archive_info['restoration_available']
                })
            
            # Enhanced statistics with deletion type breakdown
            total_soft_deleted = archives.filter(employee_still_exists=True).count() if deletion_type != 'hard' else EmployeeArchive.objects.filter(employee_still_exists=True).count()
            total_hard_deleted = archives.filter(employee_still_exists=False).count() if deletion_type != 'soft' else EmployeeArchive.objects.filter(employee_still_exists=False).count()
            
            stats = {
                'total_archived': total_count,
                'soft_deleted_archives': total_soft_deleted,
                'hard_deleted_archives': total_hard_deleted,
                'restorable_count': archives.filter(employee_still_exists=True).count(),
                'permanent_deletions': archives.filter(employee_still_exists=False).count(),
                'by_deletion_type': {
                    'soft_delete': total_soft_deleted,
                    'hard_delete': total_hard_deleted
                },
                'recent_30_days': archives.filter(deleted_at__gte=timezone.now() - timedelta(days=30)).count(),
                'by_data_quality': {}
            }
            
            # Data quality breakdown
            quality_counts = archives.values('data_quality').annotate(count=Count('data_quality'))
            for quality_data in quality_counts:
                quality_display = dict(EmployeeArchive._meta.get_field('data_quality').choices).get(
                    quality_data['data_quality'], quality_data['data_quality']
                )
                stats['by_data_quality'][quality_display] = quality_data['count']
            
            return Response({
                'count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size,
                'results': archive_data,
                'statistics': stats,
                'filters_applied': {
                    'search': bool(search),
                    'deletion_type': deletion_type,
                    'employee_still_exists': still_exists,
                    'date_range': bool(deleted_after or deleted_before)
                }
            })
            
        except Exception as e:
            logger.error(f"Get archived employees failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to get archived employees: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
    method='post',
    operation_description="Bulk hard delete employees and create archives - NO VACANCY CREATION",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['employee_ids', 'confirm_hard_delete'],
        properties={
            'employee_ids': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(type=openapi.TYPE_INTEGER),
                description='List of employee IDs to hard delete',
                example=[1, 2, 3]
            ),
            'confirm_hard_delete': openapi.Schema(
                type=openapi.TYPE_BOOLEAN, 
                description='Confirmation flag (must be true)',
                example=True
            ),
            'notes': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Additional notes about deletion',
                example='End of contract period - bulk cleanup'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description="Bulk hard deletion completed",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                    'message': openapi.Schema(type=openapi.TYPE_STRING),
                    'summary': openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            'total_requested': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'successful': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'failed': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'archives_created': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'data_permanently_deleted': openapi.Schema(type=openapi.TYPE_BOOLEAN)
                        }
                    ),
                    'results': openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                'employee_name': openapi.Schema(type=openapi.TYPE_STRING),
                                'status': openapi.Schema(type=openapi.TYPE_STRING),
                                'archive_created': openapi.Schema(type=openapi.TYPE_OBJECT)
                            }
                        )
                    )
                }
            )
        ),
        400: "Bad request - validation errors"
    }
)
    @action(detail=False, methods=['post'], url_path='bulk-hard-delete-with-archives')
    def bulk_hard_delete_with_archives(self, request):
        """FIXED: Bulk hard delete with proper serializer validation - NO VACANCY CREATION"""
        
        # Use the dedicated serializer
        serializer = BulkHardDeleteSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_ids = serializer.validated_data['employee_ids']
        notes = serializer.validated_data.get('notes', '')
        
        employees = Employee.objects.filter(id__in=employee_ids)
        
        results = []
        archives_created = []
        total_direct_reports_updated = 0
        
        # Process employees individually
        for employee in employees:
            try:
                # Store info before deletion
                employee_info = {
                    'id': employee.id,
                    'employee_id': employee.employee_id,
                    'name': employee.full_name,
                    'email': employee.user.email if employee.user else None,
                    'direct_reports_count': employee.direct_reports.filter(is_deleted=False).count()
                }
                
                # FIXED: Hard delete and create archive - NO VACANCY CREATION
                archive = employee.hard_delete_with_archive(request.user)
                
                # Update archive with bulk deletion info
                if archive and notes:
                    archive.deletion_notes = f"{archive.deletion_notes}\n\nBulk hard deletion notes: {notes}"
                    archive.save()
                
                results.append({
                    'employee_id': employee_info['id'],
                    'original_employee_id': employee_info['employee_id'],
                    'employee_name': employee_info['name'],
                    'status': 'success',
                    'archive_created': {
                        'id': archive.id if archive else None,
                        'reference': archive.get_archive_reference() if archive else None
                    },
                    'direct_reports_updated': employee_info['direct_reports_count'],
                    'vacancy_created': None,  # FIXED: No vacancy for hard delete
                    'data_permanently_deleted': True
                })
                
                if archive:
                    archives_created.append(archive)
                total_direct_reports_updated += employee_info['direct_reports_count']
                
            except Exception as e:
                results.append({
                    'employee_id': employee.id,
                    'original_employee_id': employee.employee_id,
                    'employee_name': employee.full_name,
                    'status': 'failed',
                    'error': str(e)
                })
        
        successful_count = len([r for r in results if r['status'] == 'success'])
        failed_count = len([r for r in results if r['status'] == 'failed'])
        
        return Response({
            'success': True,
            'message': f'Bulk hard deletion completed: {successful_count} successful, {failed_count} failed',
            'summary': {
                'total_requested': len(employee_ids),
                'successful': successful_count,
                'failed': failed_count,
                'archives_created': len(archives_created),
                'total_direct_reports_updated': total_direct_reports_updated,
                'vacancies_created': 0,  # FIXED: No vacancies for hard delete
                'data_permanently_deleted': True,
                'cannot_restore': True
            },
            'results': results,
            'notes': notes,
            'deletion_type': 'bulk_hard_delete_with_archives_only'
        })
    
    @swagger_auto_schema(
    method='post',
    operation_description="Bulk restore soft-deleted employees with vacancy cleanup",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['employee_ids'],
        properties={
            'employee_ids': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(type=openapi.TYPE_INTEGER),
                description='List of soft-deleted employee IDs to restore',
                example=[1, 2, 3]
            ),
            'restore_to_active': openapi.Schema(
                type=openapi.TYPE_BOOLEAN,
                description='Set status to active after restore',
                example=False,
                default=False
            )
        }
    ),
    responses={
        200: openapi.Response(
            description="Bulk restoration completed with vacancy cleanup",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                    'message': openapi.Schema(type=openapi.TYPE_STRING),
                    'summary': openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            'total_requested': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'successful': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'failed': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'total_vacancies_removed': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'total_archives_deleted': openapi.Schema(type=openapi.TYPE_INTEGER)
                        }
                    ),
                    'results': openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                'employee_name': openapi.Schema(type=openapi.TYPE_STRING),
                                'status': openapi.Schema(type=openapi.TYPE_STRING),
                                'restored_to_active': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                'vacancies_removed': openapi.Schema(type=openapi.TYPE_INTEGER),
                                'archives_deleted': openapi.Schema(type=openapi.TYPE_INTEGER)
                            }
                        )
                    )
                }
            )
        ),
        400: "Bad request - validation errors"
    }
)
    @action(detail=False, methods=['post'], url_path='bulk-restore-employees')
    def bulk_restore_employees(self, request):
        """FIXED: Bulk restore soft-deleted employees with proper vacancy cleanup and archive deletion"""
        try:
            employee_ids = request.data.get('employee_ids', [])
            restore_to_active = request.data.get('restore_to_active', False)
            
            if not employee_ids:
                return Response(
                    {'error': 'employee_ids list is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Use all_objects to include soft-deleted employees
            employees = Employee.all_objects.filter(id__in=employee_ids, is_deleted=True)
            
            if employees.count() != len(employee_ids):
                return Response(
                    {'error': 'Some employee IDs were not found or are not deleted'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            results = []
            active_status = None
            vacancies_removed = []
            archives_deleted = []  # FIXED: Track deleted archives
            
            if restore_to_active:
                active_status = EmployeeStatus.objects.filter(status_type='ACTIVE', is_active=True).first()
            
            # Process each employee individually
            for employee in employees:
                try:
                    # Store info before restore
                    employee_info = {
                        'id': employee.id,
                        'employee_id': employee.employee_id,
                        'name': employee.full_name,
                        'was_deleted_at': employee.deleted_at
                    }
                    
                    with transaction.atomic():
                        # FIXED: Find and remove vacancies using original_employee_pk
                        related_vacancies = VacantPosition.objects.filter(
                            original_employee_pk=employee.pk,  # Use PK for exact match
                            is_filled=False
                        )
                        
                        vacancy_info = []
                        for vacancy in related_vacancies:
                            vacancy_info.append({
                                'id': vacancy.id,
                                'position_id': vacancy.position_id,
                                'job_title': vacancy.job_title
                            })
                            vacancy.delete()  # Remove the vacancy
                            vacancies_removed.append(vacancy_info[-1])
                        
                        # FIXED: Find and DELETE the soft delete archive record
                        soft_delete_archives = EmployeeArchive.objects.filter(
                            original_employee_id=employee.employee_id,
                      
                            employee_still_exists=True  # Only soft delete archives
                        ).order_by('-deleted_at')
                        
                        archive_info = []
                        for archive in soft_delete_archives:
                            archive_data = {
                                'id': archive.id,
                                'reference': archive.get_archive_reference(),
                                'deleted_at': archive.deleted_at.isoformat() if archive.deleted_at else None
                            }
                            archive_info.append(archive_data)
                            archive.delete()  # DELETE the archive since employee is restored
                            archives_deleted.append(archive_data)
                           
                        
                        # Restore the employee
                        employee.restore()
                        
                        # Set to active if requested
                        if restore_to_active and active_status:
                            employee.status = active_status
                            employee.save()
                        
                        # Log activity
                        EmployeeActivity.objects.create(
                            employee=employee,
                            activity_type='RESTORED',
                            description=f"Employee {employee.full_name} bulk restored from soft deletion. {len(vacancy_info)} vacancies removed. {len(archive_info)} archives deleted.",
                            performed_by=request.user,
                            metadata={
                                'bulk_restoration': True,
                                'restored_from_deletion': True,
                                'originally_deleted_at': employee_info['was_deleted_at'].isoformat() if employee_info['was_deleted_at'] else None,
                                'restored_to_active': restore_to_active,
                                'restoration_method': 'bulk_restore',
                                'vacancies_removed': vacancy_info,
                                'archives_deleted': archive_info,  # FIXED: Include deleted archives
                                'archive_updated': len(archive_info) > 0,
                                'original_employee_pk_restored': employee.pk
                            }
                        )
                    
                    results.append({
                        'employee_id': employee_info['id'],
                        'employee_name': employee_info['name'],
                        'status': 'success',
                        'original_employee_id': employee_info['employee_id'],
                        'was_deleted_at': employee_info['was_deleted_at'],
                        'restored_to_active': restore_to_active,
                        'vacancies_removed': len(vacancy_info),
                        'archives_deleted': len(archive_info)  # FIXED: Include in results
                    })
                    
                except Exception as e:
                    results.append({
                        'employee_id': employee.id,
                        'employee_name': employee.full_name,
                        'status': 'failed',
                        'error': str(e)
                    })
            
            successful_count = len([r for r in results if r['status'] == 'success'])
            failed_count = len([r for r in results if r['status'] == 'failed'])
            
            return Response({
                'success': True,
                'message': f'Bulk restoration completed: {successful_count} successful, {failed_count} failed',
                'summary': {
                    'total_requested': len(employee_ids),
                    'successful': successful_count,
                    'failed': failed_count,
                    'restored_to_active': restore_to_active,
                    'total_vacancies_removed': len(vacancies_removed),
                    'total_archives_deleted': len(archives_deleted)  # FIXED: Include archive deletion count
                },
                'results': results,
                'restoration_type': 'bulk_restore_with_vacancy_and_archive_cleanup'  # FIXED: Updated type
            })
            
        except Exception as e:
            logger.error(f"Bulk restore failed: {str(e)}")
            return Response(
                {'error': f'Bulk restore failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
    method='post',
    operation_description="Bulk soft delete employees, create vacancies and archive data",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['employee_ids'],
        properties={
            'employee_ids': openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(type=openapi.TYPE_INTEGER),
                description='List of employee IDs to soft delete',
                example=[1, 2, 3]
            ),
            'reason': openapi.Schema(
                type=openapi.TYPE_STRING,
                description='Reason for bulk deletion',
                example='Department restructuring'
            )
        }
    ),
    responses={
        200: openapi.Response(
            description="Bulk soft deletion completed with vacancies created",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                    'message': openapi.Schema(type=openapi.TYPE_STRING),
                    'summary': openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            'total_requested': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'successful': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'failed': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'vacancies_created': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'archives_created': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'data_preserved': openapi.Schema(type=openapi.TYPE_BOOLEAN)
                        }
                    ),
                    'results': openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                'employee_name': openapi.Schema(type=openapi.TYPE_STRING),
                                'status': openapi.Schema(type=openapi.TYPE_STRING),
                                'vacancy_created': openapi.Schema(type=openapi.TYPE_OBJECT),
                                'archive_created': openapi.Schema(type=openapi.TYPE_OBJECT)
                            }
                        )
                    )
                }
            )
        ),
        400: "Bad request - validation errors"
    }
)
    @action(detail=False, methods=['post'], url_path='bulk-soft-delete-with-vacancies')
    def bulk_soft_delete_with_vacancies(self, request):
        """FIXED: Bulk soft delete employees, create vacancies AND archive all data with PK preservation"""
        try:
            employee_ids = request.data.get('employee_ids', [])
            reason = request.data.get('reason', 'Bulk restructuring')
            
            
            
            if not employee_ids:
                return Response(
                    {'error': 'employee_ids list is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get employees outside transaction first
            employees = Employee.objects.filter(id__in=employee_ids, is_deleted=False)
            found_employee_ids = list(employees.values_list('id', flat=True))
            missing_employee_ids = [emp_id for emp_id in employee_ids if emp_id not in found_employee_ids]
            
            if missing_employee_ids:
                return Response(
                    {
                        'error': 'Some employee IDs were not found or already deleted',
                        'missing_employee_ids': missing_employee_ids,
                        'found_employee_ids': found_employee_ids,
                        'total_requested': len(employee_ids),
                        'found_count': len(found_employee_ids)
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            results = []
            vacancies_created = []
            archives_created = []
            total_direct_reports_updated = 0
            
            # Process each employee individually with separate transactions
            for employee in employees:
                employee_info = {
                    'id': employee.id,
                    'employee_id': employee.employee_id,
                    'name': employee.full_name,
                    'job_title': employee.job_title,
                    'department': employee.department.name if employee.department else None,
                    'direct_reports_count': employee.direct_reports.filter(is_deleted=False).count()
                }
                
                try:
                    # Use individual transaction for each employee
                    with transaction.atomic():
                        vacancy, archive = self._manual_bulk_soft_delete_with_archive(employee, request.user, reason)
                        
                        results.append({
                            'employee_id': employee_info['id'],
                            'employee_name': employee_info['name'],
                            'employee_hc_id': employee_info['employee_id'],
                            'status': 'success',
                            'vacancy_created': {
                                'id': vacancy.id,
                                'position_id': vacancy.position_id,
                                'job_title': vacancy.job_title,
                                'original_employee_pk': vacancy.original_employee_pk  # FIXED: Include PK reference
                            } if vacancy else None,
                            'archive_created': {
                                'id': archive.id if archive else None,
                                'reference': archive.get_archive_reference() if archive else None,
                                'status': archive.get_deletion_summary()['data_quality'] if archive else None
                            },
                            'direct_reports_updated': employee_info['direct_reports_count'],
                            'original_employee_pk': employee_info['id']  # FIXED: Store for restoration
                        })
                        
                        if vacancy:
                            vacancies_created.append(vacancy)
                        if archive:
                            archives_created.append(archive)
                        total_direct_reports_updated += employee_info['direct_reports_count']
                        
                except Exception as e:
                    logger.error(f"Failed to process employee {employee.employee_id}: {str(e)}")
                    results.append({
                        'employee_id': employee.id,
                        'employee_name': employee.full_name,
                        'employee_hc_id': employee.employee_id,
                        'status': 'failed',
                        'error': str(e),
                        'error_details': {
                            'exception_type': type(e).__name__,
                            'full_error': str(e)
                        }
                    })
            
            successful_count = len([r for r in results if r['status'] == 'success'])
            failed_count = len([r for r in results if r['status'] == 'failed'])
            
        
            
            return Response({
                'success': True,
                'message': f'Bulk soft deletion completed: {successful_count} successful, {failed_count} failed',
                'summary': {
                    'total_requested': len(employee_ids),
                    'employees_found': len(employees),
                    'successful': successful_count,
                    'failed': failed_count,
                    'vacancies_created': len(vacancies_created),
                    'archives_created': len(archives_created),
                    'total_direct_reports_updated': total_direct_reports_updated,
                    'data_preserved': True,
                    'can_restore': True
                },
                'results': results,
                'reason': reason,
                'deletion_type': 'bulk_soft_delete_with_vacancies_and_archives'
            })
            
        except Exception as e:
            logger.error(f"Bulk soft delete failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {
                    'error': f'Bulk soft delete failed: {str(e)}',
                    'error_type': type(e).__name__
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _manual_bulk_soft_delete_with_archive(self, employee, user, reason):
        """
        FIXED: Manual bulk soft delete with proper PK preservation
        """
        try:
            # Store employee data for vacancy creation before any database operations
            employee_data = {
                'job_title': employee.job_title,
                'original_employee_pk': employee.pk,  # FIXED: Store original PK
                'business_function': employee.business_function,
                'department': employee.department,
                'unit': employee.unit,
                'job_function': employee.job_function,
                'position_group': employee.position_group,
                'grading_level': employee.grading_level,
                'reporting_to': employee.line_manager,
                'is_visible_in_org_chart': employee.is_visible_in_org_chart,
                'notes': f"Position vacated by {employee.full_name} ({employee.employee_id}) on {timezone.now().date()}. Reason: {reason}"
            }
            
            # Create vacant position from employee data
            vacancy = VacantPosition.objects.create(
                job_title=employee_data['job_title'],
                original_employee_pk=employee_data['original_employee_pk'],  # FIXED: Set original PK
                business_function=employee_data['business_function'],
                department=employee_data['department'],
                unit=employee_data['unit'],
                job_function=employee_data['job_function'],
                position_group=employee_data['position_group'],
                grading_level=employee_data['grading_level'],
                reporting_to=employee_data['reporting_to'],
                include_in_headcount=True,
                is_visible_in_org_chart=employee_data['is_visible_in_org_chart'],
                notes=employee_data['notes'],
                created_by=user
            )
            
            # Update direct reports to report to this employee's manager
            if employee.line_manager:
                direct_reports = employee.direct_reports.filter(is_deleted=False)
                for report in direct_reports:
                    report.line_manager = employee.line_manager
                    report.updated_by = user
                    report.save()
            
            # Create archive record BEFORE soft deletion
            archive = employee._create_archive_record(
                deletion_notes=f"Employee bulk soft deleted and vacancy {vacancy.position_id} created. Reason: {reason}",
                deleted_by=user,
                preserve_original_data=True
            )
            
            # Soft delete the employee
            employee.soft_delete(user)
            
            # Log the soft delete activity
            EmployeeActivity.objects.create(
                employee=employee,
                activity_type='SOFT_DELETED',
                description=f"Employee {employee.full_name} bulk soft deleted, vacancy {vacancy.position_id} created, and archived",
                performed_by=user,
                metadata={
                    'delete_type': 'bulk_soft_with_vacancy',
                    'vacancy_created': True,
                    'vacancy_id': vacancy.id,
                    'vacancy_position_id': vacancy.position_id,
                    'employee_data_preserved': True,
                    'can_be_restored': True,
                    'archive_id': archive.id if archive else None,
                    'archive_reference': archive.get_archive_reference() if archive else None,
                    'original_employee_pk': employee.pk,  # FIXED: Store for restoration
                    'bulk_operation': True
                }
            )
            
            return vacancy, archive
            
        except Exception as e:
            logger.error(f"Manual bulk soft delete failed for employee {employee.employee_id}: {str(e)}")
            raise e
    
   
    @swagger_auto_schema(
    method='post',
    operation_description="Remove tag from multiple employees and set status to ACTIVE",
    request_body=BulkEmployeeTagUpdateSerializer,
    responses={200: "Tags removed and statuses updated successfully", 400: "Bad request"}
)
    @action(detail=False, methods=['post'], url_path='bulk-remove-tag')
    def bulk_remove_tag(self, request):
        """
        ✅ UPDATED: Remove tag from multiple employees and automatically set status to ACTIVE
        """
        serializer = BulkEmployeeTagUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_ids = serializer.validated_data['employee_ids']
        tag_id = serializer.validated_data['tag_id']
        
        try:
            tag = EmployeeTag.objects.get(id=tag_id)
            employees = Employee.objects.filter(id__in=employee_ids)
            
            # Get ACTIVE status
            active_status = EmployeeStatus.objects.filter(
                status_type='ACTIVE',
                is_active=True
            ).first()
            
            if not active_status:
                logger.warning("ACTIVE status not found in system")
            
            removed_count = 0
            didnt_have_count = 0
            status_changed_count = 0
            results = []
            
            with transaction.atomic():
                for employee in employees:
                    result_data = {
                        'employee_id': employee.id,
                        'employee_name': employee.full_name,
                        'employee_hc_id': employee.employee_id,
                        'tag_removed': False,
                        'status_changed': False,
                        'old_status': employee.status.name if employee.status else None,
                        'new_status': None
                    }
                    
                    # Check if tag exists
                    tag_existed = employee.tags.filter(id=tag.id).exists()
                    
                    if tag_existed:
                        # Remove tag using the updated remove_tag method
                        # This will automatically change status to ACTIVE
                        tag_removed = employee.remove_tag(tag, request.user)
                        
                        if tag_removed:
                            removed_count += 1
                            result_data['tag_removed'] = True
                            result_data['status'] = 'removed'
                            
                            # Refresh to get updated status
                            employee.refresh_from_db()
                            result_data['new_status'] = employee.status.name if employee.status else None
                            
                            # Check if status was changed
                            if result_data['old_status'] != result_data['new_status']:
                                status_changed_count += 1
                                result_data['status_changed'] = True
                                
                               
                    else:
                        didnt_have_count += 1
                        result_data['status'] = 'didnt_have'
                        result_data['new_status'] = result_data['old_status']
                    
                    results.append(result_data)
            
            return Response({
                'success': True,
                'message': f"Tag '{tag.name}' removal processed for {len(employee_ids)} employees. {status_changed_count} statuses changed to ACTIVE.",
                'tag_name': tag.name,
                'total_employees': len(employee_ids),
                'removed_count': removed_count,
                'didnt_have_count': didnt_have_count,
                'status_changed_count': status_changed_count,
                'active_status_available': bool(active_status),
                'results': results
            })
            
        except EmployeeTag.DoesNotExist:
            return Response({'error': 'Tag not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Bulk remove tag failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Bulk remove tag failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    @swagger_auto_schema(
        method='post',
        operation_description="Assign line manager to single employee",
        request_body=SingleLineManagerAssignmentSerializer,
        responses={200: "Line manager assigned successfully", 400: "Bad request"}
    )
    @action(detail=False, methods=['post'], url_path='assign-line-manager')
    def assign_line_manager(self, request):
        """Assign line manager to single employee"""
        serializer = SingleLineManagerAssignmentSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_id = serializer.validated_data['employee_id']
        line_manager_id = serializer.validated_data['line_manager_id']
        
        try:
            employee = Employee.objects.get(id=employee_id)
            line_manager = Employee.objects.get(id=line_manager_id) if line_manager_id else None
            
            old_manager_name = employee.line_manager.full_name if employee.line_manager else 'None'
            new_manager_name = line_manager.full_name if line_manager else 'None'
            
            employee.change_line_manager(line_manager, request.user)
            
            return Response({
                'success': True,
                'message': f'Line manager updated for {employee.full_name}',
                'employee_id': employee.id,
                'employee_name': employee.full_name,
                'old_line_manager': old_manager_name,
                'new_line_manager': new_manager_name
            })
        except Employee.DoesNotExist:
            return Response({'error': 'Employee or line manager not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @swagger_auto_schema(
        method='post',
        operation_description="Assign line manager to multiple employees",
        request_body=BulkLineManagerAssignmentSerializer,
        responses={200: "Line managers assigned successfully", 400: "Bad request"}
    )
    @action(detail=False, methods=['post'], url_path='bulk-assign-line-manager')
    def bulk_assign_line_manager(self, request):
        """Assign line manager to multiple employees"""
        serializer = BulkLineManagerAssignmentSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_ids = serializer.validated_data['employee_ids']
        line_manager_id = serializer.validated_data['line_manager_id']
        
        try:
            line_manager = Employee.objects.get(id=line_manager_id) if line_manager_id else None
            employees = Employee.objects.filter(id__in=employee_ids)
            
            updated_count = 0
            results = []
            
            with transaction.atomic():
                for employee in employees:
                    old_manager_name = employee.line_manager.full_name if employee.line_manager else 'None'
                    employee.change_line_manager(line_manager, request.user)
                    updated_count += 1
                    
                    results.append({
                        'employee_id': employee.id,
                        'employee_name': employee.full_name,
                        'old_line_manager': old_manager_name,
                        'new_line_manager': line_manager.full_name if line_manager else 'None'
                    })
            
            return Response({
                'success': True,
                'message': f'Line manager updated for {updated_count} employees',
                'new_line_manager': line_manager.full_name if line_manager else 'None',
                'total_employees': len(employee_ids),
                'updated_count': updated_count,
                'results': results
            })
        except Employee.DoesNotExist:
            return Response({'error': 'Line manager not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @swagger_auto_schema(
        method='post',
        operation_description="Update contract for single employee",
        request_body=ContractExtensionSerializer,
        responses={200: "Contract updated successfully", 400: "Bad request"}
    )
    @action(detail=False, methods=['post'], url_path='extend-contract')
    def extend_employee_contract(self, request):
        """Update contract for single employee with new type and start date"""
        serializer = ContractExtensionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_id = serializer.validated_data['employee_id']
        new_contract_type = serializer.validated_data['new_contract_type']
        new_start_date = serializer.validated_data['new_start_date']
        reason = serializer.validated_data.get('reason', '')
        
        try:
            employee = Employee.objects.get(id=employee_id)
            
            # Store old values
            old_contract_type = employee.contract_duration
            old_start_date = employee.contract_start_date
            old_end_date = employee.contract_end_date
            
            # Update contract
            employee.contract_duration = new_contract_type
            employee.contract_start_date = new_start_date
            employee.contract_extensions += 1
            employee.last_extension_date = timezone.now().date()
            
            
            if request.user:
                employee.updated_by = request.user
            
            # Save will auto-calculate new end date
            employee.save()
            
            # Log detailed activity
            EmployeeActivity.objects.create(
                employee=employee,
                activity_type='CONTRACT_UPDATED',
                description=f"Contract updated: {old_contract_type} → {new_contract_type}. New start: {new_start_date}. Reason: {reason}",
                performed_by=request.user,
                metadata={
                    'old_contract_type': old_contract_type,
                    'new_contract_type': new_contract_type,
                    'old_start_date': str(old_start_date) if old_start_date else None,
                    'new_start_date': str(new_start_date),
                    'old_end_date': str(old_end_date) if old_end_date else None,
                    'new_end_date': str(employee.contract_end_date) if employee.contract_end_date else None,
                    'reason': reason,
                    'extension_count': employee.contract_extensions
                }
            )
            
            return Response({
                'success': True,
                'message': f'Contract updated successfully for {employee.full_name}',
                'employee_id': employee.id,
                'employee_name': employee.full_name,
                'old_contract_type': old_contract_type,
                'new_contract_type': new_contract_type,
                'old_start_date': old_start_date,
                'new_start_date': new_start_date,
                'old_end_date': old_end_date,
                'new_end_date': employee.contract_end_date,
                'extensions_count': employee.contract_extensions
            })
                
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @swagger_auto_schema(
        method='post',
        operation_description="Update contracts for multiple employees",
        request_body=BulkContractExtensionSerializer,
        responses={200: "Contracts updated successfully", 400: "Bad request"}
    )
    @action(detail=False, methods=['post'], url_path='bulk-extend-contracts')
    def bulk_extend_contracts(self, request):
        """Update contracts for multiple employees"""
        serializer = BulkContractExtensionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        employee_ids = serializer.validated_data['employee_ids']
        new_contract_type = serializer.validated_data['new_contract_type']
        new_start_date = serializer.validated_data['new_start_date']
        reason = serializer.validated_data.get('reason', '')
        
        try:
            employees = Employee.objects.filter(id__in=employee_ids)
            
            updated_count = 0
            failed_count = 0
            results = []
            
            with transaction.atomic():
                for employee in employees:
                    try:
                        old_contract_type = employee.contract_duration
                        old_start_date = employee.contract_start_date
                        old_end_date = employee.contract_end_date
                        
                        # Update contract
                        employee.contract_duration = new_contract_type
                        employee.contract_start_date = new_start_date
                        employee.contract_extensions += 1
                        employee.last_extension_date = timezone.now().date()
                       
                        
                        if request.user:
                            employee.updated_by = request.user
                        
                        # Save will auto-calculate new end date
                        employee.save()
                        
                        # Log detailed activity
                        EmployeeActivity.objects.create(
                            employee=employee,
                            activity_type='CONTRACT_UPDATED',
                            description=f"Bulk contract update: {old_contract_type} → {new_contract_type}. New start: {new_start_date}. Reason: {reason}",
                            performed_by=request.user,
                            metadata={
                                'bulk_update': True,
                                'old_contract_type': old_contract_type,
                                'new_contract_type': new_contract_type,
                                'old_start_date': str(old_start_date) if old_start_date else None,
                                'new_start_date': str(new_start_date),
                                'old_end_date': str(old_end_date) if old_end_date else None,
                                'new_end_date': str(employee.contract_end_date) if employee.contract_end_date else None,
                                'reason': reason,
                                'extension_count': employee.contract_extensions
                            }
                        )
                        
                        updated_count += 1
                        results.append({
                            'employee_id': employee.id,
                            'employee_name': employee.full_name,
                            'status': 'success',
                            'old_contract_type': old_contract_type,
                            'new_contract_type': new_contract_type,
                            'old_end_date': old_end_date,
                            'new_end_date': employee.contract_end_date,
                            'extensions_count': employee.contract_extensions
                        })
                    except Exception as e:
                        failed_count += 1
                        results.append({
                            'employee_id': employee.id,
                            'employee_name': employee.full_name,
                            'status': 'failed',
                            'error': str(e)
                        })
            
            return Response({
                'success': True,
                'message': f'Contract update completed: {updated_count} updated, {failed_count} failed',
                'total_employees': len(employee_ids),
                'updated_count': updated_count,
                'failed_count': failed_count,
                'new_contract_type': new_contract_type,
                'new_start_date': new_start_date,
                'reason': reason,
                'results': results
            })
            
        except Exception as e:
            return Response({'error': f'Bulk contract update failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
 

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """✅ COMPLETE: Get comprehensive employee statistics with vacant positions"""
        queryset = self.get_queryset()
        
        # Apply filtering
        employee_filter = ComprehensiveEmployeeFilter(queryset, request.query_params)
        queryset = employee_filter.filter()
        
        total_employees = queryset.count()
        active_employees = queryset.filter(status__affects_headcount=True).count()
        
        # By status
        status_stats = {}
        for emp_status in EmployeeStatus.objects.filter(is_active=True):
            count = queryset.filter(status=emp_status).count()
            if count > 0:
                status_stats[emp_status.name] = {
                    'count': count,
                    'color': emp_status.color,
                    'affects_headcount': emp_status.affects_headcount
                }
        
        # ✅ By business function with ALL info
        function_stats = {}
        for func in BusinessFunction.objects.filter(is_active=True):
            emp_count = queryset.filter(business_function=func).count()
            active_count = queryset.filter(
                business_function=func, 
                status__affects_headcount=True
            ).count()
            
            # Recent hires in this function
            recent_hires = queryset.filter(
                business_function=func,
                start_date__gte=date.today() - timedelta(days=30)
            ).count()
            
            function_stats[func.name] = {
                'count': emp_count,
                'active': active_count,
                'recent_hires': recent_hires
            }
        
        # ✅ Vacant positions by business function
        vacant_by_function = {}
        vacant_positions = VacantPosition.objects.filter(
            is_filled=False,
            is_deleted=False,
            include_in_headcount=True
        ).select_related('business_function')
        
        for func in BusinessFunction.objects.filter(is_active=True):
            vacant_count = vacant_positions.filter(business_function=func).count()
            vacant_by_function[func.name] = vacant_count
        
        # By position group
        position_stats = {}
        for pos in PositionGroup.objects.filter(is_active=True):
            count = queryset.filter(position_group=pos).count()
            if count > 0:
                position_stats[pos.get_name_display()] = count
        
        # Contract duration statistics
        contract_stats = {}
        try:
            contract_types = queryset.values_list('contract_duration', flat=True).distinct()
            for contract_type in contract_types:
                if contract_type:
                    count = queryset.filter(contract_duration=contract_type).count()
                    if count > 0:
                        try:
                            config = ContractTypeConfig.objects.get(
                                contract_type=contract_type, 
                                is_active=True
                            )
                            display_name = config.display_name
                        except ContractTypeConfig.DoesNotExist:
                            display_name = contract_type.replace('_', ' ').title()
                        contract_stats[display_name] = count
        except Exception as e:
            logger.error(f"Error calculating contract statistics: {e}")
            contract_stats = {}
        
        # Recent activity
        recent_hires = queryset.filter(
            start_date__gte=date.today() - timedelta(days=30)
        ).count()
        
        upcoming_contract_endings = queryset.filter(
            contract_end_date__lte=date.today() + timedelta(days=30),
            contract_end_date__gte=date.today()
        ).count()
        
        # Status update analysis
        try:
            from .status_management import EmployeeStatusManager
            employees_needing_updates = EmployeeStatusManager.get_employees_needing_update()
            status_update_stats = {
                'employees_needing_updates': len(employees_needing_updates),
                'status_transitions': {}
            }
            
            for update_info in employees_needing_updates:
                transition = f"{update_info['current_status']} → {update_info['required_status']}"
                status_update_stats['status_transitions'][transition] = \
                    status_update_stats['status_transitions'].get(transition, 0) + 1
        except Exception as e:
            status_update_stats = {
                'employees_needing_updates': 0,
                'status_transitions': {},
                'error': str(e)
            }
        
        # ✅ Total vacant positions
        total_vacant = vacant_positions.count()
        
        return Response({
            'total_employees': total_employees,
            'active_employees': active_employees,
            'inactive_employees': total_employees - active_employees,
            'total_vacant_positions': total_vacant,  # ✅ NEW
            'by_status': status_stats,
            'by_business_function': function_stats,
            'vacant_positions_by_business_function': vacant_by_function,  # ✅ NEW
            'by_position_group': position_stats,
            'by_contract_duration': contract_stats,
            'recent_hires_30_days': recent_hires,
            'upcoming_contract_endings_30_days': upcoming_contract_endings,
            'status_update_analysis': status_update_stats
        })
    @action(detail=True, methods=['get'])
    def activities(self, request, pk=None):
        """Get employee activity history"""
        employee = self.get_object()
        activities = employee.activities.all()[:50]  # Last 50 activities
        serializer = EmployeeActivitySerializer(activities, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def status_preview(self, request, pk=None):
        """Get status preview for individual employee"""
        employee = self.get_object()
        preview = employee.get_status_preview()
        
        return Response({
            'employee_id': employee.employee_id,
            'employee_name': employee.full_name,
            'preview': preview
        })
    
    @action(detail=True, methods=['get'])
    def direct_reports(self, request, pk=None):
        """Get direct reports for an employee (NEW)"""
        employee = self.get_object()
        reports = employee.direct_reports.filter(
            status__affects_headcount=True,
            is_deleted=False
        ).select_related('status', 'position_group', 'department')
        
        serializer = EmployeeListSerializer(reports, many=True)
        return Response({
            'manager': {
                'id': employee.id,
                'employee_id': employee.employee_id,
                'name': employee.full_name,
                'job_title': employee.job_title
            },
            'direct_reports_count': reports.count(),
            'direct_reports': serializer.data
        })
    
class BulkEmployeeUploadViewSet(viewsets.ViewSet):
    """Ayrı ViewSet yalnız file upload üçün"""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]  # Yalnız file upload
    
    @swagger_auto_schema(
        operation_description="Bulk create employees from uploaded Excel file",
        manual_parameters=[
            openapi.Parameter(
                'file',
                openapi.IN_FORM,
                description='Excel file (.xlsx, .xls) containing employee data',
                type=openapi.TYPE_FILE,
                required=True
            )
        ],
        consumes=['multipart/form-data'],
        responses={
            200: openapi.Response(
                description="File processed successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'total_rows': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'successful': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'failed': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'errors': openapi.Schema(
                            type=openapi.TYPE_ARRAY, 
                            items=openapi.Schema(type=openapi.TYPE_STRING)
                        ),
                        'created_employees': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    'employee_id': openapi.Schema(type=openapi.TYPE_STRING),
                                    'name': openapi.Schema(type=openapi.TYPE_STRING),
                                    'email': openapi.Schema(type=openapi.TYPE_STRING)
                                }
                            )
                        ),
                        'filename': openapi.Schema(type=openapi.TYPE_STRING)
                    }
                )
            ),
            400: openapi.Response(
                description="Bad request - file validation error",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'error': openapi.Schema(type=openapi.TYPE_STRING)
                    }
                )
            )
        }
    )
    def create(self, request):
        """Bulk create employees from uploaded Excel file"""
        
        try:
         
            
            # Check if file exists
            if 'file' not in request.FILES:
                logger.warning("No file in request.FILES")
                return Response(
                    {'error': 'No file uploaded. Please select an Excel file.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            file = request.FILES['file']
        
            
            # Validate file format
            if not file.name.endswith(('.xlsx', '.xls')):
                logger.warning(f"Invalid file format: {file.name}")
                return Response(
                    {'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate file size (max 10MB)
            if file.size > 10 * 1024 * 1024:
                logger.warning(f"File too large: {file.size} bytes")
                return Response(
                    {'error': 'File too large. Maximum size is 10MB'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Read Excel file with better error handling
            try:
                # Try multiple engines
                try:
                    df = pd.read_excel(file, sheet_name=0, engine='openpyxl')
                except:
                    try:
                        df = pd.read_excel(file, sheet_name=0, engine='xlrd')
                    except:
                        df = pd.read_excel(file, sheet_name=0)
                
             
                
            except Exception as e:
                logger.error(f"Failed to read Excel file: {str(e)}")
                return Response(
                    {'error': f'Failed to read Excel file: {str(e)}. Please check file format and content.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Check if file has data
            if df.empty:
                logger.warning("Excel file is empty after removing empty rows")
                return Response(
                    {'error': 'Excel file is empty or has no valid data'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
          
            
            # Process the data - call EmployeeViewSet method
            employee_viewset = EmployeeViewSet()
            
            # Make sure the viewset has the method
            if not hasattr(employee_viewset, '_process_bulk_employee_data_from_excel'):
                logger.error("EmployeeViewSet missing _process_bulk_employee_data_from_excel method")
                return Response(
                    {'error': 'Server configuration error. Please contact administrator.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            result = employee_viewset._process_bulk_employee_data_from_excel(df, request.user)
            
         
            
            return Response({
                'message': f'File processed successfully. {result["successful"]} employees created, {result["failed"]} failed.',
                'total_rows': result['total_rows'],
                'successful': result['successful'],
                'failed': result['failed'],
                'errors': result['errors'],
                'created_employees': result['created_employees'],
                'filename': file.name
            })
            
        except Exception as e:
   
            return Response(
                {'error': f'Failed to process request: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        operation_description="Download Excel template for bulk employee creation",
        responses={
            200: openapi.Response(
                description="Excel template file",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        }
    )
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download Excel template for bulk employee creation"""
        try:
         
            
            # Call EmployeeViewSet template method
            employee_viewset = EmployeeViewSet()
            
            # Make sure the method exists
            if not hasattr(employee_viewset, '_generate_bulk_template'):
                logger.error("EmployeeViewSet missing _generate_bulk_template method")
                return Response(
                    {'error': 'Template generation not available. Please contact administrator.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            response = employee_viewset._generate_bulk_template()
       
            return response
            
        except Exception as e:

            return Response(
                {'error': f'Failed to generate template: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )    
       
class OrgChartFilter:
    """
    Comprehensive filter system for organizational chart
    Handles all org chart specific filtering needs
    """
    
    def __init__(self, queryset, params):
        self.queryset = queryset
        self.params = params
    
    def parse_comma_separated(self, param_value):
        """Parse comma-separated string into list of cleaned values"""
        if not param_value:
            return []
        
        if isinstance(param_value, list):
            result = []
            for item in param_value:
                if isinstance(item, str) and ',' in item:
                    result.extend([val.strip() for val in item.split(',') if val.strip()])
                elif item:
                    result.append(str(item).strip())
            return result
        elif isinstance(param_value, str):
            return [val.strip() for val in param_value.split(',') if val.strip()]
        else:
            return [str(param_value).strip()] if param_value else []
    
    def get_filter_values(self, param_name):
        """Get filter values, handling both getlist() and comma-separated strings"""
        if hasattr(self.params, 'getlist'):
            values = self.params.getlist(param_name)
            if values:
                all_values = []
                for value in values:
                    all_values.extend(self.parse_comma_separated(value))
                return all_values
        
        single_value = self.params.get(param_name)
        if single_value:
            return self.parse_comma_separated(single_value)
        
        return []
    
    def get_int_filter_values(self, param_name):
        """Get integer filter values"""
        string_values = self.get_filter_values(param_name)
        int_values = []
        for val in string_values:
            try:
                int_values.append(int(val))
            except (ValueError, TypeError):
                continue
        return int_values
    
    def filter(self):
        queryset = self.queryset
        
        print(f"🔍 ORG CHART FILTER DEBUG: Raw params = {dict(self.params)}")
        
        # ===========================================
        # 1. SEARCH FILTERS
        # ===========================================
        
        # ✅ DÜZƏLDILMIŞ: Employee ID search
        employee_id_search = self.params.get('employee_id_search')
        if employee_id_search:
            print(f"🔍 Applying employee ID search: {employee_id_search}")
            queryset = queryset.filter(employee_id__icontains=employee_id_search)
        
        # General search across multiple fields
        search = self.params.get('search')
        if search:
            print(f"🔍 Applying org chart search: {search}")
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(employee_id__icontains=search) |  # ✅ Employee ID də search-ə daxildir
                Q(user__email__icontains=search) |
                Q(job_title__icontains=search) |
                Q(business_function__name__icontains=search) |
                Q(department__name__icontains=search) |
                Q(unit__name__icontains=search) |
                Q(father_name__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Job title search
        job_title_search = self.params.get('job_title_search')
        if job_title_search:
            print(f"🔍 Applying job title search: {job_title_search}")
            queryset = queryset.filter(job_title__icontains=job_title_search)
        
        # Department search
        department_search = self.params.get('department_search')
        if department_search:
            print(f"🔍 Applying department search: {department_search}")
            queryset = queryset.filter(department__name__icontains=department_search)
        
        # ===========================================
        # 2. ORGANIZATIONAL STRUCTURE FILTERS
        # ===========================================
        
        # ✅ DÜZƏLDILMIŞ: Business Functions (array)
        business_function_ids = self.get_int_filter_values('business_function')
        if business_function_ids:
            print(f"🏭 Applying business function filter: {business_function_ids}")
            queryset = queryset.filter(business_function__id__in=business_function_ids)
        
        # ✅ DÜZƏLDILMIŞ: Departments (array)
        department_ids = self.get_int_filter_values('department')
        if department_ids:
            print(f"🏢 Applying department filter: {department_ids}")
            queryset = queryset.filter(department__id__in=department_ids)
        
        # ✅ DÜZƏLDILMIŞ: Units (array)
        unit_ids = self.get_int_filter_values('unit')
        if unit_ids:
            print(f"🏢 Applying unit filter: {unit_ids}")
            queryset = queryset.filter(unit__id__in=unit_ids)
        
        # ✅ DÜZƏLDILMIŞ: Job Functions (array)
        job_function_ids = self.get_int_filter_values('job_function')
        if job_function_ids:
            print(f"💼 Applying job function filter: {job_function_ids}")
            queryset = queryset.filter(job_function__id__in=job_function_ids)
        
        # ✅ DÜZƏLDILMIŞ: Position Groups (array)
        position_group_ids = self.get_int_filter_values('position_group')
        if position_group_ids:
            print(f"📊 Applying position group filter: {position_group_ids}")
            queryset = queryset.filter(position_group__id__in=position_group_ids)
        
        # ===========================================
        # 3. HIERARCHY FILTERS
        # ===========================================
        
        # ✅ DÜZƏLDILMIŞ: Line Managers (array)
        line_manager_ids = self.get_int_filter_values('line_manager')
        if line_manager_ids:
            print(f"👨‍💼 Applying line manager filter: {line_manager_ids}")
            queryset = queryset.filter(line_manager__id__in=line_manager_ids)
        
        # Top level managers only (no line manager)
        show_top_level_only = self.params.get('show_top_level_only')
        if show_top_level_only and show_top_level_only.lower() == 'true':
            print(f"👑 Showing top level managers only")
            queryset = queryset.filter(line_manager__isnull=True)
        
        # Specific manager's team (direct reports)
        manager_team = self.params.get('manager_team')
        if manager_team:
            try:
                manager_id = int(manager_team)
                print(f"👥 Showing team for manager ID: {manager_id}")
                queryset = queryset.filter(line_manager__id=manager_id)
            except (ValueError, TypeError):
                pass
        
        # Hierarchy level filters
        max_hierarchy_level = self.params.get('max_hierarchy_level')
        if max_hierarchy_level:
            try:
                max_level = int(max_hierarchy_level)
                print(f"📈 Applying max hierarchy level: {max_level}")
                queryset = queryset.filter(position_group__hierarchy_level__lte=max_level)
            except (ValueError, TypeError):
                pass
        
        min_hierarchy_level = self.params.get('min_hierarchy_level')
        if min_hierarchy_level:
            try:
                min_level = int(min_hierarchy_level)
                print(f"📈 Applying min hierarchy level: {min_level}")
                queryset = queryset.filter(position_group__hierarchy_level__gte=min_level)
            except (ValueError, TypeError):
                pass
        
        # ===========================================
        # 4. EMPLOYMENT STATUS FILTERS
        # ===========================================
        
        # ✅ DÜZƏLDILMIŞ: Employment Status (array)
        status_values = self.get_filter_values('status')
        if status_values:
            print(f"🎯 Applying status filter: {status_values}")
            status_q = Q()
            for status_val in status_values:
                try:
                    status_id = int(status_val)
                    status_q |= Q(status__id=status_id)
                except (ValueError, TypeError):
                    status_q |= Q(status__name=status_val)
            
            if status_q:
                queryset = queryset.filter(status_q)
        
        # ✅ DÜZƏLDILMIŞ: Grading Levels (array)
        grading_levels = self.get_filter_values('grading_level')
        if grading_levels:
            print(f"📈 Applying grading level filter: {grading_levels}")
            queryset = queryset.filter(grading_level__in=grading_levels)
        
        # ===========================================
        # 5. ORG CHART SPECIFIC FILTERS
        # ===========================================
        
        # Exclude employees without teams (no direct reports)
        hide_individual_contributors = self.params.get('hide_individual_contributors')
        if hide_individual_contributors and hide_individual_contributors.lower() == 'true':
       
            queryset = queryset.annotate(
                direct_reports_count=Count(
                    'direct_reports',
                    filter=Q(direct_reports__status__affects_headcount=True, direct_reports__is_deleted=False)
                )
            ).filter(direct_reports_count__gt=0)
        
        # Show only managers (have direct reports)
        managers_only = self.params.get('managers_only')
        if managers_only and managers_only.lower() == 'true':
        
            queryset = queryset.annotate(
                direct_reports_count=Count(
                    'direct_reports',
                    filter=Q(direct_reports__status__affects_headcount=True, direct_reports__is_deleted=False)
                )
            ).filter(direct_reports_count__gt=0)
        
        # ===========================================
        # 6. DEMOGRAPHIC FILTERS
        # ===========================================
        
        # ✅ DÜZƏLDILMIŞ: Gender filter (array)
        genders = self.get_filter_values('gender')
        if genders:
            print(f"👤 Applying gender filter: {genders}")
            queryset = queryset.filter(gender__in=genders)
        
        final_count = queryset.count()
        print(f"✅ Filter complete: {final_count} employees after all filters")
        
        return queryset

class OrgChartViewSet(viewsets.ViewSet):  
   
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Get complete organizational chart tree including vacant positions with filtering",
        manual_parameters=[
            # Search Parameters
            openapi.Parameter('search', openapi.IN_QUERY, description="General search", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('employee_id_search', openapi.IN_QUERY, description="Search by employee ID", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('job_title_search', openapi.IN_QUERY, description="Search by job title", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('department_search', openapi.IN_QUERY, description="Search by department", type=openapi.TYPE_STRING, required=False),
            
            # Organizational Structure Filters
            openapi.Parameter('business_function', openapi.IN_QUERY, description="Filter by business function IDs (comma-separated)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('department', openapi.IN_QUERY, description="Filter by department IDs (comma-separated)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('unit', openapi.IN_QUERY, description="Filter by unit IDs (comma-separated)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('job_function', openapi.IN_QUERY, description="Filter by job function IDs (comma-separated)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('position_group', openapi.IN_QUERY, description="Filter by position group IDs (comma-separated)", type=openapi.TYPE_STRING, required=False),
            
            # Hierarchy Filters
            openapi.Parameter('line_manager', openapi.IN_QUERY, description="Filter by line manager IDs (comma-separated)", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('show_top_level_only', openapi.IN_QUERY, description="Show only top-level managers", type=openapi.TYPE_BOOLEAN, required=False),
            openapi.Parameter('manager_team', openapi.IN_QUERY, description="Show direct reports of specific manager", type=openapi.TYPE_INTEGER, required=False),
            openapi.Parameter('managers_only', openapi.IN_QUERY, description="Show only employees with direct reports", type=openapi.TYPE_BOOLEAN, required=False),
            
            # Employment Status Filters
            openapi.Parameter('status', openapi.IN_QUERY, description="Filter by employment status", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('grading_level', openapi.IN_QUERY, description="Filter by grading levels", type=openapi.TYPE_STRING, required=False),
            openapi.Parameter('gender', openapi.IN_QUERY, description="Filter by gender", type=openapi.TYPE_STRING, required=False),
            
            # Other
            openapi.Parameter('ordering', openapi.IN_QUERY, description="Order results by field", type=openapi.TYPE_STRING, required=False),
        ],
        responses={
            200: openapi.Response(
                description="Organizational chart tree data",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'org_chart': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(type=openapi.TYPE_OBJECT)
                        ),
                        'statistics': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'metadata': openapi.Schema(type=openapi.TYPE_OBJECT)
                    }
                )
            )
        }
    )
    @action(detail=False, methods=['get'], url_path='tree')
    def get_full_tree(self, request):
        """
        ✅ MAIN ENDPOINT: Get complete organizational chart tree with vacancies and filtering
        
        This replaces the old list() and full_tree_with_vacancies() endpoints
        """
        try:
            # Get base queryset
            employees = Employee.objects.filter(
                status__allows_org_chart=True,
                is_visible_in_org_chart=True,
                is_deleted=False
            ).select_related(
                'user', 'business_function', 'department', 'unit', 'job_function',
                'position_group', 'status', 'line_manager'
            ).prefetch_related('tags').order_by('position_group__hierarchy_level', 'employee_id')
            
            # Apply filtering
            org_filter = OrgChartFilter(employees, request.query_params)
            employees = org_filter.filter()
            
            # Apply sorting if specified
            ordering = request.query_params.get('ordering', '')
            if ordering:
                sort_params = [param.strip() for param in ordering.split(',') if param.strip()]
                if sort_params:
                    employees = employees.order_by(*sort_params)
            
            # Serialize employees
            serializer = OrgChartNodeSerializer(employees, many=True, context={'request': request})
            employee_data = serializer.data
            
            # Get vacant positions
            vacancies = VacantPosition.objects.filter(
                is_visible_in_org_chart=True,
                is_filled=False,
                is_deleted=False
            ).select_related(
                'business_function', 'department', 'unit', 'job_function',
                'position_group', 'vacancy_status', 'reporting_to'
            )
            
            # Apply same filters to vacancies
            business_function_values = request.query_params.getlist('business_function')
            if business_function_values:
                try:
                    bf_ids = []
                    for bf_val in business_function_values:
                        if ',' in bf_val:
                            bf_ids.extend([int(id.strip()) for id in bf_val.split(',') if id.strip().isdigit()])
                        elif bf_val.isdigit():
                            bf_ids.append(int(bf_val))
                    
                    if bf_ids:
                        vacancies = vacancies.filter(business_function__id__in=bf_ids)
                except (ValueError, TypeError):
                    pass
            
            department_values = request.query_params.getlist('department')
            if department_values:
                try:
                    dept_ids = []
                    for dept_val in department_values:
                        if ',' in dept_val:
                            dept_ids.extend([int(id.strip()) for id in dept_val.split(',') if id.strip().isdigit()])
                        elif dept_val.isdigit():
                            dept_ids.append(int(dept_val))
                    
                    if dept_ids:
                        vacancies = vacancies.filter(department__id__in=dept_ids)
                except (ValueError, TypeError):
                    pass
            
      
            
            # Convert vacancies to org chart format
            vacancy_data = []
            for vacancy in vacancies:
                if not vacancy.business_function or not vacancy.department or not vacancy.job_title:
                    logger.warning(f"⚠️ Vacancy {vacancy.id} missing required fields - skipping")
                    continue
                
                vac_data = {
                    'id': vacancy.id,
                    'employee_id': vacancy.position_id,
                    'name': f"[VACANT] {vacancy.job_title}",
                    'title': vacancy.job_title,
                    'avatar': 'VA',
                    'department': vacancy.department.name if vacancy.department else 'N/A',
                    'unit': vacancy.unit.name if vacancy.unit else None,
                    'business_function': vacancy.business_function.name if vacancy.business_function else 'N/A',
                    'position_group': vacancy.position_group.get_name_display() if vacancy.position_group else 'N/A',
                    'email': 'recruitment@company.com',
                    'phone': 'Position Open',
                    'line_manager_id': vacancy.reporting_to.employee_id if vacancy.reporting_to else None,
                    'direct_reports': 0,
                    'direct_reports_details': [],
                    'status_color': vacancy.vacancy_status.color if vacancy.vacancy_status else '#F97316',
                    'profile_image_url': None,
                    'level_to_ceo': 0,
                    'total_subordinates': 0,
                    'colleagues_in_unit': 0,
                    'colleagues_in_business_function': 0,
                    'manager_info': {
                        'id': vacancy.reporting_to.id,
                        'employee_id': vacancy.reporting_to.employee_id,
                        'name': vacancy.reporting_to.full_name,
                        'title': vacancy.reporting_to.job_title,
                        'avatar': self._generate_avatar(vacancy.reporting_to.full_name)
                    } if vacancy.reporting_to else None,
                    'employee_details': {
                        'internal_id': vacancy.id,
                        'employee_id': vacancy.position_id,
                        'is_vacancy': True,
                        'original_employee_pk': vacancy.original_employee_pk,
                        'is_visible_in_org_chart': vacancy.is_visible_in_org_chart,
                        'include_in_headcount': vacancy.include_in_headcount
                    }
                }
                vacancy_data.append(vac_data)
            
            # Combine data
            all_org_data = employee_data + vacancy_data
            
            # Statistics
            total_employees = employees.count()
            total_vacancies = len(vacancy_data)
            
           
            
            return Response({
                'org_chart': all_org_data,
                'statistics': {
                    'total_employees': total_employees,
                    'total_vacancies': total_vacancies,
                    'total_positions': total_employees + total_vacancies,
                    'filters_applied': len([k for k, v in request.query_params.items() if v and k not in ['format']]),
                    'filter_summary': self._get_filter_summary(employees, request.query_params)
                },
                'metadata': {
                    'generated_at': timezone.now(),
                    'includes_vacancies': True,
                    'filters_applied': {
                        'allows_org_chart': True,
                        'is_visible': True,
                        'is_deleted': False
                    }
                }
            })
            
        except Exception as e:
            logger.error(f"Error in org chart tree: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve org chart tree: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        operation_description="Get detailed information for a specific employee in org chart context",
        responses={
            200: openapi.Response(
                description="Employee detail in org chart context",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'employee': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'org_context': openapi.Schema(type=openapi.TYPE_OBJECT),
                        'hierarchy': openapi.Schema(type=openapi.TYPE_OBJECT)
                    }
                )
            ),
            404: "Employee not found"
        }
    )
    @action(detail=False, methods=['get'], url_path='detail/(?P<employee_pk>[^/.]+)')
    def get_employee_detail(self, request, employee_pk=None):
        """
        ✅ NEW: Get detailed employee information in org chart context
        
        This provides comprehensive details about a specific employee including:
        - Full employee data
        - Organizational context (manager, direct reports, peers)
        - Hierarchy information (level to CEO, subordinates count)
        """
        try:
            # Get employee
            try:
                employee = Employee.objects.select_related(
                    'user', 'business_function', 'department', 'unit', 'job_function',
                    'position_group', 'status', 'line_manager'
                ).prefetch_related('tags').get(pk=employee_pk)
            except Employee.DoesNotExist:
                return Response(
                    {'error': 'Employee not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Serialize main employee data
            serializer = OrgChartNodeSerializer(employee, context={'request': request})
            employee_data = serializer.data
            
            # Get organizational context
            
            # 1. Direct Reports (Team)
            direct_reports = Employee.objects.filter(
                line_manager=employee,
                status__allows_org_chart=True,
                is_deleted=False
            ).select_related('position_group', 'department')
            
            team_data = []
            for report in direct_reports:
                team_data.append({
                    'id': report.id,
                    'employee_id': report.employee_id,
                    'name': report.full_name,
                    'title': report.job_title,
                    'department': report.department.name if report.department else None,
                    'position_group': report.position_group.get_name_display() if report.position_group else None,
                    'profile_image_url': self._get_profile_image_url(report, request)
                })
            
            # 2. Manager Chain (Path to CEO)
            manager_chain = []
            current = employee.line_manager
            visited = set()
            max_depth = 10
            
            while current and current.id not in visited and len(manager_chain) < max_depth:
                visited.add(current.id)
                manager_chain.append({
                    'id': current.id,
                    'employee_id': current.employee_id,
                    'name': current.full_name,
                    'title': current.job_title,
                    'department': current.department.name if current.department else None,
                    'level': current.position_group.hierarchy_level if current.position_group else 0
                })
                current = current.line_manager
            
            # 3. Peers (Same manager)
            peers = []
            if employee.line_manager:
                peer_employees = Employee.objects.filter(
                    line_manager=employee.line_manager,
                    status__allows_org_chart=True,
                    is_deleted=False
                ).exclude(id=employee.id).select_related('position_group')[:10]
                
                for peer in peer_employees:
                    peers.append({
                        'id': peer.id,
                        'employee_id': peer.employee_id,
                        'name': peer.full_name,
                        'title': peer.job_title,
                        'position_group': peer.position_group.get_name_display() if peer.position_group else None
                    })
            
            # 4. Department colleagues
            department_colleagues_count = 0
            if employee.department:
                department_colleagues_count = Employee.objects.filter(
                    department=employee.department,
                    status__allows_org_chart=True,
                    is_deleted=False
                ).exclude(id=employee.id).count()
            
            # Calculate hierarchy metrics
            def count_total_subordinates(emp, visited=None):
                if visited is None:
                    visited = set()
                if emp.id in visited:
                    return 0
                visited.add(emp.id)
                
                reports = Employee.objects.filter(
                    line_manager=emp,
                    status__allows_org_chart=True,
                    is_deleted=False
                )
                total = reports.count()
                for report in reports:
                    total += count_total_subordinates(report, visited.copy())
                return total
            
            return Response({
                'employee': employee_data,
                'org_context': {
                    'direct_reports': {
                        'count': len(team_data),
                        'employees': team_data
                    },
                    'manager_chain': {
                        'levels_to_top': len(manager_chain),
                        'chain': manager_chain
                    },
                    'peers': {
                        'count': len(peers),
                        'sample': peers
                    },
                    'department_colleagues': department_colleagues_count
                },
                'hierarchy': {
                    'level_to_ceo': len(manager_chain),
                    'total_subordinates': count_total_subordinates(employee),
                    'direct_reports_count': len(team_data),
                    'has_team': len(team_data) > 0,
                    'is_top_level': employee.line_manager is None
                }
            })
            
        except Exception as e:
            logger.error(f"Error getting employee detail: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Failed to retrieve employee detail: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _get_filter_summary(self, queryset, params):
        """Get summary of applied filters"""
        summary = {}
        
        # Business function breakdown
        bf_counts = queryset.values('business_function__name').annotate(
            count=Count('id')
        ).order_by('-count')
        summary['by_business_function'] = {
            item['business_function__name']: item['count'] 
            for item in bf_counts if item['business_function__name']
        }
        
        # Department breakdown
        dept_counts = queryset.values('department__name').annotate(
            count=Count('id')
        ).order_by('-count')
        summary['by_department'] = {
            item['department__name']: item['count'] 
            for item in dept_counts if item['department__name']
        }
        
        # Position group breakdown
        pos_counts = queryset.values('position_group__name').annotate(
            count=Count('id')
        ).order_by('position_group__hierarchy_level')
        summary['by_position_group'] = {
            item['position_group__name']: item['count'] 
            for item in pos_counts if item['position_group__name']
        }
        
        return summary
    
    def _generate_avatar(self, full_name):
        """Generate avatar initials from full name"""
        if not full_name:
            return 'NA'
        
        words = full_name.strip().split()
        if len(words) >= 2:
            return f"{words[0][0]}{words[1][0]}".upper()
        elif len(words) == 1:
            return words[0][:2].upper()
        return 'NA'
    
    def _get_profile_image_url(self, employee, request):
        """Get profile image URL safely"""
        if employee.profile_image:
            try:
                if hasattr(employee.profile_image, 'url'):
                    return request.build_absolute_uri(employee.profile_image.url)
            except:
                pass
        return None

class ProfileImageViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    @swagger_auto_schema(
        operation_description="Upload or update employee profile image",
        manual_parameters=[
            openapi.Parameter(
                'employee_id',
                openapi.IN_FORM,
                description='Employee ID',
                type=openapi.TYPE_INTEGER,
                required=True
            ),
            openapi.Parameter(
                'profile_image',
                openapi.IN_FORM,
                description='Profile image file',
                type=openapi.TYPE_FILE,
                required=True
            )
        ],
        consumes=['multipart/form-data'],
        responses={
            200: openapi.Response(
                description="Profile image updated successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'profile_image_url': openapi.Schema(type=openapi.TYPE_STRING, nullable=True),
                    }
                )
            )
        }
    )
    @action(detail=False, methods=['post'])
    def upload(self, request):
        """Upload or update employee profile image"""
        try:
      
            
            # Validate request data
            if 'employee_id' not in request.data:
                return Response(
                    {'error': 'employee_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if 'profile_image' not in request.FILES:
                return Response(
                    {'error': 'profile_image file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
    
            
            serializer = ProfileImageUploadSerializer(data=request.data, context={'request': request})
            if not serializer.is_valid():
                logger.error(f"Profile image upload validation failed: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            employee = serializer.save()
            
            # Refresh employee from database to get the saved image
            employee.refresh_from_db()
            

            
            # Build profile image URL with better error handling
            profile_image_url = None
            if employee.profile_image:
                try:
                    # Check if the file exists and has a URL
                    if hasattr(employee.profile_image, 'url') and employee.profile_image.name:
                        profile_image_url = request.build_absolute_uri(employee.profile_image.url)
                       
                    else:
                        logger.warning(f"Profile image exists but no URL available: {employee.profile_image}")
                except Exception as e:
                  
                    # Try to construct URL manually
                    if employee.profile_image.name:
                        profile_image_url = request.build_absolute_uri(f"/media/{employee.profile_image.name}")
                     
            else:
                logger.warning("No profile image found after save")
            
     
            
            return Response({
                'success': True,
                'message': f'Profile image updated for {employee.full_name}',
                'employee_id': employee.id,
                'employee_name': employee.full_name,
                'profile_image_url': profile_image_url,
                'debug_info': {
                    'has_profile_image': bool(employee.profile_image),
                    'image_name': employee.profile_image.name if employee.profile_image else None,
                    'image_size': employee.profile_image.size if employee.profile_image else None
                }
            })
            
        except Exception as e:
            logger.error(f"Profile image upload failed: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Profile image upload failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    @swagger_auto_schema(
        operation_description="Delete employee profile image",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['employee_id'],
            properties={
                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER, description='Employee ID')
            }
        ),
        responses={200: "Profile image deleted successfully"}
    )
    @action(detail=False, methods=['post'], parser_classes=[JSONParser])
    def delete(self, request):
        """Delete employee profile image"""
        try:
            serializer = ProfileImageDeleteSerializer(data=request.data, context={'request': request})
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            employee = serializer.save()
            
            return Response({
                'success': True,
                'message': f'Profile image deleted for {employee.full_name}',
                'employee_id': employee.id,
                'employee_name': employee.full_name
            })
            
        except Exception as e:
            logger.error(f"Profile image delete failed: {str(e)}")
            return Response(
                {'error': f'Profile image delete failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
  
        """Get employee profile image URL"""
        try:
            employee_id = request.query_params.get('employee_id')
            if not employee_id:
                return Response(
                    {'error': 'employee_id parameter is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            try:
                employee = Employee.objects.get(id=employee_id)
            except Employee.DoesNotExist:
                return Response(
                    {'error': 'Employee not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
           
            
            profile_image_url = None
            has_image = False
            
            if employee.profile_image:
                try:
                    if hasattr(employee.profile_image, 'url') and employee.profile_image.name:
                        profile_image_url = request.build_absolute_uri(employee.profile_image.url)
                        has_image = True
                        logger.info(f"Found profile image URL: {profile_image_url}")
                    else:
                        logger.warning(f"Profile image exists but no URL: {employee.profile_image}")
                except Exception as e:
                    logger.error(f"Error getting profile image URL: {e}")
                    # Try manual construction
                    if employee.profile_image.name:
                        profile_image_url = request.build_absolute_uri(f"/media/{employee.profile_image.name}")
                        has_image = True
                        logger.info(f"Manually constructed URL: {profile_image_url}")
            
            return Response({
                'employee_id': employee.id,
                'employee_name': employee.full_name,
                'profile_image_url': profile_image_url,
                'has_image': has_image,
                'debug_info': {
                    'image_field_value': str(employee.profile_image),
                    'image_name': employee.profile_image.name if employee.profile_image else None,
                    'image_size': employee.profile_image.size if employee.profile_image else None
                }
            })
            
        except Exception as e:
            logger.error(f"Get profile image failed: {str(e)}")
            return Response(
                {'error': f'Failed to get profile image: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )