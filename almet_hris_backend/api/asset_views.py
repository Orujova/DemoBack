# api/asset_views.py - COMPLETE REWRITE

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count
from django.db import transaction
from django.utils import timezone
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
import logging
import pandas as pd
import traceback

logger = logging.getLogger(__name__)

from .asset_models import (
    AssetCategory, AssetBatch, Asset, AssetAssignment, 
    AssetActivity, EmployeeOffboarding, AssetTransferRequest
)
from .asset_serializers import (
    AssetCategorySerializer, 
    AssetBatchListSerializer, AssetBatchDetailSerializer, AssetBatchCreateSerializer,
    AssetListSerializer, AssetDetailSerializer, AssetCreateSerializer, AssetCreateMultipleSerializer,
    AssetAssignmentSerializer, AssetAssignmentCreateSerializer,
    AssetActivitySerializer,
    AssetAcceptanceSerializer, AssetClarificationRequestSerializer,
    AssetCancellationSerializer, AssetClarificationProvisionSerializer,
    EmployeeOffboardingSerializer,
    AssetTransferRequestSerializer, AssetTransferRequestCreateSerializer,
    AssetBulkUploadSerializer
)
from .asset_permissions import (
    get_asset_access_level, filter_assets_by_access, filter_batches_by_access,
    require_asset_permission, can_user_manage_asset, get_access_summary
)
from .models import Employee
from .system_email_service import system_email_service


# ============================================
# CATEGORY VIEWSET
# ============================================
class AssetCategoryViewSet(viewsets.ModelViewSet):

    
    queryset = AssetCategory.objects.all()
    serializer_class = AssetCategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Kateqoriya √ºzr…ô statistika"""
        category = self.get_object()
        
        total_batches = category.batches.count()
        total_assets = Asset.objects.filter(category=category).count()
        
        # Status breakdown
        status_breakdown = {}
        for status_choice in Asset.STATUS_CHOICES:
            status_code = status_choice[0]
            count = Asset.objects.filter(category=category, status=status_code).count()
            if count > 0:
                status_breakdown[status_code] = {
                    'label': status_choice[1],
                    'count': count
                }
        
        return Response({
            'category': category.name,
            'total_batches': total_batches,
            'total_assets': total_assets,
            'status_breakdown': status_breakdown
        })


# ============================================
# BATCH VIEWSET
# ============================================
class AssetBatchViewSet(viewsets.ModelViewSet):

    
    queryset = AssetBatch.objects.select_related('category', 'created_by').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'status']
    search_fields = ['batch_number', 'asset_name', 'supplier']
    ordering_fields = ['created_at', 'asset_name', 'available_quantity']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AssetBatchListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return AssetBatchCreateSerializer
        return AssetBatchDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return filter_batches_by_access(self.request.user, queryset)
    
    @swagger_auto_schema(
        request_body=AssetBatchCreateSerializer,
        responses={
            201: openapi.Response(
                description="Batch yaradƒ±ldƒ±",
                schema=AssetBatchDetailSerializer
            )
        }
    )
    @require_asset_permission('create')
    def create(self, request, *args, **kwargs):
        """
        üéØ Batch yaratma - SAY BURADAN QEYD EDƒ∞Lƒ∞R
        
        N√ºmun…ô:
        {
            "asset_name": "Dell Latitude 5420",
            "category": 1,
            "initial_quantity": 10,  üëà BURADAN SAY
            "unit_price": 1500.00,
            "purchase_date": "2024-01-15",
            "useful_life_years": 5
        }
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        batch = serializer.save(created_by=request.user)
        
        logger.info(
            f"‚úÖ Batch yaradƒ±ldƒ±: {batch.batch_number} - {batch.asset_name} | "
            f"Quantity: {batch.initial_quantity} | "
            f"Value: {batch.total_value}"
        )
        
        return Response({
            'success': True,
            'message': f'Batch yaradƒ±ldƒ±: {batch.batch_number}',
            'batch': AssetBatchDetailSerializer(batch, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetCreateMultipleSerializer,
        responses={
            200: openapi.Response(
                description="Asset-l…ôr yaradƒ±ldƒ±",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        'message': openapi.Schema(type=openapi.TYPE_STRING),
                        'created_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'assets': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT))
                    }
                )
            )
        }
    )
    @action(detail=True, methods=['post'], url_path='create-assets')
    @require_asset_permission('manage')
    def create_assets_from_batch(self, request, pk=None):
        """
        üéØ Batch-d…ôn asset yaratma
        
        N√ºmun…ô:
        {
            "serial_numbers": ["SN001", "SN002", "SN003"]
        }
        
        Prosess:
        1. Batch-in available_quantity yoxlanƒ±lƒ±r
        2. Asset-l…ôr yaradƒ±lƒ±r
        3. Batch-in available_quantity azalƒ±r
        """
        try:
            batch = self.get_object()
            
            # Serialize
            serializer = AssetCreateMultipleSerializer(
                data={**request.data, 'batch_id': batch.id},
                context={'request': request}
            )
            serializer.is_valid(raise_exception=True)
            
            # Create assets
            created_assets = serializer.save()
            
            return Response({
                'success': True,
                'message': f'{len(created_assets)} asset yaradƒ±ldƒ±',
                'created_count': len(created_assets),
                'batch': {
                    'batch_number': batch.batch_number,
                    'available_quantity': batch.available_quantity,
                    'assigned_quantity': batch.assigned_quantity
                },
                'assets': AssetListSerializer(created_assets, many=True, context={'request': request}).data
            })
            
        except Exception as e:
            logger.error(f"‚ùå Batch-d…ôn asset yaratma x…ôtasƒ±: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def assets(self, request, pk=None):
        """Batch-d…ôki b√ºt√ºn asset-l…ôr"""
        batch = self.get_object()
        assets = batch.assets.all()
        
        # Status breakdown
        status_summary = {}
        for choice in Asset.STATUS_CHOICES:
            count = assets.filter(status=choice[0]).count()
            if count > 0:
                status_summary[choice[0]] = {
                    'status': choice[1],
                    'count': count
                }
        
        return Response({
            'batch_number': batch.batch_number,
            'batch_name': batch.asset_name,
            'quantity_summary': batch.get_quantity_summary(),
            'total_assets': assets.count(),
            'status_summary': status_summary,
            'assets': AssetListSerializer(assets, many=True, context={'request': request}).data
        })
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Batch statistikasƒ±"""
        queryset = self.get_queryset()
        
        total_batches = queryset.count()
        active_batches = queryset.filter(status='ACTIVE').count()
        
        # Quantity summary
        quantity_summary = queryset.aggregate(
            total_initial=Sum('initial_quantity'),
            total_available=Sum('available_quantity'),
            total_assigned=Sum('assigned_quantity'),
            total_out_of_stock=Sum('out_of_stock_quantity')
        )
        
        # Financial summary
        financial_summary = queryset.aggregate(
            total_value=Sum('total_value')
        )
        
        return Response({
            'total_batches': total_batches,
            'active_batches': active_batches,
            'quantity_summary': {
                'total_initial': quantity_summary['total_initial'] or 0,
                'total_available': quantity_summary['total_available'] or 0,
                'total_assigned': quantity_summary['total_assigned'] or 0,
                'total_out_of_stock': quantity_summary['total_out_of_stock'] or 0
            },
            'total_value': float(financial_summary['total_value'] or 0)
        })


# ============================================
# ASSET VIEWSET - MAIN
# ============================================
class AssetViewSet(viewsets.ModelViewSet):

    queryset = Asset.objects.select_related(
        'batch', 'category', 'assigned_to', 'created_by', 'updated_by'
    ).all()
    
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'category', 'batch', 'assigned_to']
    search_fields = ['asset_number', 'serial_number', 'asset_name', 'batch__batch_number']
    ordering_fields = ['created_at', 'asset_name', 'status', 'updated_at']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return AssetListSerializer
        elif self.action == 'create':
            return AssetCreateSerializer
        return AssetDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return filter_assets_by_access(self.request.user, queryset)
    
    @swagger_auto_schema(
        request_body=AssetCreateSerializer,
        responses={201: AssetDetailSerializer}
    )
    @require_asset_permission('create')
    def create(self, request, *args, **kwargs):
        """
        üéØ Asset yaratma
        
        N√ºmun…ô:
        {
            "batch_id": 5,
            "serial_number": "SN123456"
        }
        
        ‚ö†Ô∏è Asset yaradanda batch-in available_quantity avtomatik azalƒ±r
        """
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            asset = serializer.save()
            
            # Log activity
            AssetActivity.objects.create(
                asset=asset,
                activity_type='CREATED',
                description=f"Asset batch-d…ôn yaradƒ±ldƒ±: {asset.batch.batch_number}",
                performed_by=request.user,
                metadata={
                    'batch_number': asset.batch.batch_number,
                    'batch_id': asset.batch.id,
                    'creation_method': 'manual'
                }
            )
            
            logger.info(f"‚úÖ Asset yaradƒ±ldƒ±: {asset.asset_number} from {asset.batch.batch_number}")
            
            return Response({
                'success': True,
                'message': f'Asset yaradƒ±ldƒ±: {asset.asset_number}',
                'asset': AssetDetailSerializer(asset, context={'request': request}).data
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"‚ùå Asset yaratma x…ôtasƒ±: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetAssignmentCreateSerializer,
        responses={200: openapi.Response(description="Assets assigned")}
    )
    @action(detail=False, methods=['post'], url_path='assign-to-employee')
    @require_asset_permission('manage')
    def assign_to_employee(self, request):
        """
        üéØ Asset-l…ôri i≈ü√ßiy…ô t…ôyin etm…ô
        
        N√ºmun…ô:
        {
            "asset_ids": ["uuid1", "uuid2"],
            "employee_id": 123,
            "check_out_date": "2024-01-15",
            "check_out_notes": "Yeni laptop",
            "condition_on_checkout": "GOOD"
        }
        
        Prosess:
        1. Asset status: IN_STOCK ‚Üí ASSIGNED
        2. Asset assigned_to = employee
        3. Assignment record yaradƒ±lƒ±r
        4. Email g√∂nd…ôrilir
        
        ‚ö†Ô∏è Batch quantity burada d…ôyi≈ümir (artƒ±q create-d…ô azalƒ±b)
        """
        try:
            serializer = AssetAssignmentCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            employee = serializer.validated_data['employee']
            assets = serializer.validated_data['assets']
            check_out_date = serializer.validated_data['check_out_date']
            check_out_notes = serializer.validated_data.get('check_out_notes', '')
            condition = serializer.validated_data['condition_on_checkout']
            
            assignments_created = []
            
            with transaction.atomic():
                for asset in assets:
                    # Create assignment
                    assignment = AssetAssignment.objects.create(
                        asset=asset,
                        employee=employee,
                        check_out_date=check_out_date,
                        check_out_notes=check_out_notes,
                        condition_on_checkout=condition,
                        assigned_by=request.user
                    )
                    
                    # Update asset
                    asset.status = 'ASSIGNED'
                    asset.assigned_to = employee
                    asset.updated_by = request.user
                    asset.save()
                    
                    # Log activity
                    AssetActivity.objects.create(
                        asset=asset,
                        activity_type='ASSIGNED',
                        description=f"ƒ∞≈ü√ßiy…ô t…ôyin edildi: {employee.full_name} - t…ôsdiq g√∂zl…ônilir",
                        performed_by=request.user,
                        metadata={
                            'employee_id': employee.employee_id,
                            'employee_name': employee.full_name,
                            'check_out_date': check_out_date.isoformat(),
                            'condition': condition
                        }
                    )
                    
                    assignments_created.append(assignment)
            
            # Send email notification
            self._send_assignment_email(employee, assets, request.user)
            
            logger.info(f"‚úÖ {len(assets)} asset t…ôyin edildi ‚Üí {employee.full_name}")
            
            return Response({
                'success': True,
                'message': f'{len(assets)} asset t…ôyin edildi: {employee.full_name}',
                'employee': {
                    'id': employee.id,
                    'name': employee.full_name,
                    'employee_id': employee.employee_id
                },
                'assignments': AssetAssignmentSerializer(assignments_created, many=True, context={'request': request}).data
            })
            
        except Exception as e:
            logger.error(f"‚ùå Assignment x…ôtasƒ±: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _send_assignment_email(self, employee, assets, assigned_by):
        """Email notification"""
        try:
            if not employee.user or not employee.user.email:
                logger.warning(f"‚ö†Ô∏è Employee {employee.full_name} - email yoxdur")
                return
            
            asset_list = '<ul>' + ''.join([
                f'<li><strong>{asset.asset_name}</strong> - {asset.serial_number}</li>'
                for asset in assets
            ]) + '</ul>'
            
            html_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #2563eb;">Asset T…ôyinatƒ±</h2>
                <p>H√∂rm…ôtli {employee.full_name},</p>
                <p>Siz…ô a≈üaƒüƒ±daki asset-l…ôr t…ôyin edilmi≈üdir:</p>
                {asset_list}
                <p><strong>Z…ôhm…ôt olmasa t…ôsdiq edin:</strong></p>
                <ul>
                    <li>‚úÖ Q…ôbul et - H…ôr ≈üey d√ºzg√ºnd√ºrs…ô</li>
                    <li>‚ùì Aydƒ±nla≈üdƒ±rma sorƒüusu - Sualƒ±nƒ±z varsa</li>
                </ul>
                <p>T…ôyin ed…ôn: <strong>{assigned_by.get_full_name() or assigned_by.username}</strong></p>
                <p>HRIS sistemin…ô daxil olub t…ôsdiq ed…ô bil…ôrsiniz.</p>
                <hr>
                <p style="color: #6b7280; font-size: 12px;">Bu avtomatik mesajdƒ±r</p>
            </body>
            </html>
            """
            
            system_email_service.send_email_as_system(
                from_email='myalmet@almettrading.com',
                to_email=employee.user.email,
                subject=f'Asset T…ôyinatƒ± - {len(assets)} …ô≈üya',
                body_html=html_body
            )
            
            logger.info(f"‚úÖ Email g√∂nd…ôrildi ‚Üí {employee.user.email}")
            
        except Exception as e:
            logger.error(f"‚ùå Email x…ôtasƒ±: {str(e)}")
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetAcceptanceSerializer,
        responses={200: openapi.Response(description="Asset accepted")}
    )
    @action(detail=False, methods=['post'], url_path='accept-assignment')
    def accept_assignment(self, request):
        """
        üéØ ƒ∞≈ü√ßi asset-i q…ôbul edir
        
        Prosess:
        1. Asset status: ASSIGNED ‚Üí IN_USE
        2. Activity log
        """
        try:
            asset_id = request.data.get('asset_id')
            comments = request.data.get('comments', '')
            
            if not asset_id:
                return Response(
                    {'error': 'asset_id t…ôl…ôb olunur'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            asset = Asset.objects.get(id=asset_id)
            
            # Check permission
            access = get_asset_access_level(request.user)
            if not access['employee'] or asset.assigned_to != access['employee']:
                return Response(
                    {'error': 'Bu asset-i yalnƒ±z siz…ô t…ôyin edilmi≈ü olsa q…ôbul ed…ô bil…ôrsiniz'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            if not asset.can_be_approved():
                return Response(
                    {'error': f'Asset q…ôbul edil…ô bilm…ôz. Status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                asset.status = 'IN_USE'
                asset.updated_by = request.user
                asset.save()
                
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='ACCEPTED',
                    description=f"ƒ∞≈ü√ßi t…ôr…ôfind…ôn q…ôbul edildi: {access['employee'].full_name}",
                    performed_by=request.user,
                    metadata={
                        'comments': comments,
                        'accepted_at': timezone.now().isoformat()
                    }
                )
            
            logger.info(f"‚úÖ Asset q…ôbul edildi: {asset.asset_number} by {access['employee'].full_name}")
            
            return Response({
                'success': True,
                'message': 'Asset uƒüurla q…ôbul edildi',
                'asset_id': str(asset.id),
                'asset_number': asset.asset_number
            })
            
        except Asset.DoesNotExist:
            return Response({'error': 'Asset tapƒ±lmadƒ±'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"‚ùå Accept x…ôtasƒ±: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


    @action(detail=False, methods=['get'], url_path='assignments')
    def assignment_history_list(self, request):
        """
        üéØ Assignment History - B√ºt√ºn assignment-l…ôr
        
        Filters:
        - employee_id: ƒ∞≈ü√ßiy…ô g√∂r…ô
        - asset_id: Asset-…ô g√∂r…ô
        - date_from: Ba≈ülanƒüƒ±c tarixi
        - date_to: Bitm…ô tarixi
        - is_active: Aktiv/completed
        """
        try:
            # Get all assignments
            queryset = AssetAssignment.objects.select_related(
                'asset', 'asset__category', 'asset__batch',
                'employee', 'assigned_by', 'checked_in_by'
            ).all()
            
            # Apply filters
            employee_id = request.query_params.get('employee_id')
            if employee_id:
                queryset = queryset.filter(employee_id=employee_id)
            
            asset_id = request.query_params.get('asset_id')
            if asset_id:
                queryset = queryset.filter(asset_id=asset_id)
            
            date_from = request.query_params.get('date_from')
            if date_from:
                queryset = queryset.filter(check_out_date__gte=date_from)
            
            date_to = request.query_params.get('date_to')
            if date_to:
                queryset = queryset.filter(check_out_date__lte=date_to)
            
            is_active = request.query_params.get('is_active')
            if is_active == 'true':
                queryset = queryset.filter(check_in_date__isnull=True)
            elif is_active == 'false':
                queryset = queryset.filter(check_in_date__isnull=False)
            
            # Search
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(asset__asset_name__icontains=search) |
                    Q(asset__serial_number__icontains=search) |
                    Q(employee__full_name__icontains=search) |
                    Q(employee__employee_id__icontains=search)
                )
            
            # Ordering
            queryset = queryset.order_by('-check_out_date')
            
            # Pagination
            page_size = int(request.query_params.get('page_size', 15))
            page = int(request.query_params.get('page', 1))
            
            from django.core.paginator import Paginator
            paginator = Paginator(queryset, page_size)
            page_obj = paginator.get_page(page)
            
            return Response({
                'count': paginator.count,
                'total_pages': paginator.num_pages,
                'current_page': page,
                'page_size': page_size,
                'results': AssetAssignmentSerializer(
                    page_obj, 
                    many=True, 
                    context={'request': request}
                ).data
            })
            
        except Exception as e:
            logger.error(f"‚ùå Assignment history error: {str(e)}")
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='assignments/export')
    def export_assignments(self, request):
        """
        üéØ Export Assignments to Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse
            
            # Get filtered queryset (same filters as assignment_history_list)
            queryset = AssetAssignment.objects.select_related(
                'asset', 'employee', 'assigned_by', 'checked_in_by'
            ).all()
            
            # Apply filters from request body
            employee_id = request.data.get('employee_id')
            if employee_id:
                queryset = queryset.filter(employee_id=employee_id)
            
            date_from = request.data.get('date_from')
            if date_from:
                queryset = queryset.filter(check_out_date__gte=date_from)
            
            date_to = request.data.get('date_to')
            if date_to:
                queryset = queryset.filter(check_out_date__lte=date_to)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Assignments"
            
            # Header style
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Headers
            headers = [
                'Asset Name', 'Serial Number', 'Category',
                'Employee', 'Employee ID', 'Department',
                'Check Out Date', 'Check In Date', 'Duration (days)',
                'Condition Out', 'Condition In', 'Status',
                'Assigned By', 'Checked In By'
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, assignment in enumerate(queryset, start=2):
                ws.cell(row=row_idx, column=1, value=assignment.asset.asset_name)
                ws.cell(row=row_idx, column=2, value=assignment.asset.serial_number)
                ws.cell(row=row_idx, column=3, value=assignment.asset.category.name)
                ws.cell(row=row_idx, column=4, value=assignment.employee.full_name)
                ws.cell(row=row_idx, column=5, value=assignment.employee.employee_id)
                ws.cell(row=row_idx, column=6, value=assignment.employee.department or 'N/A')
                ws.cell(row=row_idx, column=7, value=assignment.check_out_date.strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=8, value=assignment.check_in_date.strftime('%Y-%m-%d') if assignment.check_in_date else 'Active')
                ws.cell(row=row_idx, column=9, value=assignment.duration_days)
                ws.cell(row=row_idx, column=10, value=assignment.condition_on_checkout)
                ws.cell(row=row_idx, column=11, value=assignment.condition_on_checkin or 'N/A')
                ws.cell(row=row_idx, column=12, value='Active' if assignment.is_active else 'Completed')
                ws.cell(row=row_idx, column=13, value=assignment.assigned_by.get_full_name())
                ws.cell(row=row_idx, column=14, value=assignment.checked_in_by.get_full_name() if assignment.checked_in_by else 'N/A')
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=assignments_{timezone.now().strftime("%Y%m%d")}.xlsx'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"‚ùå Export assignments error: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='activities')
    def activities(self, request, pk=None):
        """
        üéØ Asset Activities Log
        """
        asset = self.get_object()
        activities = asset.activities.select_related(
            'performed_by'
        ).order_by('-performed_at')
        
        return Response({
            'asset': {
                'id': str(asset.id),
                'asset_number': asset.asset_number,
                'asset_name': asset.asset_name
            },
            'activities': AssetActivitySerializer(
                activities, 
                many=True, 
                context={'request': request}
            ).data
        })
        
    @swagger_auto_schema(
        method='post',
        request_body=AssetClarificationRequestSerializer,
        responses={200: openapi.Response(description="Clarification requested")}
    )
    @action(detail=False, methods=['post'], url_path='request-clarification')
    def request_clarification(self, request):
        """
        üéØ ƒ∞≈ü√ßi aydƒ±nla≈üdƒ±rma sorƒüusu g√∂nd…ôrir
        
        Prosess:
        1. Asset status: ASSIGNED ‚Üí NEED_CLARIFICATION
        2. Clarification m…ôlumatlarƒ± saxlanƒ±lƒ±r
        """
        try:
            asset_id = request.data.get('asset_id')
            reason = request.data.get('clarification_reason')
            
            if not asset_id or not reason:
                return Response(
                    {'error': 'asset_id v…ô clarification_reason t…ôl…ôb olunur'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            asset = Asset.objects.get(id=asset_id)
            
            access = get_asset_access_level(request.user)
            if not access['employee'] or asset.assigned_to != access['employee']:
                return Response(
                    {'error': 'Bu asset-i yalnƒ±z siz…ô t…ôyin edilmi≈ü olsa aydƒ±nla≈üdƒ±rma sorƒüusu g√∂nd…ôr…ô bil…ôrsiniz'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            if asset.status not in ['ASSIGNED', 'NEED_CLARIFICATION']:
                return Response(
                    {'error': f'Aydƒ±nla≈üdƒ±rma sorƒüusu g√∂nd…ôril…ô bilm…ôz. Status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                asset.status = 'NEED_CLARIFICATION'
                asset.clarification_requested_reason = reason
                asset.clarification_requested_at = timezone.now()
                asset.clarification_requested_by = request.user
                asset.clarification_response = None
                asset.clarification_provided_at = None
                asset.clarification_provided_by = None
                asset.updated_by = request.user
                asset.save()
                
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='CLARIFICATION_REQUESTED',
                    description=f"Aydƒ±nla≈üdƒ±rma sorƒüusu: {access['employee'].full_name}",
                    performed_by=request.user,
                    metadata={'reason': reason}
                )
            
            logger.info(f"‚úÖ Aydƒ±nla≈üdƒ±rma sorƒüusu: {asset.asset_number}")
            
            return Response({
                'success': True,
                'message': 'Aydƒ±nla≈üdƒ±rma sorƒüusu g√∂nd…ôrildi',
                'asset_id': str(asset.id)
            })
            
        except Asset.DoesNotExist:
            return Response({'error': 'Asset tapƒ±lmadƒ±'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"‚ùå Clarification x…ôtasƒ±: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetClarificationProvisionSerializer,
        responses={200: openapi.Response(description="Clarification provided")}
    )
    @action(detail=False, methods=['post'], url_path='provide-clarification')
    @require_asset_permission('manage')
    def provide_clarification(self, request):
        """
        üéØ Admin/Manager aydƒ±nla≈üdƒ±rma cavabƒ± verir
        
        Prosess:
        1. Asset status: NEED_CLARIFICATION ‚Üí ASSIGNED
        2. Cavab saxlanƒ±lƒ±r
        """
        try:
            asset_id = request.data.get('asset_id')
            response_text = request.data.get('clarification_response')
            
            if not asset_id or not response_text:
                return Response(
                    {'error': 'asset_id v…ô clarification_response t…ôl…ôb olunur'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            asset = Asset.objects.get(id=asset_id)
            
            if asset.status != 'NEED_CLARIFICATION':
                return Response(
                    {'error': f'Asset aydƒ±nla≈üdƒ±rma g√∂zl…ômir. Status: {asset.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            with transaction.atomic():
                asset.status = 'ASSIGNED'
                asset.clarification_response = response_text
                asset.clarification_provided_at = timezone.now()
                asset.clarification_provided_by = request.user
                asset.updated_by = request.user
                asset.save()
                
                AssetActivity.objects.create(
                    asset=asset,
                    activity_type='CLARIFICATION_PROVIDED',
                    description=f"Aydƒ±nla≈üdƒ±rma cavabƒ± verildi",
                    performed_by=request.user,
                    metadata={'response': response_text}
                )
            
            logger.info(f"‚úÖ Aydƒ±nla≈üdƒ±rma cavabƒ±: {asset.asset_number}")
            
            return Response({
                'success': True,
                'message': 'Aydƒ±nla≈üdƒ±rma cavabƒ± verildi',
                'asset_id': str(asset.id)
            })
            
        except Asset.DoesNotExist:
            return Response({'error': 'Asset tapƒ±lmadƒ±'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"‚ùå Provide clarification x…ôtasƒ±: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def assignment_history(self, request, pk=None):
        """Asset-in assignment tarix√ß…ôsi"""
        asset = self.get_object()
        assignments = asset.assignments.all().order_by('-check_out_date')
        
        return Response({
            'asset': {
                'id': str(asset.id),
                'asset_number': asset.asset_number,
                'asset_name': asset.asset_name,
                'serial_number': asset.serial_number
            },
            'assignments': AssetAssignmentSerializer(assignments, many=True, context={'request': request}).data
        })
    
    @action(detail=False, methods=['get'])
    def my_assets(self, request):
        """ƒ∞stifad…ô√ßinin √∂z asset-l…ôri"""
        access = get_asset_access_level(request.user)
        
        if not access['employee']:
            return Response({'assets': [], 'message': 'Sizin i≈ü√ßi profiliniz yoxdur'})
        
        assets = Asset.objects.filter(
            assigned_to=access['employee']
        ).select_related('batch', 'category')
        
        return Response({
            'employee': {
                'id': access['employee'].id,
                'name': access['employee'].full_name,
                'employee_id': access['employee'].employee_id
            },
            'total_assets': assets.count(),
            'assets': AssetListSerializer(assets, many=True, context={'request': request}).data
        })
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetBulkUploadSerializer,
        manual_parameters=[
            openapi.Parameter(
                'file',
                openapi.IN_FORM,
                description='Excel/CSV file',
                type=openapi.TYPE_FILE,
                required=True
            )
        ],
        consumes=['multipart/form-data']
    )
    @action(detail=False, methods=['post'], url_path='bulk-upload', parser_classes=[MultiPartParser, FormParser])
    @require_asset_permission('create')
    def bulk_upload(self, request):
        """
        üéØ Excel/CSV-d…ôn bulk upload
        
        Excel format:
        | asset_name | category | quantity | serial_numbers | unit_price | purchase_date | supplier |
        """
        try:
            serializer = AssetBulkUploadSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            file = serializer.validated_data['file']
            
            # Read file
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            required_columns = ['asset_name', 'category', 'quantity', 'unit_price', 'purchase_date']
            missing = set(required_columns) - set(df.columns)
            
            if missing:
                return Response(
                    {'error': f'Lazƒ±mi s√ºtunlar yoxdur: {", ".join(missing)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            results = {'success': 0, 'failed': 0, 'errors': []}
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Get or create category
                        category, _ = AssetCategory.objects.get_or_create(
                            name=row['category'],
                            defaults={'created_by': request.user}
                        )
                        
                        # Create batch
                        batch = AssetBatch.objects.create(
                            asset_name=row['asset_name'],
                            category=category,
                            initial_quantity=int(row['quantity']),
                            available_quantity=int(row['quantity']),
                            unit_price=float(row['unit_price']),
                            purchase_date=pd.to_datetime(row['purchase_date']).date(),
                            useful_life_years=int(row.get('useful_life_years', 5)),
                            supplier=row.get('supplier', ''),
                            created_by=request.user
                        )
                        
                        results['success'] += 1
                        
                    except Exception as e:
                        results['failed'] += 1
                        results['errors'].append(f"S…ôtir {index + 2}: {str(e)}")
            
            logger.info(f"‚úÖ Bulk upload: {results['success']} uƒüurlu, {results['failed']} uƒüursuz")
            
            return Response({
                'success': True,
                'imported': results['success'],
                'failed': results['failed'],
                'errors': results['errors'][:10]
            })
            
        except Exception as e:
            logger.error(f"‚ùå Bulk upload x…ôtasƒ±: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def access_info(self, request):
        """ƒ∞stifad…ô√ßinin icaz…ô m…ôlumatlarƒ±"""
        return Response(get_access_summary(request.user))
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Asset statistikasƒ±"""
        queryset = self.get_queryset()
        
        total_assets = queryset.count()
        
        # Status breakdown
        status_breakdown = {}
        for choice in Asset.STATUS_CHOICES:
            count = queryset.filter(status=choice[0]).count()
            if count > 0:
                status_breakdown[choice[0]] = {
                    'label': choice[1],
                    'count': count,
                    'percentage': round((count / total_assets * 100), 1) if total_assets > 0 else 0
                }
        
        # Category breakdown
        category_breakdown = {}
        categories = queryset.values('category__name').annotate(count=Count('id'))
        for cat in categories:
            if cat['category__name']:
                category_breakdown[cat['category__name']] = cat['count']
        
        # Assignment breakdown
        assigned_count = queryset.filter(assigned_to__isnull=False).count()
        unassigned_count = total_assets - assigned_count
        
        return Response({
            'total_assets': total_assets,
            'status_breakdown': status_breakdown,
            'category_breakdown': category_breakdown,
            'assignment_summary': {
                'assigned': assigned_count,
                'unassigned': unassigned_count,
                'assignment_rate': round((assigned_count / total_assets * 100), 1) if total_assets > 0 else 0
            }
        })

class EmployeeOffboardingViewSet(viewsets.ModelViewSet):

    queryset = EmployeeOffboarding.objects.select_related(
        'employee', 'created_by', 'approved_by', 'it_handover_completed_by'
    ).all()
    serializer_class = EmployeeOffboardingSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-created_at']
    
    def get_queryset(self):
        access = get_asset_access_level(self.request.user)
        queryset = super().get_queryset()
        
        if access['can_view_all_assets']:
            return queryset
        
        if access['accessible_employee_ids']:
            return queryset.filter(employee_id__in=access['accessible_employee_ids'])
        
        return queryset.none()
    
    @swagger_auto_schema(
        method='post',
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'employee_id': openapi.Schema(type=openapi.TYPE_INTEGER),
                'last_working_day': openapi.Schema(type=openapi.TYPE_STRING, format='date'),
                'offboarding_type': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    enum=['TRANSFER', 'RETURN']
                ),
                'notes': openapi.Schema(type=openapi.TYPE_STRING)
            }
        )
    )
    @action(detail=False, methods=['post'], url_path='initiate')
    @require_asset_permission('manage')
    def initiate_offboarding(self, request):
        """Offboarding prosesini ba≈ülat"""
        try:
            employee_id = request.data.get('employee_id')
            last_working_day = request.data.get('last_working_day')
            offboarding_type = request.data.get('offboarding_type', 'RETURN')
            
            employee = Employee.objects.get(id=employee_id, is_deleted=False)
            
            # Count assets
            assets = Asset.objects.filter(
                assigned_to=employee, 
                status__in=['ASSIGNED', 'IN_USE']
            )
            total_assets = assets.count()
            
            offboarding = EmployeeOffboarding.objects.create(
                employee=employee,
                last_working_day=last_working_day,
                offboarding_type=offboarding_type,
                total_assets=total_assets,
                notes=request.data.get('notes', ''),
                created_by=request.user
            )
            
            # üìß Send email to IT based on type
            self._send_it_offboarding_notification(offboarding, assets)
            
            logger.info(
                f"‚úÖ Offboarding started: {employee.full_name} - "
                f"{total_assets} assets - Type: {offboarding_type}"
            )
            
            return Response({
                'success': True,
                'offboarding_id': offboarding.id,
                'employee': employee.full_name,
                'total_assets': total_assets,
                'offboarding_type': offboarding_type
            })
            
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employee not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"‚ùå Offboarding error: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _send_it_offboarding_notification(self, offboarding, assets):
        """
        üìß Send email to IT department about offboarding
        """
        try:
            # Get IT team emails
            from .role_models import EmployeeRole
            it_roles = EmployeeRole.objects.filter(
                role__name__icontains='IT',
                role__is_active=True,
                is_active=True
            ).select_related('employee', 'employee__user')
            
            it_emails = []
            for role in it_roles:
                if role.employee.user and role.employee.user.email:
                    it_emails.append(role.employee.user.email)
            
            if not it_emails:
                logger.warning("‚ö†Ô∏è No IT emails found for offboarding notification")
                return
            
            # Generate asset list
            asset_list_html = '<ul>'
            for asset in assets:
                asset_list_html += f'''
                <li>
                    <strong>{asset.asset_name}</strong><br>
                    Serial: {asset.serial_number}<br>
                    Asset #: {asset.asset_number}
                </li>
                '''
            asset_list_html += '</ul>'
            
            # Generate handover link
            handover_link = f"https://yourdomain.com/assets/offboarding/{offboarding.id}/complete"
            
            if offboarding.offboarding_type == 'TRANSFER':
                # TRANSFER - Assets will be transferred
                subject = f"Asset Transfer Required - {offboarding.employee.full_name} Offboarding"
                html_body = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2 style="color: #2563eb;">Asset Transfer - Employee Offboarding</h2>
                    
                    <p><strong>Employee:</strong> {offboarding.employee.full_name} (ID: {offboarding.employee.employee_id})</p>
                    <p><strong>Last Working Day:</strong> {offboarding.last_working_day.strftime('%B %d, %Y')}</p>
                    <p><strong>Total Assets:</strong> {offboarding.total_assets}</p>
                    
                    <hr>
                    
                    <h3>Assets to be Transferred:</h3>
                    {asset_list_html}
                    
                    <p style="background-color: #FEF3C7; padding: 15px; border-left: 4px solid #F59E0B;">
                        <strong>‚ö†Ô∏è Action Required:</strong><br>
                        These assets will be transferred to other employees. 
                        Transfer requests will be created separately.
                    </p>
                    
                    <p>Please coordinate with the employee to collect the assets before their last working day.</p>
                    
                    <hr>
                    <p style="color: #6b7280; font-size: 12px;">
                        This is an automated notification from HRIS Asset Management System
                    </p>
                </body>
                </html>
                """
            
            else:
                # RETURN - Assets return to IT
                subject = f"Asset Return Required - {offboarding.employee.full_name} Offboarding"
                html_body = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2 style="color: #DC2626;">Asset Return - Employee Offboarding</h2>
                    
                    <p><strong>Employee:</strong> {offboarding.employee.full_name} (ID: {offboarding.employee.employee_id})</p>
                    <p><strong>Last Working Day:</strong> {offboarding.last_working_day.strftime('%B %d, %Y')}</p>
                    <p><strong>Total Assets:</strong> {offboarding.total_assets}</p>
                    
                    <hr>
                    
                    <h3>Assets to Return:</h3>
                    {asset_list_html}
                    
                    <p style="background-color: #FEE2E2; padding: 15px; border-left: 4px solid #DC2626;">
                        <strong>üö® Action Required:</strong><br>
                        Please collect all assets from this employee before their last working day.
                    </p>
                    
                    <h3>After Collecting Assets:</h3>
                    <ol>
                        <li>Verify all assets are received</li>
                        <li>Check asset conditions</li>
                        <li>Complete handover in system:
                            <br><br>
                            <a href="{handover_link}" 
                               style="background-color: #2563eb; color: white; padding: 12px 24px; 
                                      text-decoration: none; border-radius: 6px; display: inline-block;">
                                Complete Handover ‚úì
                            </a>
                        </li>
                    </ol>
                    
                    <hr>
                    <p style="color: #6b7280; font-size: 12px;">
                        This is an automated notification from HRIS Asset Management System
                    </p>
                </body>
                </html>
                """
            
            # Send email
            system_email_service.send_email_as_system(
                from_email='myalmet@almettrading.com',
                to_email=it_emails,
                subject=subject,
                body_html=html_body
            )
            
            logger.info(f"‚úÖ IT offboarding email sent - Type: {offboarding.offboarding_type}")
            
        except Exception as e:
            logger.error(f"‚ùå IT offboarding email error: {str(e)}")
    
    @action(detail=True, methods=['post'], url_path='complete-handover')
    @require_asset_permission('complete_handover')
    def complete_handover(self, request, pk=None):
        """
        ‚úÖ IT confirms asset handover completed
        """
        try:
            offboarding = self.get_object()
            
            if offboarding.it_handover_completed:
                return Response(
                    {'error': 'Handover already completed'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            offboarding.it_handover_completed = True
            offboarding.it_handover_completed_at = timezone.now()
            offboarding.it_handover_completed_by = request.user
            offboarding.status = 'COMPLETED'
            offboarding.completed_at = timezone.now()
            offboarding.save()
            
            logger.info(
                f"‚úÖ Handover completed: {offboarding.employee.full_name} by "
                f"{request.user.get_full_name()}"
            )
            
            return Response({
                'success': True,
                'message': 'Asset handover completed successfully',
                'completed_by': request.user.get_full_name(),
                'completed_at': offboarding.it_handover_completed_at
            })
            
        except Exception as e:
            logger.error(f"‚ùå Complete handover error: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def assets(self, request, pk=None):
        """Offboarding √º√ß√ºn asset-l…ôr"""
        offboarding = self.get_object()
        assets = Asset.objects.filter(
            assigned_to=offboarding.employee,
            status__in=['ASSIGNED', 'IN_USE']
        )
        
        return Response({
            'employee': offboarding.employee.full_name,
            'offboarding_type': offboarding.offboarding_type,
            'total_assets': assets.count(),
            'it_handover_completed': offboarding.it_handover_completed,
            'assets': AssetListSerializer(assets, many=True, context={'request': request}).data
        })


# ============================================
# TRANSFER REQUEST VIEWSET - UPDATED
# ============================================
class AssetTransferRequestViewSet(viewsets.ModelViewSet):

    
    queryset = AssetTransferRequest.objects.select_related(
        'asset', 'from_employee', 'to_employee', 'requested_by', 'approved_by'
    ).all()
    serializer_class = AssetTransferRequestSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-requested_at']
    
    @swagger_auto_schema(
        method='post',
        request_body=AssetTransferRequestCreateSerializer
    )
    @action(detail=False, methods=['post'], url_path='create')
    @require_asset_permission('create_transfer')  # üîê Admin/IT only
    def create_transfer(self, request):
        """
        üéØ Transfer sorƒüusu yarat (Admin/IT only)
        """
        try:
            serializer = AssetTransferRequestCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            asset = serializer.validated_data['asset']
            from_employee = serializer.validated_data['from_employee']
            to_employee = serializer.validated_data['to_employee']
            
            # Get or create offboarding
            offboarding, _ = EmployeeOffboarding.objects.get_or_create(
                employee=from_employee,
                status__in=['PENDING', 'IN_PROGRESS'],
                defaults={
                    'last_working_day': timezone.now().date(),
                    'offboarding_type': 'TRANSFER',
                    'total_assets': Asset.objects.filter(assigned_to=from_employee).count(),
                    'created_by': request.user
                }
            )
            
            transfer = AssetTransferRequest.objects.create(
                offboarding=offboarding,
                asset=asset,
                from_employee=from_employee,
                to_employee=to_employee,
                transfer_notes=serializer.validated_data.get('transfer_notes', ''),
                requested_by=request.user
            )
            
            # üìß Send notification emails
            self._send_transfer_notifications(transfer)
            
            logger.info(
                f"‚úÖ Transfer request created: {asset.asset_number} ‚Üí "
                f"{to_employee.full_name} by {request.user.get_full_name()}"
            )
            
            return Response({
                'success': True,
                'transfer_id': transfer.id,
                'message': 'Transfer request created successfully'
            })
            
        except Exception as e:
            logger.error(f"‚ùå Transfer request error: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _send_transfer_notifications(self, transfer):
        """
        üìß Send transfer notifications:
        1. To employee (must approve)
        2. From employee (informational)
        """
        try:
            to_employee_email = transfer.to_employee.user.email if transfer.to_employee.user else None
            from_employee_email = transfer.from_employee.user.email if transfer.from_employee.user else None
            
            # üìß Email to NEW employee (must approve)
            if to_employee_email:
                approve_link = f"https://yourdomain.com/assets/transfers/{transfer.id}/approve"
                
                html_body_to = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2 style="color: #2563eb;">Asset Transfer - Your Approval Required</h2>
                    
                    <p>Dear {transfer.to_employee.full_name},</p>
                    
                    <p>An asset has been assigned to you as part of an employee offboarding process:</p>
                    
                    <div style="background-color: #F3F4F6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <p><strong>Asset Details:</strong></p>
                        <ul>
                            <li><strong>Asset:</strong> {transfer.asset.asset_name}</li>
                            <li><strong>Serial Number:</strong> {transfer.asset.serial_number}</li>
                            <li><strong>Asset Number:</strong> {transfer.asset.asset_number}</li>
                            <li><strong>Previous User:</strong> {transfer.from_employee.full_name}</li>
                            <li><strong>Requested by:</strong> {transfer.requested_by.get_full_name()}</li>
                        </ul>
                    </div>
                    
                    <p style="background-color: #DBEAFE; padding: 15px; border-left: 4px solid #2563eb;">
                        <strong>‚ö†Ô∏è Action Required:</strong><br>
                        Please confirm that you accept this asset transfer.
                    </p>
                    
                    <p style="text-align: center; margin: 30px 0;">
                        <a href="{approve_link}" 
                           style="background-color: #10B981; color: white; padding: 14px 28px; 
                                  text-decoration: none; border-radius: 6px; display: inline-block; 
                                  font-weight: bold;">
                            Approve Transfer ‚úì
                        </a>
                    </p>
                    
                    {f"<p><strong>Notes:</strong> {transfer.transfer_notes}</p>" if transfer.transfer_notes else ""}
                    
                    <p>If you have any questions, please contact IT department.</p>
                    
                    <hr>
                    <p style="color: #6b7280; font-size: 12px;">
                        This is an automated notification from HRIS Asset Management System
                    </p>
                </body>
                </html>
                """
                
                system_email_service.send_email_as_system(
                    from_email='myalmet@almettrading.com',
                    to_email=to_employee_email,
                    subject=f'Asset Transfer Approval Required - {transfer.asset.asset_name}',
                    body_html=html_body_to
                )
                
                logger.info(f"‚úÖ Transfer approval email sent to {transfer.to_employee.full_name}")
            
            # üìß Email to OLD employee (informational)
            if from_employee_email:
                html_body_from = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2 style="color: #6B7280;">Asset Transfer Notification</h2>
                    
                    <p>Dear {transfer.from_employee.full_name},</p>
                    
                    <p>As part of your offboarding process, one of your assigned assets is being transferred:</p>
                    
                    <div style="background-color: #F3F4F6; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <p><strong>Asset Details:</strong></p>
                        <ul>
                            <li><strong>Asset:</strong> {transfer.asset.asset_name}</li>
                            <li><strong>Serial Number:</strong> {transfer.asset.serial_number}</li>
                            <li><strong>Asset Number:</strong> {transfer.asset.asset_number}</li>
                            <li><strong>Transferring to:</strong> {transfer.to_employee.full_name}</li>
                        </ul>
                    </div>
                    
                    <p>IT department will contact you to coordinate the asset handover.</p>
                    
                    {f"<p><strong>Notes:</strong> {transfer.transfer_notes}</p>" if transfer.transfer_notes else ""}
                    
                    <hr>
                    <p style="color: #6b7280; font-size: 12px;">
                        This is an automated notification from HRIS Asset Management System
                    </p>
                </body>
                </html>
                """
                
                system_email_service.send_email_as_system(
                    from_email='myalmet@almettrading.com',
                    to_email=from_employee_email,
                    subject=f'Asset Transfer - {transfer.asset.asset_name}',
                    body_html=html_body_from
                )
                
                logger.info(f"‚úÖ Transfer notification email sent to {transfer.from_employee.full_name}")
            
        except Exception as e:
            logger.error(f"‚ùå Transfer notification error: {str(e)}")
    
    @swagger_auto_schema(
        method='post',
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'approved': openapi.Schema(type=openapi.TYPE_BOOLEAN),
                'comments': openapi.Schema(type=openapi.TYPE_STRING)
            }
        )
    )
    @action(detail=True, methods=['post'], url_path='employee-approve')
    def employee_approve_transfer(self, request, pk=None):
        """
        ‚úÖ Employee approves incoming transfer
        """
        try:
            transfer = self.get_object()
            approved = request.data.get('approved', False)
            comments = request.data.get('comments', '')
            
          
            
            if transfer.status != 'PENDING':
                return Response(
                    {'error': f'Transfer cannot be approved. Status: {transfer.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if approved:
                # ‚úÖ Approve and complete transfer
                with transaction.atomic():
                    asset = transfer.asset
                    old_employee = transfer.from_employee
                    new_employee = transfer.to_employee
                    
                    # Check in from old employee
                    active_assignment = asset.assignments.filter(
                        check_in_date__isnull=True
                    ).first()
                    if active_assignment:
                        active_assignment.check_in_date = timezone.now().date()
                        active_assignment.checked_in_by = request.user
                        active_assignment.check_in_notes = f"Transfer to {new_employee.full_name}"
                        active_assignment.save()
                    
                    # Create new assignment
                    AssetAssignment.objects.create(
                        asset=asset,
                        employee=new_employee,
                        check_out_date=timezone.now().date(),
                        assigned_by=transfer.requested_by,
                        check_out_notes=f'Transfer from {old_employee.full_name}'
                    )
                    
                    # Update asset
                    asset.assigned_to = new_employee
                    asset.status = 'IN_USE'  # Directly IN_USE (already approved)
                    asset.updated_by = request.user
                    asset.save()
                    
                    # Update transfer
                    transfer.status = 'COMPLETED'
                    transfer.employee_approved = True
                    transfer.employee_approved_at = timezone.now()
                    transfer.completed_at = timezone.now()
                    transfer.save()
                    
                    # Update offboarding
                    transfer.offboarding.assets_transferred += 1
                    transfer.offboarding.save()
                    
                    # Log activity
                    AssetActivity.objects.create(
                        asset=asset,
                        activity_type='TRANSFERRED',
                        description=f'Transfer completed: {old_employee.full_name} ‚Üí {new_employee.full_name}',
                        performed_by=request.user,
                        metadata={
                            'from_employee': old_employee.full_name,
                            'to_employee': new_employee.full_name,
                            'transfer_id': transfer.id,
                            'comments': comments
                        }
                    )
                
                logger.info(f"‚úÖ Transfer approved: {asset.asset_number} by {new_employee.full_name}")
                
                return Response({
                    'success': True,
                    'message': 'Transfer approved and completed successfully'
                })
            
            else:
                # ‚ùå Reject transfer
                transfer.status = 'REJECTED'
                transfer.rejection_reason = comments or 'Rejected by employee'
                transfer.save()
                
                logger.info(f"‚ùå Transfer rejected: {transfer.asset.asset_number} by {transfer.to_employee.full_name}")
                
                return Response({
                    'success': True,
                    'message': 'Transfer rejected'
                })
            
        except Exception as e:
            logger.error(f"‚ùå Employee approve transfer error: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='my-pending')
    def my_pending_transfers(self, request):
        """
        Get transfers pending my approval
        """
        access = get_asset_access_level(request.user)
        
        if not access['employee']:
            return Response({'transfers': []})
        
        transfers = AssetTransferRequest.objects.filter(
            to_employee=access['employee'],
            status='PENDING'
        ).select_related('asset', 'from_employee', 'requested_by')
        
        return Response({
            'count': transfers.count(),
            'transfers': AssetTransferRequestSerializer(
                transfers, 
                many=True, 
                context={'request': request}
            ).data
        })